import frappe
import requests, json
from frappe.utils import flt
from frappe.utils import now_datetime , add_days, get_datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO

# ------------------------------------------------------------------
# Stage 1: Send to Item & BOM Creation
# ------------------------------------------------------------------
@frappe.whitelist()
def send_to_item_bom_creation(docname):
    """
    Transition Product Return Form from Draft to 'Item & BOM Creation' stage.

    Validations:
        - Document must be in Draft status
        - is_jewlex_credit_note must be checked
        - At least one child row must have a tag_no

    Updates:
        - status = 'Item & BOM Creation'
        - Saves document
    """
    doc = frappe.get_doc("Product Return Order Form", docname)

    if doc.status != "Draft":
        frappe.throw(f"Status must be Draft to send to Item & BOM Creation. Current status: {frappe.bold(doc.status)}")

    if not doc.is_jewlex_credit_note:
        frappe.throw("This action is only available for Jewlex Credit Notes.")

    has_tag = any(row.tag_no for row in doc.items)
    if not has_tag:
        frappe.throw("At least one item row must have a Tag No before proceeding.")

    doc.status = "Item & BOM Creation"

#     error_dict = {}
#     success_dict = {}
#     repair_order = frappe.new_doc("Repair Order Form")

#     repair_order.source_prf = doc.name
#     repair_order.company = doc.company
#     repair_order.branch = doc.branch
#     repair_order.customer_code = doc.customer
#     repair_order.department ='Information Technology - GEPL'
#     repair_order.salesman_name='Chintankumar Dilipbhai Ramani'
#     repair_order.order_date=now_datetime()
#     repair_order.delivery_date = get_datetime(
#     add_days(repair_order.order_date.date(), 1).strftime("%Y-%m-%d") + " 11:00:00"
# )
#     repair_order.due_days=1
#     repair_order.status = "Draft"
#     repair_order.items = []

    
#     for row in doc.items:   # Replace 'items' with your source child table fieldname
#         repair_order.append("order_details", {
#             "tag_no1": row.tag_no,
#             "item_code": row.item_code,
#             "category": row.item_category,
#             "subcategory": row.item_subcategory,
#             "metal_target": row.gold_rate,
#             "diamond_target": row.diamond_weight,
#             "metal_touch":row.metal_touch,
#             "metal_purity":row.metal_purity,
#             "metal_colour":row.metal_colour,
#             "setting_type":row.setting_type
            
#             # Add any other fields you want to copy
#         })
#     repair_order.flags.ignore_mandatory = True
#     repair_order.insert(ignore_permissions=True)

    doc.save(ignore_permissions=True)
#     frappe.msgprint(
#     f'Repair Order Form is Created <a href="/app/repair-order-form/{repair_order.name}" target="_blank">{repair_order.name}</a>'
# )
    return doc.status


# ------------------------------------------------------------------
# Stage 2: Update Items from Tag (Auto-link items)
# ------------------------------------------------------------------
@frappe.whitelist()
def update_items_from_tag(docname):
    """
    For each child row in Product Return Form Item that has a tag_no,
    find the matching Item where item.old_tag_no == child.tag_no
    and set child.item_code = item.name.

    After successful update, transitions status to 'Send to Pricing'
    and triggers automatic pricing calculation.

    Validations:
        - Document must be in 'Item & BOM Creation' status
        - Each tag_no must match exactly one enabled Item
    """
    doc = frappe.get_doc("Product Return  Order Form", docname)

    if doc.status != "Item & BOM Creation":
        frappe.throw(
            f"Status must be 'Item & BOM Creation' to update items from tag. "
            f"Current status: {frappe.bold(doc.status)}"
        )

    updated_count = 0
    not_found_tags = []

    for row in doc.items:
        if not row.tag_no:
            continue

        # Find Item where old_tag_no matches child.tag_no
        matching_items = frappe.get_all(
            "Item",
            filters={"old_tag_no": row.tag_no, "disabled": 0},
            fields=["name", "master_bom"],
            limit=1,
        )

        if not matching_items:
            not_found_tags.append(f"Row {row.idx}: Tag No {frappe.bold(row.tag_no)}")
            continue

        item_name = matching_items[0].name
        master_bom = matching_items[0].master_bom
        row.item_code = item_name

        if not master_bom:
            # Also fetch the BOM for this item (latest active default BOM)
            default_bom = frappe.db.get_value(
                "BOM",
                {"item": item_name, "is_active": 1, "is_default": 1},
                "name",
            )
            if default_bom:
                master_bom = default_bom
        
        if not master_bom:
            frappe.throw(
                f"No BOM found for Item {frappe.bold(item_name)}"
            )
        row.bom = master_bom

        updated_count += 1

    if not_found_tags:
        frappe.throw(
            "No matching Item found for the following Tag Numbers:<br>"
            + "<br>".join(not_found_tags)
        )

    if updated_count == 0:
        frappe.throw("No rows with Tag No found to update.")

    # Transition to pricing stage
    doc.status = "Send to Pricing"

    # Trigger pricing calculation
    trigger_pricing_calculation(doc)

    doc.save(ignore_permissions=True)
    frappe.msgprint(
        f"Successfully linked {updated_count} item(s) from Tag No. "
        f"Pricing calculation completed.",
        indicator="green",
        alert=True,
    )

    return doc.status


# ------------------------------------------------------------------
# Stage 3: Trigger Pricing Calculation (BOM-based)
# ------------------------------------------------------------------
def trigger_pricing_calculation(doc):
    """
    BOM-based pricing calculation for Jewelex credit notes.
    Routes to the correct calculation logic based on credit_note_type
    and credit_note_subtype, mirroring the validate() dispatch in
    ProductReturnForm.

    Calculation mapping:
        (Actual, Sale Without Payment-Actual)     → _calc_bom_standard
        (Actual, Sale With Payment-Actual)         → _calc_bom_standard
        (Repair, QC Fail-Repair)                   → _calc_bom_standard
        (Repair, Physical-Repair)                  → _calc_bom_physical_repair
        (Consignment, Finish Goods-Consignment)    → _calc_bom_standard
        (Consignment, Raw Material-Consignment)    → _calc_bom_raw_material_consignment
    """
    if doc.status in ["Draft", "Item & BOM Creation"]:
        return

    credit_note_key = (doc.credit_note_type, doc.return_subtype)

    calc_mapping = {
        ("Return", "Sale Without Payment-Return"): _calc_bom_pcpm,
        ("Return", "Sale With Payment-Return"): _calc_bom_bbpm,
        ("Repair", "QC Fail-Repair"): _calc_bom_pcpm,
        ("Repair", "Physical-Repair"): _calc_bom_physical_repair,
        ("Consignment", "Finish Goods-Consignment"): _calc_bom_pcpm,
        ("Consignment", "Raw Material-Consignment"): _calc_bom_raw_material_consignment,
    }

    calc_fn = calc_mapping.get(credit_note_key)
    if not calc_fn:
        frappe.throw(
            f"Pricing calculation not supported for credit note type: "
            f"{frappe.bold(doc.credit_note_type)} / {frappe.bold(doc.credit_note_subtype)}"
        )

    calc_fn(doc)

    # Recalculate taxes and totals using existing class methods
    doc.calculate_taxes_and_totals()
    doc.set_total_in_words()


# ==================================================================
# SHARED HELPERS
# ==================================================================
def _get_common_setup(doc):
    """Setup variables common to all BOM-based calculations."""
    gold_gst_rate = flt(
        frappe.db.get_single_value("Jewellery Settings", "gold_gst_rate") or 0
    )
    customer_group = frappe.db.get_value("Customer", doc.customer, "customer_group")
    gemstone_price_list_type = frappe.db.get_value(
        "Customer", doc.customer, "custom_gemstone_price_list_type"
    )
    return gold_gst_rate, customer_group, gemstone_price_list_type


def _calc_jewelex_row_amounts(doc, row):
    """
    Populate a Jewelex-tag row's amount fields from a live Jwelex API fetch.
    Triggers purely off `row.jewelex_tag` being set (works regardless of the
    `is_jewelex_tag` checkbox) and needs no BOM.

    Returns True if the row was a Jewelex row and was handled, so callers can
    skip their BOM-based calculation for it; False otherwise.
    """
    if not row.get("jewelex_tag"):
        return False

    data = get_data_from_jwelex(row.jewelex_tag)
    if not data:
        frappe.throw(f"No JWELEX data found for Tag No {frappe.bold(row.jewelex_tag)}")

    totals = data.get("totals", {})
    charges_info = data.get("charges_info", {})
    summary_totals = data.get("summary_totals", {})

    making_amount = flt(charges_info.get("metal_making_amount")) + flt(charges_info.get("chain_making_amount"))
    if doc.making_charges_type == "Without":
        making_amount = 0
    elif doc.making_charges_type == "Half":
        making_amount = making_amount * 0.5
    # "With" (or unset) keeps the full making_amount

    row.metal_amount = flt(totals.get("metal_totals", {}).get("total_amount"))
    row.diamond_amount = flt(totals.get("diamond_totals", {}).get("total_amount"))
    row.gemstone_amount = flt(totals.get("stone_totals", {}).get("total_amount"))
    row.finding_amount = flt(totals.get("finding_totals", {}).get("total_amount"))
    row.other_material_amount = flt(totals.get("other_totals", {}).get("total_amount"))
    row.making_amount = making_amount
    row.certification_amount = flt(charges_info.get("certificate_charges"))
    row.hallmarking_amount = flt(charges_info.get("hm_charges"))
    row.gross_weight = summary_totals.get("gross_wt")

    # Note: hallmarking_amount/certification_amount are set above but not
    # folded into rate/amount here — validate() does that once, centrally,
    # gated by product_hallmarking/product_certification, so they're never
    # double-counted.
    row.rate = (
        row.metal_amount
        + row.finding_amount
        + row.diamond_amount
        + row.gemstone_amount
        + row.making_amount
    )
    row.amount = row.rate * flt(row.qty or 1)

    return True


def _calc_metal_amount(doc, bom_doc, mc_name, gold_gst_rate):
    """
    Calculate metal amount from BOM metal_detail.
    Returns (metal_total, wastage_total).
    """
    row_metal_amt_total = 0
    row_wastage_amt_total = 0

    if not hasattr(bom_doc, "metal_detail"):
        return row_metal_amt_total, row_wastage_amt_total

    for md_row in bom_doc.metal_detail:
        sub_info = frappe.db.get_value(
            "Making Charge Price Item Subcategory",
            {
                "parent": mc_name,
                "subcategory": bom_doc.item_subcategory,
            },
            [
                "rate_per_gm",
                "rate_per_pc",
                "wastage",
                "rate_per_gm_threshold",
            ],
            as_dict=True,
        )

        if not sub_info:
            frappe.throw(
                f"Making Charge Subcategory {bom_doc.item_subcategory} not found"
            )

        threshold = flt(sub_info.rate_per_gm_threshold) or 2
        weight_for_calc = flt(bom_doc.metal_and_finding_weight)

        if weight_for_calc < threshold:
            wastage_rate = 0
        else:
            wastage_rate = flt(sub_info.wastage) / 100

        customer_metal_purity = frappe.db.get_value(
            "Metal Criteria",
            {
                "parent": doc.customer,
                "metal_type": md_row.metal_type,
                "metal_touch": md_row.metal_touch,
            },
            "metal_purity",
        )

        if not customer_metal_purity:
            frappe.throw("Metal Purity not found for Customer")

        calculated_gold_rate = (
            flt(customer_metal_purity) * doc.gold_rate_with_gst
        ) / (100 + gold_gst_rate)

        line_gold_amt = round(calculated_gold_rate * md_row.quantity, 2)
        line_wastage_amt = line_gold_amt * wastage_rate

        row_metal_amt_total += line_gold_amt
        row_wastage_amt_total += line_wastage_amt

    return row_metal_amt_total, row_wastage_amt_total


def _calc_finding_amount(doc, bom_doc, mc_name, gold_gst_rate):
    """
    Calculate finding amount from BOM finding_detail.
    Returns (finding_total, wastage_total).
    """
    row_finding_amt_total = 0
    row_wastage_amt_total = 0

    if not hasattr(bom_doc, "finding_detail"):
        return row_finding_amt_total, row_wastage_amt_total

    for fd_row in bom_doc.finding_detail:
        find = frappe.db.get_all(
            "Making Charge Price Finding Subcategory",
            filters={
                "parent": mc_name,
                "subcategory": fd_row.finding_type,
            },
            fields=["rate_per_gm", "rate_per_pc"],
            limit=1,
        )

        # Fallback to Item Subcategory
        if not find:
            find = frappe.db.get_all(
                "Making Charge Price Item Subcategory",
                filters={
                    "parent": mc_name,
                    "subcategory": bom_doc.item_subcategory,
                },
                fields=["rate_per_gm", "rate_per_pc"],
                limit=1,
            )

        if not find:
            frappe.throw(
                f"Finding rate not found for {fd_row.finding_type}"
            )

        find_info = find[0]

        customer_metal_purity = frappe.db.sql(
            f"""select metal_purity from `tabMetal Criteria`
            where parent = '{doc.customer}'
            and metal_type = '{fd_row.metal_type}'
            and metal_touch = '{fd_row.metal_touch}'""",
            as_dict=True,
        )[0]["metal_purity"]

        calculated_gold_rate = (
            float(customer_metal_purity) * doc.gold_rate_with_gst
        ) / (100 + int(gold_gst_rate))

        line_finding_rate = round(calculated_gold_rate, 2)
        line_finding_amt = round(line_finding_rate * fd_row.quantity, 2)

        finding_weight = getattr(bom_doc, "metal_and_finding_weight", None)
        if finding_weight is not None and finding_weight < 2:
            wastage_rate = 0
        else:
            wastage_rate = find_info.get("wastage", 0) / 100.0

        row_finding_amt_total += line_finding_amt
        row_wastage_amt_total += wastage_rate * line_finding_amt

    return row_finding_amt_total, row_wastage_amt_total


def _calc_making_from_bom(doc, bom_doc, mc_name):
    """
    Calculate making charges from BOM metal_detail + finding_detail.
    Used for Physical Repair type (accumulates from BOM, not invoice).
    Returns making_total.
    """
    row_making_amt_total = 0
    # frappe.msgprint(f"MC Name: {mc_name}")
    if hasattr(bom_doc, "metal_detail"):
        for md_row in bom_doc.metal_detail:
            sub_info = frappe.db.get_value(
                "Making Charge Price Item Subcategory",
                {
                    "parent": mc_name,
                    "subcategory": bom_doc.item_subcategory,
                },
                [
                    "rate_per_gm",
                    "rate_per_pc",
                    "rate_per_gm_threshold",
                ],
                as_dict=True,
            )

            if not sub_info:
                continue

            threshold = flt(sub_info.rate_per_gm_threshold) or 2
            weight_for_calc = flt(bom_doc.metal_and_finding_weight)

            if weight_for_calc < threshold:
                making_rate = flt(sub_info.rate_per_pc)
                line_making_amt = making_rate
            else:
                making_rate = flt(sub_info.rate_per_gm)
                line_making_amt = making_rate * md_row.quantity

            row_making_amt_total += line_making_amt

    if hasattr(bom_doc, "finding_detail"):
        for fd_row in bom_doc.finding_detail:
            find = frappe.db.get_all(
                "Making Charge Price Finding Subcategory",
                filters={
                    "parent": mc_name,
                    "subcategory": fd_row.finding_type,
                },
                fields=["rate_per_gm", "rate_per_pc"],
                limit=1,
            )

            if not find:
                find = frappe.db.get_all(
                    "Making Charge Price Item Subcategory",
                    filters={
                        "parent": mc_name,
                        "subcategory": bom_doc.item_subcategory,
                    },
                    fields=["rate_per_gm", "rate_per_pc"],
                    limit=1,
                )

            if not find:
                continue

            find_info = find[0]
            finding_weight = getattr(bom_doc, "metal_and_finding_weight", None)

            if finding_weight is not None and finding_weight < 2:
                making_rate = find_info.get("rate_per_pc", 0)
                line_making_amt = making_rate
            else:
                making_rate = find_info.get("rate_per_gm", 0)
                line_making_amt = making_rate * fd_row.quantity

            row_making_amt_total += line_making_amt

    return row_making_amt_total


def _calc_diamond_amount(doc, bom_doc, customer_group, use_handling_charges=False):
    """
    Calculate diamond amount from BOM diamond_detail.

    Args:
        use_handling_charges: If True (Physical-Repair), apply outright/outwork
            handling charges to diamond rate. If False (BBPM/PCPM), use base rate only.
    Returns diamond_total.
    """
    row_diamond_amt_total = 0

    if not hasattr(bom_doc, "diamond_detail"):
        return row_diamond_amt_total

    for diamond_row in bom_doc.diamond_detail:
        customer_price_list = frappe.db.sql(
            """
            SELECT diamond_price_list
            FROM `tabDiamond Price List Table`
            WHERE parent = %s AND diamond_shape = %s
            """,
            (doc.customer, diamond_row.stone_shape),
            as_dict=True,
        )

        if not customer_price_list:
            continue

        diamond_price_list = customer_price_list[0].diamond_price_list

        common_filters = {
            "price_list": "Standard Selling",
            "price_list_type": diamond_price_list,
            "customer": doc.customer,
            "diamond_type": diamond_row.diamond_type,
            "stone_shape": diamond_row.stone_shape,
            "diamond_quality": diamond_row.quality,
        }

        weight_per_piece = (
            diamond_row.quantity / diamond_row.pcs
            if diamond_row.pcs else 0
        )
        weight_per_piece = round(weight_per_piece, 3)

        diamond_price_row = None

        if diamond_price_list == "Sieve Size Range":
            diamond_price_row = frappe.db.get_value(
                "Diamond Price List",
                {**common_filters, "sieve_size_range": diamond_row.sieve_size_range},
                [
                    "rate",
                    "outright_handling_charges_rate",
                    "outwork_handling_charges_rate",
                    "outright_handling_charges_in_percentage",
                    "outwork_handling_charges_in_percentage",
                    "supplier_fg_purchase_rate",
                ],
                as_dict=True,
            )

        elif diamond_price_list == "Weight (in cts)":
            filter_conditions = " AND ".join(
                [f"{key} = %s" for key in common_filters]
            )
            rate_result = frappe.db.sql(
                f"""
                SELECT
                    name,
                    rate,
                    outright_handling_charges_rate,
                    outright_handling_charges_in_percentage,
                    outwork_handling_charges_rate,
                    outwork_handling_charges_in_percentage,
                    supplier_fg_purchase_rate
                FROM `tabDiamond Price List`
                WHERE {filter_conditions}
                AND %s BETWEEN from_weight AND to_weight
                LIMIT 1
                """,
                list(common_filters.values()) + [weight_per_piece],
                as_dict=True,
            )
            diamond_price_row = rate_result[0] if rate_result else None

        elif diamond_price_list == "Size (in mm)":
            diamond_price_row = frappe.db.get_value(
                "Diamond Price List",
                {**common_filters, "diamond_size_in_mm": diamond_row.diamond_sieve_size},
                [
                    "rate",
                    "outright_handling_charges_rate",
                    "outwork_handling_charges_rate",
                    "outright_handling_charges_in_percentage",
                    "outwork_handling_charges_in_percentage",
                    "supplier_fg_purchase_rate",
                ],
                as_dict=True,
            )

        if not diamond_price_row:
            continue

        # -------------------------------------------------
        # RATE CALCULATION
        # -------------------------------------------------
        base_rate = diamond_price_row.get("rate", 0)

        if use_handling_charges:
            # Physical-Repair: apply outright/outwork handling
            outright_rate = diamond_price_row.get("outright_handling_charges_rate", 0)
            outright_pct = diamond_price_row.get("outright_handling_charges_in_percentage", 0)
            is_customer_item = getattr(diamond_row, "is_customer_item", False)

            if is_customer_item:
                total_rate = outright_rate or (base_rate * (outright_pct / 100))
            else:
                if outright_rate:
                    total_rate = base_rate + outright_rate
                else:
                    total_rate = base_rate + (base_rate * (outright_pct / 100))
        else:
            # BBPM/PCPM: use base rate only
            total_rate = base_rate

        # -------------------------------------------------
        # COMPANY & CUSTOMER GROUP LOGIC
        # -------------------------------------------------
        if (
            doc.company == "KG GK Jewellers Private Limited"
            and customer_group == "Internal"
        ):
            diamond_rate = diamond_row.se_rate
            quantity = round(diamond_row.quantity, 3)
            diamond_amount = round(quantity * diamond_rate, 2)

        elif (
            doc.company == "Gurukrupa Export Private Limited"
            and customer_group == "Internal"
        ):
            diamond_rate = diamond_price_row.get("supplier_fg_purchase_rate", 0)
            quantity = round(diamond_row.quantity, 3)
            diamond_amount = round(quantity * diamond_rate, 2)

        else:
            diamond_rate = round(total_rate, 2)
            quantity = round(diamond_row.quantity, 3)
            diamond_amount = round(quantity * diamond_rate, 2)

        row_diamond_amt_total += diamond_amount

    return row_diamond_amt_total


def _calc_gemstone_amount(doc, bom_doc, customer_group, gemstone_price_list_type):
    """Calculate gemstone amount from BOM gemstone_detail. Returns gemstone_total."""
    row_gemstone_amt_total = 0

    if not (hasattr(bom_doc, "gemstone_detail") and bom_doc.gemstone_detail):
        return row_gemstone_amt_total

    def calculate_percentage_amount(rate, base_value):
        return round((flt(rate) / 100) * flt(base_value), 2)

    for gs_row in bom_doc.gemstone_detail:

        # Internal customer – Company specific
        if doc.company == "Gurukrupa Export Private Limited" and customer_group == "Internal":
            rate = gs_row.fg_purchase_rate
            row_gemstone_amt_total += calculate_percentage_amount(rate, gs_row.quantity)
            continue

        if doc.company == "KG GK Jewellers Private Limited" and customer_group == "Internal":
            rate = gs_row.se_rate
            row_gemstone_amt_total += calculate_percentage_amount(rate, gs_row.quantity)
            continue

        # Fixed Price List – Non-Retail
        if gemstone_price_list_type == "Fixed" and customer_group != "Retail":
            price_list = frappe.get_all(
                "Gemstone Price List",
                filters={
                    "customer": doc.customer,
                    "price_list_type": gemstone_price_list_type,
                    "gemstone_grade": gs_row.gemstone_grade,
                    "cut_or_cab": gs_row.cut_or_cab,
                    "gemstone_type": gs_row.gemstone_type,
                    "stone_shape": gs_row.stone_shape,
                },
                fields=["rate", "handling_rate"],
                limit=1,
            )

            if not price_list:
                frappe.throw("No Gemstone Price List found")

            rate = price_list[0].rate
            row_gemstone_amt_total += calculate_percentage_amount(rate, gs_row.quantity)
            continue

        # Retail Customer – Fixed Price
        if customer_group == "Retail":
            price_list = frappe.get_all(
                "Gemstone Price List",
                filters={
                    "is_retail_customer": 1,
                    "price_list_type": gemstone_price_list_type,
                    "gemstone_grade": gs_row.gemstone_grade,
                    "cut_or_cab": gs_row.cut_or_cab,
                    "gemstone_type": gs_row.gemstone_type,
                    "stone_shape": gs_row.stone_shape,
                },
                fields=["rate", "outwork_handling_charges_rate"],
                limit=1,
            )

            if not price_list:
                frappe.throw("No Retail Gemstone Price List found")

            rate = (
                price_list[0].outwork_handling_charges_rate
                if gs_row.is_customer_item
                else price_list[0].rate
            )
            row_gemstone_amt_total += calculate_percentage_amount(rate, gs_row.quantity)
            continue

        # Diamond Range – Non-Retail
        if gemstone_price_list_type == "Diamond Range" and customer_group != "Retail":
            price_list = frappe.get_all(
                "Gemstone Price List",
                filters={
                    "customer": doc.customer,
                    "price_list_type": gemstone_price_list_type,
                    "cut_or_cab": gs_row.cut_or_cab,
                    "gemstone_grade": gs_row.gemstone_grade,
                    "from_gemstone_pr_rate": ["<=", gs_row.gemstone_pr],
                    "to_gemstone_pr_rate": [">=", gs_row.gemstone_pr],
                },
                fields=["name"],
                limit=1,
            )

            if not price_list:
                frappe.throw("Gemstone Diamond Range price list not found")

            price_list_doc = frappe.get_doc("Gemstone Price List", price_list[0].name)

            rate = 0.0
            for mul in price_list_doc.gemstone_multiplier:
                if (
                    mul.gemstone_type == gs_row.gemstone_type
                    and flt(mul.from_weight) <= flt(gs_row.gemstone_pr) <= flt(mul.to_weight)
                ):
                    if gs_row.is_customer_item:
                        rate = {
                            "Precious": mul.outwork_precious_percentage,
                            "Semi-Precious": mul.outwork_semi_precious_percentage,
                            "Synthetic": mul.outwork_synthetic_percentage,
                        }.get(gs_row.gemstone_quality, 0)
                    else:
                        rate = {
                            "Precious": mul.precious_percentage,
                            "Semi-Precious": mul.semi_precious_percentage,
                            "Synthetic": mul.synthetic_percentage,
                        }.get(gs_row.gemstone_quality, 0)
                    break

            row_gemstone_amt_total += calculate_percentage_amount(rate, gs_row.gemstone_pr)
            continue

        # Diamond Range – Retail
        if gemstone_price_list_type == "Diamond Range" and customer_group == "Retail":
            price_list = frappe.get_all(
                "Gemstone Price List",
                filters={
                    "is_retail_customer": 1,
                    "price_list_type": gemstone_price_list_type,
                    "cut_or_cab": gs_row.cut_or_cab,
                    "gemstone_grade": gs_row.gemstone_grade,
                    "from_gemstone_pr_rate": ["<=", gs_row.gemstone_pr],
                    "to_gemstone_pr_rate": [">=", gs_row.gemstone_pr],
                },
                fields=["name"],
                limit=1,
            )

            if not price_list:
                frappe.throw("Retail Gemstone Diamond Range price list not found")

            price_list_doc = frappe.get_doc("Gemstone Price List", price_list[0].name)

            rate = 0.0
            for mul in price_list_doc.get("gemstone_multiplier", []):
                if (
                    mul.gemstone_type == gs_row.gemstone_type
                    and flt(mul.from_weight) <= flt(gs_row.gemstone_pr) <= flt(mul.to_weight)
                ):
                    rate = {
                        "Precious": mul.precious_percentage,
                        "Semi-Precious": mul.semi_precious_percentage,
                        "Synthetic": mul.synthetic_percentage,
                    }.get(gs_row.gemstone_quality, 0)
                    break

            row_gemstone_amt_total += calculate_percentage_amount(rate, gs_row.gemstone_pr)

    return row_gemstone_amt_total


def _get_mc_name(doc, bom_doc):
    """Lookup Making Charge Price for the given doc and BOM."""
    mc = frappe.get_all(
        "Making Charge Price",
        filters={
            "customer": doc.customer,
            "metal_type": bom_doc.metal_type,
            "setting_type": bom_doc.setting_type,
            "from_gold_rate": ["<=", doc.gold_rate_with_gst],
            "to_gold_rate": [">=", doc.gold_rate_with_gst],
            "metal_touch": bom_doc.metal_touch,
        },
        fields=["name"],
        limit=1,
    )

    if not mc:
        frappe.throw(
            f"Create a valid Making Charge Price for "
            f"{bom_doc.metal_type} / {bom_doc.metal_touch}"
        )

    return mc[0]["name"]


# ==================================================================
# CALCULATION TYPE: STANDARD (BBPM)
# ==================================================================
def _calc_bom_bbpm(doc):
    """
    - Gold rate should be taken as the current day rate. (on the date of the creditnote).
    - Diamond and gemstone rates should be applied as per the updated live price list.
    - Making charges must be applied exactly as per the original invoice rate or as selected in the form.
    """
    gold_gst_rate, customer_group, gemstone_price_list_type = _get_common_setup(doc)

    total_taxable = 0
    total_gst = 0
    doc.sales_taxes_and_charges = []

    for row in doc.items:
        if _calc_jewelex_row_amounts(doc, row):
            total_taxable += row.amount
            continue

        if not row.item_code or not row.bom:
            continue

        bom_doc = frappe.get_doc("BOM", row.bom)
        mc_name = _get_mc_name(doc, bom_doc)

        row_metal_amt, _ = _calc_metal_amount(doc, bom_doc, mc_name, gold_gst_rate)
        row_finding_amt, _ = _calc_finding_amount(doc, bom_doc, mc_name, gold_gst_rate)
        row_diamond_amt = _calc_diamond_amount(doc, bom_doc, customer_group, use_handling_charges=False)
        row_gemstone_amt = _calc_gemstone_amount(doc, bom_doc, customer_group, gemstone_price_list_type)

        row.metal_amount = row_metal_amt
        row.finding_amount = row_finding_amt
        row.diamond_amount = row_diamond_amt
        row.gemstone_amount = row_gemstone_amt
        making_amount = 0

        row.rate = (
            row.metal_amount
            + row.finding_amount
            + row.diamond_amount
            + row.gemstone_amount
            + making_amount
        )
        row.amount = row.rate * row.qty

        total_taxable += row.amount

    doc.total_taxes_and_charges = total_gst
    doc.grand_total = total_taxable + total_gst

# ==================================================================
# CALCULATION TYPE: PCPM, QC Fail-Repair, Finish Goods-Consignment
# ==================================================================
def _calc_bom_pcpm(doc):
    """
    same product rate mentioned in the invoice must be used
    No new or revised rates should be applied.
    """
    for row_item in doc.items:
        _calc_jewelex_row_amounts(doc, row_item)

# ==================================================================
# CALCULATION TYPE: PHYSICAL REPAIR
# ==================================================================
def _calc_bom_physical_repair(doc):
    """
    Physical Repair BOM-based calculation.
    Same as standard BUT:
      - Making charges accumulated from BOM (not invoice)
      - Diamond has outright/outwork handling charges applied

    Used for:
      - Physical-Repair
    """
    gold_gst_rate, customer_group, gemstone_price_list_type = _get_common_setup(doc)

    total_taxable = 0
    total_gst = 0
    doc.sales_taxes_and_charges = []

    for row in doc.items:
        if not row.item_code or not row.bom:
            continue

        bom_doc = frappe.get_doc("BOM", row.bom)
        mc_name = _get_mc_name(doc, bom_doc)

        row_metal_amt, _ = _calc_metal_amount(doc, bom_doc, mc_name, gold_gst_rate)
        row_finding_amt, _ = _calc_finding_amount(doc, bom_doc, mc_name, gold_gst_rate)
        row_diamond_amt = _calc_diamond_amount(doc, bom_doc, customer_group, use_handling_charges=True)
        row_gemstone_amt = _calc_gemstone_amount(doc, bom_doc, customer_group, gemstone_price_list_type)

        row.metal_amount = row_metal_amt
        row.finding_amount = row_finding_amt
        row.diamond_amount = row_diamond_amt
        row.gemstone_amount = row_gemstone_amt

        # Making charges accumulated from BOM
        row_making_amt_total = _calc_making_from_bom(doc, bom_doc, mc_name)

        # Making charges Type
        if doc.making_charges_type == "Without":
            row.making_amount = 0
        elif doc.making_charges_type == "With":
            row.making_amount = row_making_amt_total
        elif doc.making_charges_type == "Half":
            row.making_amount = row_making_amt_total * 0.5
        else:
            row.making_amount = 0

        row.rate = (
            row.metal_amount
            + row.finding_amount
            + row.diamond_amount
            + row.gemstone_amount
            + row.making_amount
        )
        row.amount = row.rate * row.qty

        total_taxable += row.amount

    doc.total_taxes_and_charges = total_gst
    doc.grand_total = total_taxable + total_gst


# ==================================================================
# CALCULATION TYPE: RAW MATERIAL CONSIGNMENT
# ==================================================================
def _calc_bom_raw_material_consignment(doc):
    """
    Raw Material Consignment BOM-based calculation.
    Special rules:
      - making_charges_type forced to 'With'
      - return_material_type forced to 'Diamond-Gemstone'
      - Making = 50% of BOM-based making
      - Material type filter: zero out metal/finding for Diamond-Gemstone
      - Custom Duty = (making + diamond + gemstone) × 6% × 50%
      - Rate = (making + diamond + gemstone) - custom_duty

    Used for:
      - Raw Material-Consignment
    """

    doc.making_charges_type = "With"
    doc.return_material_type = "Diamond-Gemstone"

    jwelex_data = doc.get("jwelex_credit_note_data")
    if not jwelex_data:
        frappe.throw(f"JWELEX Data is required for {frappe.bold(doc.credit_note_subtype)} calculation")
    
    jwelex_data = json.loads(jwelex_data)

    # TODO: Implement PCPM calculation logic
    # Use jwelex_data to calculate rates
    # Apply same product rate as invoice
    # No new or revised rates
    
    for row_item in doc.items:
        factor = flt(row_item.qty)
        tag_no = row_item.get("tag_no")
        if not tag_no:
            continue
        
        # Find matching item in jwelex_data
        matching_item = jwelex_data.get(tag_no)
        
        if not matching_item:
            frappe.throw(f"Tag No {tag_no} not found in JWELEX Data")
        
        # frappe.throw(f"Matching Item: {matching_item}")

        charges_info = matching_item.get("charges_info", {})
        making_amount = charges_info.get("chain_making_amount", 0) + charges_info.get("metal_making_amount", 0)
        hm_charges = charges_info.get("hm_charges", 0)
        certificate_charges = charges_info.get("certificate_charges", 0)

        materials = matching_item.get("materials", {})
        diamond_details = materials.get("diamond_details", [])
        finding_details = materials.get("finding_details", [])
        metal_details = materials.get("metal_details", [])
        other_details = materials.get("other_details", [])
        stone_details = materials.get("stone_details", [])


        metal_amount = 0
        for metal in metal_details:
            metal_amount += metal.get("Amount", 0)
            
        diamon_amount = 0
        for diamond in diamond_details:
            diamon_amount += diamond.get("Amount", 0)
        
        finding_amount = 0
        for finding in finding_details:
            finding_amount += finding.get("Amount", 0)
        
        stone_amount = 0
        for stone in stone_details:
            stone_amount += stone.get("Amount", 0)
        
        other_amount = 0
        for other in other_details:
            other_amount += other.get("Amount", 0)


        # Making charges Type
        if doc.making_charges_type == "Without":
            making_amount = 0
        elif doc.making_charges_type == "With":
            making_amount = making_amount
        elif doc.making_charges_type == "Half":
            making_amount = making_amount * 0.5
        else:
            making_amount = 0
        
        proportional_fields = [
            "amount",
            "base_amount",
            "metal_amount",
            "diamond_amount",
            "finding_amount",
            "making_amount",
            "certification_amount",
            "freight_amount",
            "gemstone_amount",
            "other_material_amount",
            "hallmarking_amount",
            "custom_duty_amount",
            "other_amount",
            ]

        for field in proportional_fields:
            row_item.set(field, 0)
        
        

        # makeing charge: 50% labour charges
        row_item.set("making_amount", making_amount * factor * 0.5)
        
        # diamond and gemstone amount as invoice item amount
        row_item.set("diamond_amount", diamon_amount * factor)
        row_item.set("gemstone_amount", stone_amount * factor)

        if doc.return_material_type == "Metal-Finding":
            row_item.set("diamond_amount", 0)
            row_item.set("gemstone_amount", 0)
        elif doc.return_material_type == "Diamond-Gemstone":
            row_item.set("metal_amount", 0)
            row_item.set("finding_amount", 0)

        total_amount = row_item.diamond_amount + row_item.gemstone_amount + row_item.making_amount - row_item.custom_duty_amount
        # frappe.msgprint(f"Total Amount: {total_amount}")
        
        row_item.rate = total_amount / row_item.get("qty", 1)
        row_item.amount = total_amount


@frappe.whitelist()
def get_data_from_jwelex(tag_no):

    url = f"http://3.108.219.130:8001//credit-note?tag_no={tag_no}"
    
    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()  # raises exception for 4xx/5xx
        
        data = response.json()
        # frappe.msgprint(frappe.as_json(data))
        return data
    except requests.exceptions.RequestException:
        return None
    except ValueError:
        return None
    
    
    
import frappe

@frappe.whitelist()
def create_credit_note_orders(source_doctype, source_name, child_fieldname):
    """
    Loops through child rows of the source document, fetches Jwellex data
    for each tag_no, and creates a Credit Note Order record per row.
    """
    source_doc = frappe.get_doc(source_doctype, source_name)
    child_rows = source_doc.get(child_fieldname)

    created = []
    errors = []

    for row in child_rows:
        tag_no = row.get("tag_no")
        if not tag_no:
            continue

        try:
            data = get_data_from_jwelex(tag_no=tag_no)  # reuse existing function directly
            if not data:
                errors.append({"tag_no": tag_no, "error": "No data found"})
                continue

            cno = frappe.new_doc("Credit Note Order")

            materials = data.get("materials", {})
            totals = data.get("totals", {})
            summary = data.get("summary_totals", {})

            metal_details = materials.get("metal_details", [])
            diamond_details = materials.get("diamond_details", [])
            finding_details = materials.get("finding_details", [])
            stone_details = materials.get("stone_details", [])
            other_details = materials.get("other_details", [])

            # --- Parent-level summary fields ---
            cno.total_metal_weight = totals.get("metal_totals", {}).get("total_weight")

            cno.total_finding_pcs = len(finding_details)
            cno.total_finding_weightin_gms = totals.get("finding_totals", {}).get("total_weight")

            total_dia_pcs = sum((d.get("Pcs") or 0) for d in diamond_details)
            cno.total_diamond_pcs = total_dia_pcs
            cno.total_diamond_weight = totals.get("diamond_totals", {}).get("total_weight")
            cno.total_diamond_weightin_gms = totals.get("diamond_totals", {}).get("total_gross_wt")
            cno.total_diamond_weight_in_gram = totals.get("diamond_totals", {}).get("total_gross_wt")

            cno.total_gemstone_pcs = len(stone_details)
            cno.total_gemstone_weight = totals.get("stone_totals", {}).get("total_weight")
            cno.total_gemstone_weightin_gms = totals.get("stone_totals", {}).get("total_gross_wt")
            cno.total_gemstone_weight_in_gram = totals.get("stone_totals", {}).get("total_gross_wt")

            cno.total_other_pcs = len(other_details)
            cno.total_other_weight = totals.get("other_totals", {}).get("total_weight")

            cno.metal_weight_in_gram = summary.get("metal_weight")
            cno.gross_weight_in_gram = summary.get("gross_wt")
            cno.net_weight_in_gram = (summary.get("metal_weight") or 0) + (summary.get("chain_weight") or 0)
            cno.diamond_weight_in_carat = summary.get("diamond_weight")
            cno.gemstone_weight_in_carat = summary.get("stone_weight")
            cno.other_weight_in_gram = summary.get("other_weight")
            cno.finding_weight_in_gram = summary.get("chain_weight")
            cno.avg_diamond_weightin_carat = (
                (summary.get("diamond_weight") / total_dia_pcs) if total_dia_pcs else 0
            )

            # --- Child tables: mapped individually ---
            _append_metal_detail(cno, metal_details)
            _append_diamond_detail(cno, diamond_details)
            _append_finding_detail(cno, finding_details)
            _append_gemstone_detail(cno, stone_details)
            _append_other_detail(cno, other_details)

            cno.insert(ignore_permissions=True)
            created.append({"tag_no": tag_no, "name": cno.name})

        except Exception as e:
            frappe.log_error(frappe.get_traceback(), f"Credit Note Order creation failed for {tag_no}")
            errors.append({"tag_no": tag_no, "error": str(e)})

    frappe.db.commit()
    return {"created": created, "errors": errors}


def _append_metal_detail(cno, items):
    """Table: metal_detail -> Order BOM Metal Detail"""
    for m in items:
        cno.append("metal_detail", {
            "metal_type": m.get("Meterial"),
            "shape": m.get("Shape_Name"),
            "metal_purity": "76.0",
            "metal_colour": "Yellow",
            "size": m.get("Size_Name"),
            "code": m.get("Code_Name"),
            "pcs": m.get("Pcs"),
            "finish_product_weight": m.get("Weight"),
            "gross_weight": m.get("Gross_Wt"),
            "rate": m.get("Rate"),
            "amount": m.get("Amount"),
        })


def _append_diamond_detail(cno, items):
    """Table: diamond_detail -> Order BOM Diamond Detail"""
    for d in items:
        cno.append("diamond_detail", {
            "material": d.get("Meterial"),
            "stone_shape": d.get("Shape_Name"),
            "sub_setting_type": "Prong Setting",
            "purity": d.get("Purity_Name"),
            "size": d.get("Size_Name"),
            "code": d.get("Code_Name"),
            "pcs": d.get("Pcs"),
            "weight": d.get("Weight"),
            "gross_weight": d.get("Gross_Wt"),
            "rate": d.get("Rate"),
            "amount": d.get("Amount"),
        })


def _append_finding_detail(cno, items):
    """Table: finding_detail -> Order BOM Finding Detail"""
    for f in items:
        cno.append("finding_detail", {
            "material": f.get("Meterial"),
            "shape": f.get("Shape_Name"),
            "purity": f.get("Purity_Name"),
            "size": f.get("Size_Name"),
            "code": f.get("Code_Name"),
            "pcs": f.get("Pcs"),
            "weight": f.get("Weight"),
            "gross_weight": f.get("Gross_Wt"),
            "rate": f.get("Rate"),
            "amount": f.get("Amount"),
        })


def _append_gemstone_detail(cno, items):
    """Table: gemstone_detail -> Order BOM Gemstone Detail (source: stone_details)"""
    for s in items:
        cno.append("gemstone_detail", {
            "material": s.get("Meterial"),
            "shape": s.get("Shape_Name"),
            "purity": s.get("Purity_Name"),
            "size": s.get("Size_Name"),
            "code": s.get("Code_Name"),
            "pcs": s.get("Pcs"),
            "weight": s.get("Weight"),
            "gross_weight": s.get("Gross_Wt"),
            "rate": s.get("Rate"),
            "amount": s.get("Amount"),
        })


def _append_other_detail(cno, items):
    """Table: other_detail -> Order BOM Other Detail"""
    for o in items:
        cno.append("other_detail", {
            "material": o.get("Meterial"),
            "shape": o.get("Shape_Name"),
            "purity": o.get("Purity_Name"),
            "size": o.get("Size_Name"),
            "code": o.get("Code_Name"),
            "pcs": o.get("Pcs"),
            "weight": o.get("Weight"),
            "gross_weight": o.get("Gross_Wt"),
            "rate": o.get("Rate"),
            "amount": o.get("Amount"),
        })


# ==================================================================
# EXCEL PREVIEW — Product Return Order Form
# ==================================================================
@frappe.whitelist()
def xl_preview_product_return_order_form(docname):
    """
    Excel Preview for Product Return Order Form, mirroring the Sales Order
    "Excel Preview" button — one row per Product Return Order created
    against this form. Every linked Product Return Order must be in the
    "Send For Approval" workflow_state, or the whole export is blocked.
    """
    rows = frappe.get_all(
        "Product Return Order",
        filters={"product_return_order_form": docname},
        fields=[
            "name", "workflow_state", "index", "item_code", "serial_no", "jewelex_tag",
            "bom", "item_category", "item_subcategory", "metal_touch", "metal_purity",
            "metal_colour", "setting_type", "net_weight", "gross_weight", "diamond_weight",
            "total_diamond_pcs", "total_gemstone_pcs", "total_finding_pcs",
            "metal_amount", "diamond_amount", "finding_amount", "gemstone_amount",
            "making_amount", "hallmarking_amount", "certification_amount", "rate", "amount",
        ],
        order_by="`index` asc",
    )

    if not rows:
        frappe.throw(f"No Product Return Order found for {frappe.bold(docname)}")

    for row in rows:
        if row.workflow_state != "Send For Approval":
            frappe.throw(
                f"Product Return Order {frappe.bold(row.name)} is not in Send For Approval State"
            )

    columns = [
        "Index", "Item Code", "Serial No", "Jewelex Tag", "BOM", "Item Category",
        "Item Subcategory", "Metal Touch", "Metal Purity", "Metal Colour", "Setting Type",
        "Net Weight", "Gross Weight", "Diamond Weight", "Diamond Pcs", "Gemstone Pcs",
        "Finding Pcs", "Metal Amount", "Diamond Amount", "Finding Amount", "Gemstone Amount",
        "Making Amount", "Hallmarking Amount", "Certification Amount", "Rate", "Amount",
    ]

    data_rows = []
    for row in rows:
        data_rows.append([
            row.index, row.item_code, row.serial_no, row.jewelex_tag, row.bom,
            row.item_category, row.item_subcategory, row.metal_touch, row.metal_purity,
            row.metal_colour, row.setting_type,
            flt(row.net_weight), flt(row.gross_weight), flt(row.diamond_weight),
            flt(row.total_diamond_pcs), flt(row.total_gemstone_pcs), flt(row.total_finding_pcs),
            flt(row.metal_amount), flt(row.diamond_amount), flt(row.finding_amount),
            flt(row.gemstone_amount), flt(row.making_amount), flt(row.hallmarking_amount),
            flt(row.certification_amount), flt(row.rate), flt(row.amount),
        ])

    sum_row = [""] * len(columns)
    for idx in range(11, len(columns)):
        sum_row[idx] = round(sum(flt(r[idx]) for r in data_rows), 3)
    data_rows.append(sum_row)

    wb = Workbook()
    ws = wb.active
    ws.title = "Product Return Order"

    company_name = frappe.db.get_value("Product Return Order Form", docname, "company") or ""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    cell = ws.cell(row=1, column=1, value=company_name)
    cell.font = Font(bold=True, size=15)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_num, column_title in enumerate(columns, 1):
        c = ws.cell(row=2, column=col_num, value=column_title)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")

    for row_num, row_data in enumerate(data_rows, 3):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    for i in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    frappe.local.response.filecontent = output.read()
    frappe.local.response.filename = f"Product_Return_Order_{docname}.xlsx"
    frappe.local.response.type = "download"


# ==================================================================
# EXCEL FROM PRINT FORMAT — renders whichever Print Format the user
# picks, live, and converts that HTML into .xlsx. No hand-copied
# calculation logic here — any edit to the Print Format is picked up
# automatically the next time this is called, since it always renders
# the current version.
# ==================================================================
@frappe.whitelist()
def get_print_formats_for_product_return_order_form():
    return frappe.get_all(
        "Print Format",
        filters={"doc_type": "Product Return Order Form", "disabled": 0},
        fields=["name"],
        order_by="name asc",
    )


_HEADING_SIZE = {"h1": 16, "h2": 15, "h3": 14, "h4": 13, "h5": 12, "h6": 11}
_THIN_SIDE = Side(style="thin", color="999999")
_CELL_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)
_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")


def _is_bold_tag(tag):
    if tag.name in ("b", "strong", "h1", "h2", "h3", "h4", "h5", "h6"):
        return True
    if tag.find(["b", "strong"]):
        return True
    style = (tag.get("style") or "").replace(" ", "").lower()
    return "font-weight:bold" in style


def _is_underline_tag(tag):
    if tag.name == "u" or tag.find("u"):
        return True
    style = (tag.get("style") or "").replace(" ", "").lower()
    return "text-decoration:underline" in style


def _get_align(tag):
    style = (tag.get("style") or "").replace(" ", "").lower()
    if "text-align:center" in style:
        return "center"
    if "text-align:right" in style:
        return "right"
    return "left"


def _clean_cell_value(text):
    text = (text or "").strip()
    if not text:
        return None
    cleaned = text.replace("₹", "").replace(",", "").strip()
    try:
        value = float(cleaned)
        return int(value) if value == int(value) else value
    except ValueError:
        return text


@frappe.whitelist()
def xl_from_print_format(docname, print_format):
    """
    Render `print_format` against Product Return Order Form `docname` the
    same way Frappe renders it for PDF/print, then convert that HTML into
    an .xlsx download.
    """
    from bs4 import BeautifulSoup

    pf = frappe.db.get_value(
        "Print Format", print_format, ["name", "doc_type", "disabled"], as_dict=True
    )
    if not pf:
        frappe.throw(f"Print Format {frappe.bold(print_format)} not found")
    if pf.doc_type != "Product Return Order Form":
        frappe.throw("This Print Format is not available for Product Return Order Form")
    if pf.disabled:
        frappe.throw(f"Print Format {frappe.bold(print_format)} is disabled")

    html = frappe.get_print(
        doctype="Product Return Order Form", name=docname, print_format=print_format, as_pdf=False
    )
    soup = BeautifulSoup(html, "html.parser")
    # frappe.get_print(as_pdf=False) returns the full print-preview page
    # (toolbar, huge CSS block, etc.), and the print format's own template
    # can start with its own <html><body> — so the page ends up with two
    # nested <body> tags. soup.body would grab the outer page shell, not
    # the actual rendered content. The content is always wrapped in
    # <div class="print-format">, regardless of how the format's own
    # template is written, so anchor there instead.
    body = soup.find("div", class_="print-format") or soup.body or soup

    wb = Workbook()
    ws = wb.active
    safe_title = "".join(c for c in print_format if c not in '/\\?*[]:') or "Sheet1"
    ws.title = safe_title[:31]

    cursor = {"row": 1}
    heading_rows = []   # (row_num, align) — merged/centered once we know sheet width
    table_ranges = []   # (start_row, end_row, start_col, end_col)
    table_counter = {"count": 0}
    # The first table (item details) starts at column A. Any table after
    # that (totals/summary) is positioned to the right of the page in the
    # real print format via CSS, not by leading blank cells in the HTML —
    # since that layout info is lost once we're just reading table
    # structure, shift those later tables over to column P to approximate it.
    SUMMARY_TABLE_COL_OFFSET = 15

    def write_text_block(tag):
        text = tag.get_text(" ", strip=True)
        if not text:
            return
        r = cursor["row"]
        cell = ws.cell(row=r, column=1, value=text)
        cell.font = Font(
            bold=_is_bold_tag(tag),
            underline="single" if _is_underline_tag(tag) else None,
            size=_HEADING_SIZE.get(tag.name, 10),
        )
        align = _get_align(tag)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        heading_rows.append((r, align))
        cursor["row"] += 1

    def write_table(table):
        start_row = cursor["row"]
        # Some print formats emit "logic-only" <tr> rows with no <td>/<th> at
        # all (Jinja {% set %} accumulators with nothing rendered inside) —
        # skip those instead of letting them consume a blank grid row.
        rows_html = [
            tr for tr in table.find_all("tr")
            if tr.find_all(["td", "th"], recursive=False)
        ]
        col_offset = SUMMARY_TABLE_COL_OFFSET if table_counter["count"] > 0 else 0
        occupied = set()
        max_rows_used = 0
        max_col_used = 0
        for r_idx, tr in enumerate(rows_html):
            col_idx = 0
            for cell_tag in tr.find_all(["td", "th"], recursive=False):
                while (r_idx, col_idx) in occupied:
                    col_idx += 1
                colspan = int(cell_tag.get("colspan", 1) or 1)
                rowspan = int(cell_tag.get("rowspan", 1) or 1)
                value = _clean_cell_value(cell_tag.get_text(" ", strip=True))

                target_row = start_row + r_idx
                target_col = col_idx + 1 + col_offset
                cell = ws.cell(row=target_row, column=target_col, value=value)

                is_header_row = r_idx == 0
                cell.font = Font(bold=_is_bold_tag(cell_tag) or is_header_row)
                is_numeric = isinstance(value, (int, float))
                if is_header_row:
                    align = "center"
                elif is_numeric:
                    # Right-align single-width numeric cells (normal
                    # spreadsheet convention), but left-align ones that span
                    # multiple columns — different rows in the same table
                    # can merge a different number of columns for what's
                    # otherwise the same "value" position, and right-aligning
                    # would drift each one to a different visual column even
                    # though they all start at the same column.
                    align = "left" if colspan > 1 else "right"
                else:
                    align = "left"
                cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
                if is_numeric and not is_header_row:
                    cell.number_format = "#,##0" if isinstance(value, int) else "#,##0.00"
                if is_header_row:
                    cell.fill = _HEADER_FILL

                if colspan > 1 or rowspan > 1:
                    ws.merge_cells(
                        start_row=target_row, start_column=target_col,
                        end_row=target_row + rowspan - 1, end_column=target_col + colspan - 1,
                    )
                for rr in range(rowspan):
                    for cc in range(colspan):
                        occupied.add((r_idx + rr, col_idx + cc))
                col_idx += colspan
            max_rows_used = max(max_rows_used, r_idx + 1)
            max_col_used = max(max_col_used, col_idx)

        end_row = start_row + max_rows_used - 1

        # Weight columns/rows always show 3 decimals, regardless of the
        # generic int/2-decimal formatting applied above — detected by
        # "wt"/"weight" appearing in the column header (item table) or in
        # the row's own label (row-oriented tables like the totals block).
        weight_cols = set()
        for c in range(1 + col_offset, max_col_used + 1 + col_offset):
            header_text = str(ws.cell(row=start_row, column=c).value or "").lower()
            if "wt" in header_text or "weight" in header_text:
                weight_cols.add(c)
        for r in range(start_row, end_row + 1):
            label_text = str(ws.cell(row=r, column=1 + col_offset).value or "").lower()
            row_is_weight = "wt" in label_text or "weight" in label_text
            for c in range(1 + col_offset, max_col_used + 1 + col_offset):
                if c not in weight_cols and not row_is_weight:
                    continue
                cell = ws.cell(row=r, column=c)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.000"

        # Border every cell in the grid, including blanks, so it reads as
        # an actual table instead of loose floating values.
        for r in range(start_row, end_row + 1):
            for c in range(1 + col_offset, max_col_used + 1 + col_offset):
                ws.cell(row=r, column=c).border = _CELL_BORDER

        table_ranges.append((start_row, end_row, 1 + col_offset, max_col_used + col_offset))
        table_counter["count"] += 1
        cursor["row"] = end_row + 2

    def walk(node):
        for child in node.find_all(recursive=False):
            if child.name == "table":
                write_table(child)
            elif child.name in ("h1", "h2", "h3", "h4", "h5", "h6", "p"):
                write_text_block(child)
            elif child.name in ("div", "body", "section", "html"):
                walk(child)
            # skip <br>, <head>, <meta>, <script>, <style>, etc.

    walk(body)

    max_col = max((r[3] for r in table_ranges), default=10)

    # Merge each heading/paragraph line across the sheet's full width so it
    # reads as a banner instead of text stuck in column A.
    for row_num, align in heading_rows:
        if max_col > 1:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=max_col)
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

    # Auto-size columns to their content, capped so one long cell can't
    # blow out the whole sheet.
    col_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value in (None, ""):
                continue
            length = len(str(cell.value))
            col_widths[cell.column] = max(col_widths.get(cell.column, 0), length)
    for col_num in range(1, max_col + 1):
        width = col_widths.get(col_num, 10)
        ws.column_dimensions[get_column_letter(col_num)].width = min(max(width + 2, 10), 35)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    frappe.local.response.filecontent = output.read()
    frappe.local.response.filename = f"{safe_title}_{docname}.xlsx"
    frappe.local.response.type = "download"


