# Copyright (c) 2023, Nirali and contributors
# For license information, please see license.txt

import frappe,json
from openpyxl.styles import Alignment, Font
from frappe import _
from frappe.utils import get_link_to_form
from frappe.model.document import Document
from frappe.model.mapper import get_mapped_doc
from erpnext.controllers.item_variant import (
	ItemVariantExistsError,
	copy_attributes_to_variant,
	get_variant,
	make_variant_item_code,
	validate_item_variant_attributes,
)
from frappe.utils import (
	cint,
	cstr,
	flt,
	formatdate,
	get_link_to_form,
	getdate,
	now_datetime,
	nowtime,
	strip,
	strip_html,
)
from frappe.utils.file_manager import save_file
import openpyxl
from io import BytesIO
import requests
import os

class OrderForm(Document):
	
	def on_submit(self):
		create_cad_orders(self)
		if self.supplier:
			create_po(self)


	def on_update_after_submit(self):
		if self.updated_delivery_date:
			order_names = frappe.get_all(
				"Order",
				filters={"cad_order_form": self.name},
				pluck="name"
			)

			for order_name in order_names:
				frappe.db.set_value("Order", order_name, "updated_delivery_date", self.updated_delivery_date)
				
	# def on_cancel(self):
	# # 	delete_auto_created_cad_order(self)
	# def on_cancel(self):
	# 	if frappe.db.get_list("Order",filters={"cad_order_form":self.name},fields="name"):
	# 		for order in frappe.db.get_list("Order",filters={"cad_order_form":self.name},fields="name"):
	# 			frappe.db.set_value("Order",order["name"],"workflow_state","Cancelled")
	# 			frappe.throw(f"{order}")
	# 			if frappe.db.get_list("Timesheet",filters={"order":order["name"]},fields="name"):
	# 				for timesheet in frappe.db.get_list("Timesheet",filters={"order":order["name"]},fields="name"):
	# 					frappe.db.set_value("Timesheet",timesheet["name"],"docstatus","2")

	# 	frappe.db.set_value("Order Form",self.name,"workflow_state","Cancelled")
	# 	self.reload()

	def on_cancel(self):
		orders = frappe.db.get_list("Order", filters={"cad_order_form": self.name}, fields="name")
		if orders:
			for order in orders:
				timesheets = frappe.db.get_list("Timesheet", filters={"order": order["name"]}, fields="name")
				frappe.db.set_value("Order", order["name"], "workflow_state", "Cancelled")
				for ts in timesheets:
					frappe.db.set_value("Timesheet", ts["name"], "workflow_state", "Cancelled")
		frappe.db.set_value("Order Form", self.name, "workflow_state", "Cancelled")
		self.reload()


	def validate(self):
		self.validate_category_subcaegory()
		self.validate_filed_value()
		validate_rows_on_order_criteria(self)
		validate_design_id(self)
		validate_item_variant(self)
		validate_is_mannual(self)
		set_data(self)
		for i in self.order_details:	
			if i.metal_type == "Silver":
				i.metal_colour = "White"
				i.metal_touch = "20KT"
				i.setting_type = "Open"
				i.diamond_type = "AD"
			# return

	def validate_category_subcaegory(self):
		for row in self.get("order_details"):
			if row.subcategory:
				parent = frappe.db.get_value("Attribute Value", row.subcategory, "parent_attribute_value")
				if row.category != parent:
					frappe.throw(_(f"Category & Sub Category mismatched in row #{row.idx}"))
	
	def validate_filed_value(self):
		for row in self.get("order_details"):
			if row.design_type:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Design Type").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.design_type not in attribute_list:
					frappe.throw("Design Type is not Correct")

			if row.diamond_quality:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Diamond Quality").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.diamond_quality not in attribute_list:
					frappe.throw("Diamond Quality is not Correct")

			if row.setting_type:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Setting Type").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.setting_type not in attribute_list:
					frappe.throw("Setting Type is not Correct")
			
			if row.sub_setting_type1:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Sub Setting Type1").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.sub_setting_type1 not in attribute_list:
					frappe.throw("Setting Type 1 is not Correct")
			
			if row.sub_setting_type2:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Sub Setting Type2").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.sub_setting_type2 not in attribute_list:
					frappe.throw("Setting Type 2 is not Correct")
			
			if row.metal_type:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Metal Type").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.metal_type not in attribute_list:
					frappe.throw("Metal Type is not Correct")

			if row.metal_touch:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Metal Touch").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.metal_touch not in attribute_list:
					frappe.throw("Metal Touch is not Correct")

			if row.metal_colour:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Metal Colour").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.metal_colour not in attribute_list:
					frappe.throw("Metal Colour is not Correct")

			if row.diamond_type:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Diamond Type").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.diamond_type not in attribute_list:
					frappe.throw("Diamond Type is not Correct")
			
			if row.sizer_type:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Sizer Type").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.sizer_type not in attribute_list:
					frappe.throw("Sizer Type is not Correct")

			if row.stone_changeable:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Stone Changeable").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.stone_changeable not in attribute_list:
					frappe.throw("Stone Changeable is not Correct")

			if row.feature:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Feature").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.feature not in attribute_list:
					frappe.throw("Feature is not Correct")

			if row.rhodium:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Rhodium").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.rhodium not in attribute_list:
					frappe.throw("Rhodium is not Correct")

			if row.enamal:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Enamal").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.enamal not in attribute_list:
					frappe.throw("Enamal is not Correct")
			
			if row.gemstone_type:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Gemstone Type").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.gemstone_type not in attribute_list:
					frappe.throw("Gemstone Type is not Correct")
			
			if row.gemstone_quality:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Gemstone Quality").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.gemstone_quality not in attribute_list:
					frappe.throw("Gemstone Quality is not Correct")
			
			if row.mod_reason:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Mod Reason").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.mod_reason not in attribute_list:
					frappe.throw("Mod Reason is not Correct")
			
			if row.finding_category:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Finding Category").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.finding_category not in attribute_list:
					frappe.throw("Finding Category is not Correct")

			if row.finding_subcategory:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Finding Sub-Category").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.finding_subcategory not in attribute_list:
					frappe.throw("Finding Sub-Category is not Correct")
			
			if row.finding_size:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Finding Size").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.finding_size not in attribute_list:
					frappe.throw("Finding Size is not Correct")
			
			if row.metal_target_from_range:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Metal Target Range").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.metal_target_from_range not in attribute_list:
					frappe.throw("Metal Target Range is not Correct")

			if row.diamond_target_from_range:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Diamond Target Range").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.diamond_target_from_range not in attribute_list:
					frappe.throw("Diamond Target Range is not Correct")
			
			if row.detachable:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Detachable").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.detachable not in attribute_list:
					frappe.throw("Detachable is not Correct")
			
			if row.lock_type:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Lock Type").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.lock_type not in attribute_list:
					frappe.throw("Lock Type is not Correct")
			
			if row.capganthan:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Cap/Ganthan").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.capganthan not in attribute_list:
					frappe.throw("Cap/Ganthan is not Correct")
			
			if row.charm:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Charm").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.charm not in attribute_list:
					frappe.throw("Charm is not Correct")
			
			if row.back_chain:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Back Chain").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.back_chain not in attribute_list:
					frappe.throw("ChBack Chainarm is not Correct")

			if row.back_belt:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Back Belt").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.back_belt not in attribute_list:
					frappe.throw("Back Belt is not Correct")

			if row.black_bead:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Black Bead").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.black_bead not in attribute_list:
					frappe.throw("Black Bead is not Correct")

			if row.two_in_one:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","2 in 1").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.two_in_one not in attribute_list:
					frappe.throw("2 in 1 is not Correct")

			if row.chain_type:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Chain Type").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.chain_type not in attribute_list:
					frappe.throw("Chain Type is not Correct")

			# if row.customer_chain:
			# 	attribute_list = []
			# 	for i in frappe.get_doc("Item Attribute","Black Bead").item_attribute_values:
			# 		attribute_list.append(i.attribute_value)
			# 	if row.customer_chain not in attribute_list:
			# 		frappe.throw("Black Bead is not Correct")

			if row.nakshi_from:
				attribute_list = []
				for i in frappe.get_doc("Item Attribute","Nakshi From").item_attribute_values:
					attribute_list.append(i.attribute_value)
				if row.nakshi_from not in attribute_list:
					frappe.throw("Nakshi From is not Correct")


		



# from datetime import datetime, timedelta, time
# import frappe
# from frappe.utils import now_datetime, get_datetime
# from frappe.utils import get_link_to_form
# from frappe import _

# def create_cad_orders(self):
#     doclist = []

   
#     order_criteria = frappe.get_single("Order Criteria")
#     criteria_rows = order_criteria.get("order")
#     enabled_criteria = next((row for row in criteria_rows if not row.disable), None)

#     if not enabled_criteria:
#         frappe.throw("No enabled Order Criteria found.")

    
#     cad_days = int(enabled_criteria.cad_approval_day or 0)

   
#     cad_time_raw = enabled_criteria.cad_submission_time
#     if isinstance(cad_time_raw, time):
#         cad_time = cad_time_raw
#     elif isinstance(cad_time_raw, timedelta):
#         cad_time = (datetime.min + cad_time_raw).time()
#     elif isinstance(cad_time_raw, str):
#         try:
#             h, m, s = [int(x) for x in cad_time_raw.strip().split(".")]
#             cad_time = time(h, m, s)
#         except:
#             frappe.throw("Invalid CAD Submission Time format.")
#     else:
#         cad_time = time(0, 0, 0)

    
#     ibm_time_raw = enabled_criteria.cad_appoval_timefrom_ibm_team
#     if isinstance(ibm_time_raw, time):
#         ibm_timedelta = timedelta(hours=ibm_time_raw.hour, minutes=ibm_time_raw.minute, seconds=ibm_time_raw.second)
#     elif isinstance(ibm_time_raw, timedelta):
#         ibm_timedelta = ibm_time_raw
#     elif isinstance(ibm_time_raw, str):
#         try:
#             h, m, s = [int(x) for x in ibm_time_raw.strip().split(".")]
#             ibm_timedelta = timedelta(hours=h, minutes=m, seconds=s)
#         except:
#             frappe.throw("Invalid IBM Approval Time format.")
#     else:
#         ibm_timedelta = timedelta()

#     for row in self.order_details:
#         docname = make_cad_order(row.name, parent_doc=self)
#         if row.pre_order_form_details:
#             frappe.db.set_value("Pre Order Form Details", row.pre_order_form_details, "order_form_id", self.name)

#         order_datetime = now_datetime()
#         frappe.db.set_value("Order", docname, "order_date", order_datetime)

#         if self.delivery_date:
#             frappe.db.set_value("Order", docname, "delivery_date", self.delivery_date)

#         cad_delivery_datetime = datetime.combine(order_datetime.date() + timedelta(days=cad_days), cad_time)
#         ibm_delivery_datetime = cad_delivery_datetime + ibm_timedelta

#         frappe.db.set_value("Order", docname, "cad_delivery_date", cad_delivery_datetime)
#         frappe.db.set_value("Order", docname, "ibm_delivery_date", ibm_delivery_datetime)

       
#         doclist.append(get_link_to_form("Order", docname))

   
#     if doclist:
#         msg = _("The following {0} were created: {1}").format(
#             frappe.bold(_("Orders")), "<br>" + ", ".join(doclist)
#         )
#         frappe.msgprint(msg)


# import frappe
# from frappe.utils import now_datetime, get_datetime
# from frappe.utils import get_link_to_form
# from frappe import _
# from datetime import datetime, time, timedelta

# def create_cad_orders(self):
    
#     if self.docstatus == 0 or self.workflow_state in ["Draft","Send For Approval", "Cancelled"]:
#         frappe.msgprint(_("Order creation skipped because document is in Draft or Cancelled state."))
#         return

#     doclist = []

#     order_criteria = frappe.get_single("Order Criteria")
#     criteria_rows = order_criteria.get("order")
#     enabled_criteria = next((row for row in criteria_rows if not row.disable), None)

#     if not enabled_criteria:
#         frappe.throw("No enabled Order Criteria found.")

#     cad_days = int(enabled_criteria.cad_approval_day or 0)

#     cad_time_raw = enabled_criteria.cad_submission_time
#     if isinstance(cad_time_raw, time):
#         cad_time = cad_time_raw
#     elif isinstance(cad_time_raw, timedelta):
#         cad_time = (datetime.min + cad_time_raw).time()
#     elif isinstance(cad_time_raw, str):
#         try:
#             h, m, s = [int(x) for x in cad_time_raw.strip().split(".")]
#             cad_time = time(h, m, s)
#         except:
#             frappe.throw("Invalid CAD Submission Time format.")
#     else:
#         cad_time = time(0, 0, 0)

#     ibm_time_raw = enabled_criteria.cad_appoval_timefrom_ibm_team
#     if isinstance(ibm_time_raw, time):
#         ibm_timedelta = timedelta(hours=ibm_time_raw.hour, minutes=ibm_time_raw.minute, seconds=ibm_time_raw.second)
#     elif isinstance(ibm_time_raw, timedelta):
#         ibm_timedelta = ibm_time_raw
#     elif isinstance(ibm_time_raw, str):
#         try:
#             h, m, s = [int(x) for x in ibm_time_raw.strip().split(".")]
#             ibm_timedelta = timedelta(hours=h, minutes=m, seconds=s)
#         except:
#             frappe.throw("Invalid IBM Approval Time format.")
#     else:
#         ibm_timedelta = timedelta()

#     for row in self.order_details:
#         docname = make_cad_order(row.name, parent_doc=self)

#         if row.pre_order_form_details:
#             frappe.db.set_value("Pre Order Form Details", row.pre_order_form_details, "order_form_id", self.name)

#         order_datetime = now_datetime()
#         frappe.db.set_value("Order", docname, "order_date", order_datetime)

#         if self.delivery_date:
#             frappe.db.set_value("Order", docname, "delivery_date", self.delivery_date)

#         cad_delivery_datetime = datetime.combine(order_datetime.date() + timedelta(days=cad_days), cad_time)
#         ibm_delivery_datetime = cad_delivery_datetime + ibm_timedelta

#         frappe.db.set_value("Order", docname, "cad_delivery_date", cad_delivery_datetime)
#         frappe.db.set_value("Order", docname, "ibm_delivery_date", ibm_delivery_datetime)

#         doclist.append(get_link_to_form("Order", docname))

#     if doclist:
#         msg = _("The following {0} were created: {1}").format(
#             frappe.bold(_("Orders")), "<br>" + ", ".join(doclist)
#         )
#         frappe.msgprint(msg)


import frappe
from frappe.utils import now_datetime, get_datetime
from frappe.utils import get_link_to_form
from frappe import _
from datetime import datetime, time, timedelta

def create_cad_orders(self):
    
    if self.docstatus == 0 or self.workflow_state in ["Draft","Send For Approval", "Cancelled"]:
        frappe.msgprint(_("Order creation skipped because document is in Draft or Cancelled state."))
        return

    doclist = []

    order_criteria = frappe.get_single("Order Criteria")
    criteria_rows = order_criteria.get("order")
    enabled_criteria = next((row for row in criteria_rows if not row.disable), None)

    if not enabled_criteria:
        frappe.throw("No enabled Order Criteria found.")

    cad_days = int(enabled_criteria.cad_approval_day or 0)

    # ------------------ CAD Submission Time ------------------
    cad_time_raw = enabled_criteria.cad_submission_time
    if isinstance(cad_time_raw, time):
        cad_time = cad_time_raw
    elif isinstance(cad_time_raw, timedelta):
        cad_time = (datetime.min + cad_time_raw).time()
    elif isinstance(cad_time_raw, str):
        try:
            h, m, s = [int(x) for x in cad_time_raw.strip().split(".")]
            cad_time = time(h, m, s)
        except:
            frappe.throw("Invalid CAD Submission Time format.")
    else:
        cad_time = time(0, 0, 0)

    # ------------------ IBM Approval Time ------------------
    ibm_time_raw = enabled_criteria.cad_appoval_timefrom_ibm_team
    if isinstance(ibm_time_raw, time):
        ibm_hours_total = ibm_time_raw.hour + ibm_time_raw.minute / 60 + ibm_time_raw.second / 3600
    elif isinstance(ibm_time_raw, timedelta):
        ibm_hours_total = ibm_time_raw.total_seconds() / 3600
    elif isinstance(ibm_time_raw, str):
        try:
            h, m, s = [int(x) for x in ibm_time_raw.strip().split(".")]
            ibm_hours_total = h + m / 60 + s / 3600
        except:
            frappe.throw("Invalid IBM Approval Time format.")
    else:
        ibm_hours_total = 0

   
    valid_shift_rows = [row for row in order_criteria.department_shift if not row.disable]
    if not valid_shift_rows:
        frappe.throw("No active (enabled) rows found in Order Criteria 'department_shift' table.")
    latest_shift_row = valid_shift_rows[-1]

    def to_time(value):
        if isinstance(value, timedelta):
            total_seconds = value.total_seconds()
            h = int(total_seconds // 3600)
            m = int((total_seconds % 3600) // 60)
            s = int(total_seconds % 60)
            return time(h, m, s)
        elif isinstance(value, time):
            return value
        elif isinstance(value, str):
            parts = value.split(":")
            return time(int(parts[0]), int(parts[1]), int(parts[2]) if len(parts) > 2 else 0)
        else:
            frappe.throw("Shift times must be timedelta, time, or string in HH:MM:SS format")

    shift_start_time = to_time(latest_shift_row.shift_start_time)
    shift_end_time = to_time(latest_shift_row.shift_end_time)

    # ------------------ Create Orders ------------------
    for row in self.order_details:
        docname = make_cad_order(row.name, parent_doc=self)

        if row.pre_order_form_details:
            frappe.db.set_value("Pre Order Form Details", row.pre_order_form_details, "order_form_id", self.name)

        order_datetime = now_datetime()
        frappe.db.set_value("Order", docname, "order_date", order_datetime)

        if self.delivery_date:
            frappe.db.set_value("Order", docname, "delivery_date", self.delivery_date)

        
        cad_delivery_datetime = datetime.combine(order_datetime.date() + timedelta(days=cad_days), cad_time)

        current_datetime = cad_delivery_datetime
        remaining_hours = ibm_hours_total

        while remaining_hours > 0:
            shift_start_datetime = datetime.combine(current_datetime.date(), shift_start_time)
            shift_end_datetime = datetime.combine(current_datetime.date(), shift_end_time)

            
            if current_datetime >= shift_end_datetime:
                current_datetime = shift_start_datetime + timedelta(days=1)
                shift_start_datetime += timedelta(days=1)
                shift_end_datetime += timedelta(days=1)

            if current_datetime < shift_start_datetime:
                current_datetime = shift_start_datetime

            # Hours available in current shift
            available_hours = (shift_end_datetime - current_datetime).total_seconds() / 3600

            if remaining_hours <= available_hours:
                current_datetime += timedelta(hours=remaining_hours)
                remaining_hours = 0
            else:
                remaining_hours -= available_hours
                current_datetime = shift_start_datetime + timedelta(days=1)

        ibm_delivery_datetime = current_datetime

        frappe.db.set_value("Order", docname, "cad_delivery_date", cad_delivery_datetime)
        frappe.db.set_value("Order", docname, "ibm_delivery_date", ibm_delivery_datetime)

        doclist.append(get_link_to_form("Order", docname))

    
    if doclist:
        msg = _("The following {0} were created: {1}").format(
            frappe.bold(_("Orders")), "<br>" + ", ".join(doclist)
        )
        frappe.msgprint(msg)


def delete_auto_created_cad_order(self):
	for row in frappe.get_all("Order", filters={"order_form": self.name}):
		frappe.delete_doc("Order", row.name)

def make_cad_order(source_name, target_doc=None, parent_doc = None):
	def set_missing_values(source, target):
		target.cad_order_form_detail = source.name
		target.cad_order_form = source.parent
		target.index = source.idx
	
	
	design_type = frappe.get_doc('Order Form Detail',source_name).design_type
	item_type = frappe.get_doc('Order Form Detail',source_name).item_type
	# as_per_serial_no = frappe.get_doc('Order Form Detail',source_name).as_per_serial_no
	mod_reason = frappe.get_doc('Order Form Detail',source_name).mod_reason
	design_id = frappe.get_doc('Order Form Detail',source_name).design_id
	is_repairing = frappe.get_doc('Order Form Detail',source_name).is_repairing
	is_finding_order = frappe.get_doc('Order Form Detail',source_name).is_finding_order
	if design_type == 'Mod - Old Stylebio & Tag No':
		if is_repairing == 1:
			bom_or_cad = frappe.get_doc('Order Form Detail',source_name).bom_or_cad
			item_type = frappe.get_doc('Order Form Detail',source_name).item_type
		else:
		# 	variant_of = frappe.db.get_value("Item",design_id,"variant_of")
		# 	bom = frappe.db.get_value('Item',design_id,'master_bom')
		# 	if bom==None:
		# 		frappe.throw(f'BOM is not available for {design_id}')
		# 	attribute_list = make_atribute_list(source_name)
		# 	validate_variant_attributes(variant_of,attribute_list)
		# 	if mod_reason in ['No Design Change','Change in Metal Colour']:
		# 		item_type = "Only Variant"
		# 		# bom_or_cad = workflow_state_maker(source_name)
		# 	elif mod_reason in ['Attribute Change']:
		# 		attribute_list = make_atribute_list(source_name)
		# 		validate_variant_attributes(variant_of,attribute_list)
		# 		item_type = "Only Variant"
		# 		# bom_or_cad = 'Check'
		# 	elif mod_reason == 'Change in Metal Touch':
		# 		item_type = "No Variant No Suffix"
		# 		# bom_or_cad = workflow_state_maker(source_name)
		# 	else:
		# 		attribute_list = make_atribute_list(source_name)
		# 		validate_variant_attributes(variant_of,attribute_list)
		# 		# item_type = "Suffix Of Variant"
		# 		item_type = "Template and Variant"
		# 		# bom_or_cad = 'CAD'

			bom_or_cad = 'Check'
			# if frappe.db.get_value("Item",design_id,"Item_group") == 'Design DNU':
			# 	item_type = "Only Variant"
	elif design_type == 'Sketch Design':
		# variant_of = frappe.db.get_value("Item",design_id,"variant_of")
		# attribute_list = make_atribute_list(source_name)
		# validate_variant_attributes(variant_of,attribute_list)
		# item_type = "No Variant No Suffix"
		item_type = "Only Variant"
		bom_or_cad = 'CAD'
		# bom_or_cad = 'Check'
	elif design_type == 'As Per Design Type':
		item_type = "No Variant No Suffix"
		bom_or_cad = 'New BOM'
	elif is_finding_order:
		item_type = "No Variant No Suffix"
		bom_or_cad = 'New BOM'
	else:
		item_type = 'Template and Variant'
		bom_or_cad = 'CAD'

	doc = get_mapped_doc(
		"Order Form Detail",
		source_name,
		{
			"Order Form Detail": {
				"doctype": "Order" 
			}
		},target_doc, set_missing_values
	)

	for entity in parent_doc.get("service_type",[]):
		doc.append("service_type", {"service_type1": entity.service_type1})
	doc.parcel_place = parent_doc.parcel_place
	doc.company = parent_doc.company
	doc.form_remarks = parent_doc.remarks
	doc.india = parent_doc.india
	doc.usa = parent_doc.usa
	doc.india_states = parent_doc.india_states
	doc.item_type = item_type
	doc.bom_or_cad = bom_or_cad
	if design_type in ['New Design','Sketch Design']:
		doc.workflow_type = 'CAD'
	
	doc.save()
	if design_type == 'As Per Design Type' and item_type == "No Variant No Suffix" and bom_or_cad == 'New BOM':
		doc.submit()
		frappe.db.set_value("Order",doc.name,"workflow_state","Approved")
	return doc.name

def make_atribute_list(source_name):
	order_form_details = frappe.get_doc('Order Form Detail',source_name)
	all_variant_attribute = frappe.db.sql(
		f"""select item_attribute from `tabAttribute Value Item Attribute Detail` 
		where parent = '{order_form_details.subcategory}' and in_item_variant=1""",as_list=1
	)

	final_list = {}
	for i in all_variant_attribute:
		new_i = i[0].replace(' ','_').replace('/','').lower()
		final_list[i[0]] = order_form_details.get_value(new_i)
	return final_list
# def set_item_type(source_name):
# 	doc = frappe.get_doc('Order Form Detail',source_name)
	
# 	# all_variant_attribute = frappe.db.get_list('Attribute Value Item Attribute Detail',filters={'parent':doc.subcategory,'in_item_variant':1},fields=['item_attribute'])
# 	all_variant_attribute = frappe.db.sql(
# 		f"""select item_attribute from `tabAttribute Value Item Attribute Detail` where parent = '{doc.subcategory}' and in_item_variant=1""",as_dict=1
# 	)
# 	# suffix_attribute = frappe.db.get_list('Attribute Value Item Attribute Detail',filters={'parent':doc.subcategory,'suffix':1},fields=['item_attribute'])
# 	suffix_attribute = frappe.db.sql(
# 		f"""select item_attribute from `tabAttribute Value Item Attribute Detail` where parent = '{doc.subcategory}' and suffix=1""",as_dict=1
# 	)
# 	# print(suffix_attribute)

# 	all_attribute_list = []
# 	for variant_attribut in all_variant_attribute:
# 		item_attribute = variant_attribut['item_attribute'].lower().replace(' ','_')
# 		all_attribute_list.append(item_attribute)
	
# 	# bom_detail = frappe.db.get_value('BOM',{'is_active':1,'item':doc.design_id},all_attribute_list,as_dict=1)
# 	bom = frappe.db.get_value('Item',doc.design_id,'master_bom')
# 	if bom==None:
# 		frappe.throw(f'BOM is not available for {doc.design_id}')
# 	bom_detail = frappe.db.get_value('BOM',doc.bom,all_attribute_list,as_dict=1)
	
# 	all_item_type = []
# 	for attribute in suffix_attribute:
# 		item_attribute = attribute['item_attribute'].lower().replace(' ','_')
# 		if bom_detail[item_attribute] != frappe.db.get_value('Order Form Detail',source_name,item_attribute):
# 			item_type = 'Suffix Of Variant'
# 		else:
# 			item_type = 'Only Variant'
# 		all_item_type.append(item_type)
	
# 	if 'Suffix Of Variant' in all_item_type:
# 		item_type = 'Suffix Of Variant'
# 	else:
# 		item_type = 'Only Variant'

# 	return item_type


def workflow_state_maker(source_name):

	doc = frappe.get_doc('Order Form Detail',source_name)
	all_variant_attribute = frappe.db.sql(
		f"""select item_attribute from `tabAttribute Value Item Attribute Detail` where parent = '{doc.subcategory}' and in_item_variant=1""",as_dict=1
	)
	all_attribute_list = []
	for variant_attribut in all_variant_attribute:
		item_attribute = variant_attribut['item_attribute'].lower().replace(' ','_').replace('/','')
		all_attribute_list.append(item_attribute)
	
	bom_detail = frappe.db.get_value('BOM',doc.bom,all_attribute_list,as_dict=1)
	cad_attribute_list = frappe.db.sql(
		f"""select item_attribute from `tabAttribute Value Item Attribute Detail` where parent = '{doc.subcategory}' and in_cad_flow=1""",as_dict=1
	)
	
	all_bom_or_cad = []
	for i in cad_attribute_list:
		item_attribute = i['item_attribute'].lower().replace(' ','_')
		if str(bom_detail[item_attribute]).replace(".0","") != str(frappe.db.get_value('Order Form Detail',source_name,item_attribute)):
			bom_or_cad = 'CAD'
		else:
			bom_or_cad = 'New BOM'
		all_bom_or_cad.append(bom_or_cad)
	
	
	if 'CAD' in all_bom_or_cad:
		bom_or_cad = 'CAD'
	else:
		bom_or_cad = 'New BOM'

	return bom_or_cad


@frappe.whitelist()
def make_order_form(source_name,target_doc=None):
	if isinstance(target_doc, str):
		target_doc = json.loads(target_doc)
	if not target_doc:
		target_doc = frappe.new_doc("Order Form")
	else:
		target_doc = frappe.get_doc(target_doc)

	titan_order_form = frappe.db.get_value("Titan Order Form", source_name, "*")

	target_doc.append("order_details", {
		"category": titan_order_form.get("item_category"),
		"design_by": "Our Design",
		"design_type": "Mod - Old Stylebio & Tag No",
		"design_id": titan_order_form.get("design_code"),
		"titan_code": titan_order_form.get("titan_code"),
		"subcategory": titan_order_form.get("item_subcategory"),
		"bom": titan_order_form.get("bom"),
		"serial_no_bom": titan_order_form.get("serial_no_bom"),
		"tag_no": titan_order_form.get("serial_no"),
		"metal_type": titan_order_form.get("metal_type"),
		"metal_touch": titan_order_form.get("metal_touch"),
		"metal_purity": titan_order_form.get("metal_purity"),
		"metal_colour": titan_order_form.get("metal_colour"),
		"product_size": titan_order_form.get("product_size"),
		"setting_type": titan_order_form.get("setting_type"),
		"delivery_date": titan_order_form.get("delivery_date"),
		"enamal": titan_order_form.get("enamal"),
		"rhodium": titan_order_form.get("rhodium"),
	})


	return target_doc


# @frappe.whitelist()
# def get_metal_color_varinat(metal_colour,design_id):
# 	variant_of = frappe.db.get_value('Item',{'name':design_id},'variant_of')
# 	all_variant = frappe.db.get_all('Item',filters={'variant_of':variant_of},pluck='name')
# 	for i in all_variant:
# 		print(frappe.db.sql(f"""SELECT attribute_value  from `tabItem Variant Attribute` tiva where parent ='{i}' and `attribute`= 'Metal Colour'""",as_dict=1))
# 		if metal_colour == frappe.db.sql(f"""SELECT attribute_value  from `tabItem Variant Attribute` tiva where parent ='{i}' and `attribute`= 'Metal Colour'""",as_dict=1)[0]['attribute_value']:
# 			return i
		
# def check_varinat(source_name):
# 	row = frappe.get_doc('Order Form Detail',source_name)
# 	# variant_of,item_subcategory = frappe.db.get_value('Item',{'name':row.design_id},['variant_of','item_subcategory'])
# 	# all_variant = frappe.db.get_all('Item',filters={'variant_of':variant_of},pluck='name')

# 	item_subcategory = row.subcategory
# 	all_attribute = frappe.db.sql(f"""select item_attribute  from `tabAttribute Value Item Attribute Detail` taviad  where parent = '{item_subcategory}' and in_item_variant =1""",as_dict=1)
# 	a = []
# 	for i in all_variant:
# 		variant_attrbute_value_list = []
# 		for j in all_attribute:
# 			attribute_value = frappe.db.sql(f"""SELECT attribute_value  from `tabItem Variant Attribute` tiva where parent ='{i}' and `attribute`= '{j['item_attribute']}'""",as_dict=1)
# 			if attribute_value:
# 				variant_attrbute_value_list.append(attribute_value[0]['attribute_value'])
# 			else:
# 				variant_attrbute_value_list.append('')
# 		a.append([i,variant_attrbute_value_list])
	
# 	current_data = []
# 	for k in all_attribute:
# 		av = frappe.db.get_value('Order Form Detail',row.name,k['item_attribute'].replace(' ','_').lower())
# 		if type(av) == float:
# 			av = "{:g}".format(av)
# 		current_data.append(av)


# 	is_present = any(current_data == sublist[1] for sublist in a)
	
# 	if is_present:
# 		matching_sublist = next(sublist for sublist in a if current_data == sublist[1])
# 		first_element = matching_sublist[0]
# 		return first_element
	# else:
	# 	return row.design_id

# @frappe.whitelist()
# def get_bom_details(design_id,doc):
# 	doc = json.loads(doc)
# 	if doc['is_finding_order']:
# 		master_bom = frappe.db.get_value("BOM",{"bom_type":"Template","item":design_id},"name",order_by="creation DESC")
# 		frappe.throw(f"{master_bom}//{doc['is_finding_order']}")
	
# 	item_subcategory = frappe.db.get_value("Item",design_id,"item_subcategory")

# 	fg_bom = frappe.db.get_value("BOM",{"bom_type":"Finished Goods","item":design_id},"name",order_by="creation DESC")
# 	master_bom = fg_bom
# 	if not fg_bom:
# 		temp_bom = frappe.db.get_value("Item",design_id,"master_bom")
# 		master_bom = temp_bom
# 	if not master_bom:
# 		frappe.throw(f"Master BOM for Item <b>{get_link_to_form('Item',design_id)}</b> is not set")
# 	all_item_attributes = []

# 	for i in frappe.get_doc("Attribute Value",item_subcategory).item_attributes:
# 		all_item_attributes.append(i.item_attribute.replace(' ','_').replace('/','').lower())
	
# 	with_value = frappe.db.get_value("BOM",master_bom,all_item_attributes,as_dict=1)
# 	with_value['master_bom'] = master_bom
# 	return with_value
@frappe.whitelist()
def get_bom_details(design_id, doc):
	doc = json.loads(doc)

	if doc['is_finding_order']:
		master_bom = frappe.db.get_value("BOM", {"bom_type": "Template", "item": design_id}, "name", order_by="creation DESC")
		frappe.throw(f"{master_bom}//{doc['is_finding_order']}")

	item_subcategory = frappe.db.get_value("Item", design_id, "item_subcategory")

	fg_bom = frappe.db.get_value("BOM", {"bom_type": "Finished Goods", "item": design_id}, "name", order_by="creation DESC")
	master_bom = fg_bom or frappe.db.get_value("Item", design_id, "master_bom")

	if not master_bom:
		frappe.throw(f"Master BOM for Item <b>{get_link_to_form('Item', design_id)}</b> is not set")

	# Prepare attribute keys from subcategory
	attribute_keys = []
	attribute_pairs = []

	for attr in frappe.get_doc("Attribute Value", item_subcategory).item_attributes:
		formatted = attr.item_attribute.replace(' ', '_').replace('/', '').lower()
		attribute_keys.append(formatted)
		attribute_pairs.append((attr.item_attribute, formatted))

	# Get values from Item Variant Attribute table
	variant_attributes = frappe.db.get_all("Item Variant Attribute", filters={"parent": design_id}, fields=["attribute", "attribute_value"])
	variant_map = {d.attribute.replace(' ', '_').replace('/', '').lower(): d.attribute_value for d in variant_attributes}

	# Get fallback values from BOM
	bom_values = frappe.db.get_value("BOM", master_bom, attribute_keys, as_dict=1)

	with_value = {}
	for original_name, formatted_key in attribute_pairs:
		with_value[formatted_key] = variant_map.get(formatted_key) or bom_values.get(formatted_key)

	with_value['master_bom'] = master_bom
	return with_value

def validate_variant_attributes(variant_of,attribute_list):
	args = attribute_list
	variant = get_variant(variant_of, args)
	if variant:
		frappe.throw(
			_("Item variant <b>{0}</b> exists with same attributes").format(get_link_to_form("Item",variant)), ItemVariantExistsError
		)

@frappe.whitelist()
def get_metal_purity(metal_type,metal_touch,customer):
	metal_purity = frappe.db.sql(f"""select metal_purity from `tabMetal Criteria` where parent = '{customer}' and metal_type = '{metal_type}' and metal_touch = '{metal_touch}'""",as_dict=1)
	return metal_purity


@frappe.whitelist()
def get_sketh_details(design_id):
	
	db_data = frappe.db.sql(f"select name,attribute, attribute_value from `tabItem Variant Attribute` where parent = '{design_id}'",as_dict=1)
	final_data = {}
	sketch_order_id = frappe.db.get_value("Item",design_id,"custom_sketch_order_id")
	final_data['item_category'] = frappe.db.get_value("Item",design_id,"item_category")
	final_data['item_subcategory'] = frappe.db.get_value("Item",design_id,"item_subcategory")
	final_data['setting_type'] = frappe.db.get_value("Item",design_id,"setting_type")
	final_data['sub_setting_type1'] = frappe.db.get_value("Sketch Order",sketch_order_id,"sub_setting_type1")
	final_data['sub_setting_type2'] = frappe.db.get_value("Sketch Order",sketch_order_id,"sub_setting_type2")
	final_data['qty'] = frappe.db.get_value("Sketch Order",sketch_order_id,"qty")
	final_data['metal_type'] = frappe.db.get_value("Sketch Order",sketch_order_id,"metal_type")
	final_data['metal_touch'] = frappe.db.get_value("Sketch Order",sketch_order_id,"metal_touch")
	final_data['metal_colour'] = frappe.db.get_value("Sketch Order",sketch_order_id,"metal_colour")
	final_data['metal_target'] = frappe.db.get_value("Item",design_id,"approx_gold")
	final_data['diamond_target'] = frappe.db.get_value("Item",design_id,"approx_diamond")
	final_data['product_size'] = frappe.db.get_value("Sketch Order",sketch_order_id,"product_size")
	final_data['sizer_type'] = frappe.db.get_value("Sketch Order",sketch_order_id,"sizer_type")
	final_data['length'] = frappe.db.get_value("Sketch Order",sketch_order_id,"length")
	final_data['width'] = frappe.db.get_value("Sketch Order",sketch_order_id,"width")
	final_data['height'] = frappe.db.get_value("Sketch Order",sketch_order_id,"height")
	for i in db_data:
		if i.attribute_value in [None,'']:
			continue
		final_data[i.attribute.lower().replace(' ','_')]=i.attribute_value
	return final_data

@frappe.whitelist()
def get_item_details(item_code):
	item_code = frappe.db.sql(f"""select attribute,attribute_value from `tabItem Variant Attribute` where parent = '{item_code}'""")
	return item_code

@frappe.whitelist()
def item_attribute_query(doctype, txt, searchfield, start, page_len, filters):
	args = {
		'item_attribute': filters.get("item_attribute"),
		"txt": "%{0}%".format(txt),
	}
	condition = ''
	if filters.get("customer_code"):		
		if filters.get("item_attribute") == "Metal Touch":
			args["customer_code"] = filters.get("customer_code")
			condition += "and attribute_value in (select metal_touch from `tabMetal Criteria`  where parent = %(customer_code)s)"

	item_attribute = frappe.db.sql(f"""select attribute_value
			from `tabItem Attribute Value`
				where parent = %(item_attribute)s 
				and attribute_value like %(txt)s {condition}
			""",args)
	return item_attribute if item_attribute else []

@frappe.whitelist()
def get_customer_orderType(customer_code):
	order_type = frappe.db.sql(
		f""" select order_type from `tabOrder Type` where parent= '{customer_code}' """, as_dict=1
	)

	return order_type

@frappe.whitelist()
def get_customer_order_form(source_name, target_doc=None):
	if isinstance(target_doc, str):
		target_doc = json.loads(target_doc)
	target_doc = frappe.new_doc("Order Form") if not target_doc else frappe.get_doc(target_doc)
 
	customer_code = ""
	if source_name:
		customer_code = frappe.db.get_value("Customer Order Form", source_name, "customer_code")
		target_doc.customer_code = customer_code  # Set in Order Form (Parent)

	if source_name:
		customer_order_form = frappe.db.sql(f"""SELECT * FROM `tabCustomer Order Form Detail` 
							WHERE parent = '{source_name}' AND docstatus = 1""", as_dict=1)
	if not customer_order_form:
		frappe.throw(_("Please submit the Customer Order Form"))
		return target_doc

	for i in customer_order_form:
		item, order_id, item_bom = i.get("design_code"), i.get("order_id"), i.get("design_code_bom")
		order_data = frappe.db.sql(f"SELECT * FROM `tabOrder` WHERE name = '{order_id}'", as_dict=1)
		customer_design_code = frappe.db.sql(f"SELECT * FROM `tabBOM` WHERE item = '{item}' AND name = '{i.get('design_code_bom')}'", as_dict=1)
		item_serial = frappe.db.get_value("Serial No", {'item_code': item}, 'name')
		
		data_source = order_data if order_data else customer_design_code
		
		product_code = ''
		# frappe.throw(f"111{item} {i.get('design_code_bom')} {customer_design_code}")
		if i.get("digit14_code"):
			product_code = i.get("digit14_code")
		elif i.get("digit18_code"):
			product_code = i.get("digit18_code")
		elif i.get("digit15_code"):
			product_code = i.get("digit15_code")
		elif i.get("sku_code"):
			product_code = i.get("sku_code")
		# frappe.throw(f"{data_source}")
		if data_source:
			for j in data_source:
				target_doc.append("order_details", {
					"delivery_date": target_doc.delivery_date,
					"design_by": "Our Design",
					"design_type": "As Per Design Type",
					"qty": i.get('no_of_pcs'),
					"design_id": j.get("item", item),
					"bom": j.get("new_bom", i.get('design_code_bom')),
					"tag_no": item_serial or j.get('tag_no'),
					"diamond_quality": i.get("diamond_quality"),
					"customer_order_form": i.get("parent"),
					"category": i.get("category"),
					"subcategory": i.get("subcategory"),
					"setting_type": i.get("setting_type"),
					"product_code": product_code if product_code else '',
					"theme_code": i.get("theme_code"),
					"metal_type": i.get("metal_type"),
					"metal_touch": i.get("metal_touch"),
					"metal_colour": i.get("metal_colour"),
					"metal_target": i.get("metal_target"),
					"diamond_target": i.get("diamond_target"),
					"feature": i.get("feature"),
					"product_size": i.get("product_size"),
					"rhodium": i.get("rhodium"),
					"enamal": j.get("enamal"),
					"sub_setting_type1": j.get("sub_setting_type1"),
					"sub_setting_type2": j.get("sub_setting_type2"),
					"sizer_type": j.get("sizer_type"),
					"stone_changeable": j.get("stone_changeable"),
					"detachable": j.get("detachable"),
					"lock_type": j.get("lock_type"),
					"capganthan": j.get("capganthan"),
					"charm": j.get("charm"),
					"back_chain": j.get("back_chain"),
					"back_chain_size": j.get("back_chain_size"),
					"back_belt": j.get("back_belt"),
					"back_belt_length": j.get("back_belt_length"),
					"black_beed_line": j.get("black_beed_line"),
					"back_side_size": j.get("back_side_size"),
					"back_belt_patti": j.get("back_belt_patti"),
					"two_in_one": j.get("two_in_one"),
					"number_of_ant": j.get("number_of_ant"),
					"distance_between_kadi_to_mugappu": j.get("distance_between_kadi_to_mugappu"),
					"space_between_mugappu": j.get("space_between_mugappu"),
					"chain_type": j.get("chain_type"),
					"customer_chain": j.get("customer_chain"),
					"nakshi_weght": j.get("nakshi_weght"),
					"diamond_type":"Natural"
				})
		else: 
			frappe.throw(f"{item} has master bom {item_bom}")
	return target_doc


def validate_item_variant(self):
	for i in self.order_details:
		if i.design_type == "Sketch Design" and i.design_id:
			custom_sketch_order_id = frappe.db.get_value("Item", i.design_id, "custom_sketch_order_id")
			if custom_sketch_order_id:
				# Get all variants where variant_of = i.design_id
				variants = frappe.get_all("Item",
					filters={"variant_of": i.design_id},
					fields=["name"]
				)
				# frappe.throw(f"{variants}")
				if variants:
					variant_names = ", ".join(item.name for item in variants)
					frappe.throw(f"""
						You already created a variant for this Design ID ({i.design_id}).<br><br>
						Items found: {variant_names}<br><br>
						You cannot create another variant using <b>Sketch Design</b>.<br>
						Please select <b>Design Type = 'Mod - Old Stylebio & Tag No'</b> and select the variant in Design ID.
					""")

def validate_rows_on_order_criteria(self):
    criteria = frappe.get_single("Order Criteria")
    for row in criteria.order_form_table:
        if row.disabled:
            continue


        if self.order_type == "Sales":
            allowed_rows = row.sales
            if self.total_rows > allowed_rows:
                frappe.throw(
                    f"Order Type is 'Sales'.\n"
                    f"Allowed rows as per Order Criteria: {allowed_rows}\n"
                    f"But you are adding {self.total_rows} rows."
                )

        elif self.order_type == "Stock Order":
            allowed_rows = row.stock_order
            if self.total_rows > allowed_rows:
                frappe.throw(
                    f"Order Type is 'Stock Order'.\n"
                    f"Allowed rows as per Order Criteria: {allowed_rows}\n"
                    f"But you are adding {self.total_rows} rows."
                )

        elif self.order_type == "Purchase":
            allowed_rows = row.purchase
            if self.total_rows > allowed_rows:
                frappe.throw(
                    f"Order Type is 'Purchase'.\n"
                    f"Allowed rows as per Order Criteria: {allowed_rows}\n"
                    f"But you are adding {self.total_rows} rows."
                )




def validate_design_id(self):
	for i in self.order_details:
		# If tagno exists, find all matching enabled Items by old_tag_no
		if i.tagno:
			matching_items = frappe.db.get_all(
				"Item",
				filters={"old_tag_no": i.tagno, "disabled": 0},
				fields=["name", "master_bom", "creation"],
				order_by="creation desc"
			)

			if matching_items:
				# Pick the latest enabled item
				item = matching_items[0]
				matched_design_id = item.name

				# Set design_id only if not manually overridden
				if not i.design_id or i.design_id == matched_design_id:
					i.design_id = matched_design_id
					if not i.bom:
						i.bom = item.master_bom

		# If design_id is set, fetch Item and set master_bom to bom
		if i.design_id:
			item_doc = frappe.get_doc("Item", i.design_id)
			if item_doc.master_bom and not i.bom:
				i.bom = item_doc.master_bom

		# Continue only if design_id and bom are now set
		if i.design_id and i.bom:
			is_manual_override = (i.design_type == "Mod - Old Stylebio & Tag No")

			# Skip if mod_reason is "Change In Metal Type"
			if i.mod_reason != "Change In Metal Type":
				bom_doc = frappe.get_doc("BOM", i.bom)

				# Set metal_type and metal_touch from metal_detail
				if bom_doc.metal_detail:
					if not is_manual_override or not i.metal_type:
						i.metal_type = bom_doc.metal_detail[0].metal_type or None
					if not is_manual_override or not i.metal_touch:
						i.metal_touch = bom_doc.metal_detail[0].metal_touch or None
				else:
					frappe.msgprint(f"No metal details found for BOM {i.bom}")

				# Set setting_type, category, subcategory
				if not is_manual_override or not i.setting_type:
					i.setting_type = bom_doc.setting_type or None
				if not is_manual_override or not i.category:
					i.category = bom_doc.item_category or None
				if not is_manual_override or not i.subcategory:
					i.subcategory = bom_doc.item_subcategory or None

				# Attribute mapping
				attr_map = {
					"metal_colour": "Metal Colour",
					"diamond_target": "Diamond Target",
					"stone_changeable": "Stone Changeable",
					"gemstone_type": "Gemstone Type",
					"chain_type": "Chain Type",
					"chain_length": "Chain Length",
					"feature": "Feature",
					"rhodium": "Rhodium",
					"enamal": "Enamal",
					"detachable": "Detachable",
					"capganthan": "Cap/Ganthan",
					"two_in_one": "Two in One",
					"product_size": "Product Size",
					"sizer_type": "Sizer Type",
					"lock_type": "Lock Type",
					"black_bead_line": "Black Bead Line",
					"charm": "Charm",
					"count_of_spiral_turns": "Count of Spiral Turns",
					"number_of_ant": "Number of Ant",
					"back_side_size": "Back Side Size",
					"space_between_mugappu": "Space between Mugappu",
					"distance_between_kadi_to_mugappu": "Distance Between Kadi To Mugappu",
					"back_belt": "Back Belt",
					"back_belt_length": "Back Belt Length",
				}

				# Clear all mapped fields if not manual override
				if not is_manual_override:
					for fieldname in attr_map.keys():
						setattr(i, fieldname, None)

				# Set values from attributes if not manually overridden
				for attr in item_doc.attributes:
					for fieldname, attrname in attr_map.items():
						if attr.attribute == attrname:
							if not is_manual_override or not getattr(i, fieldname):
								setattr(i, fieldname, attr.attribute_value)

		# Final mandatory field validation
		if  i.design_type == "New Design":
			missing = []
			if not i.category:
				missing.append("Category")
			if not i.subcategory:
				missing.append("Subcategory")
			if not i.metal_type:
				missing.append("Metal Type")
			if not i.diamond_target:
				missing.append("Diamond Target")
			if not i.setting_type:
				missing.append("Setting Type")
			if not i.metal_touch:
				missing.append("Metal Touch")
			if not i.metal_colour:
				missing.append("Metal Colour")
			if not i.metal_target:
				missing.append("Metal Target")

			if missing:
				frappe.throw(f"Row {i.idx}: Please fill the following fields for 'New Design' with Manual checked: {', '.join(missing)}")


def validate_is_mannual(self):
	if self.is_mannual:
		errors = []

		for row in self.order_details:
			missing_fields = []

			if not row.stylebio:
				missing_fields.append("'Style Bio'")
			if not row.status:
				missing_fields.append("'Status'")
			if not row.order_details_and_remarks:
				missing_fields.append("'Order Details and Remark'")

			# Enhanced: handle multiple items with same tagno, pick non-disabled one
			if row.tagno:
				matching_items = frappe.db.get_all(
					"Item",
					filters={"old_tag_no": row.tagno},
					fields=["name", "master_bom", "disabled"],
					order_by="creation desc"
				)

				selected_item = next((item for item in matching_items if not item.disabled), None)

				if selected_item:
					if not row.design_id:
						row.design_id = selected_item.name
					if not row.bom:
						row.bom = selected_item.master_bom

					if selected_item.master_bom:
						diamond_type = frappe.db.get_value(
							"BOM Diamond Detail", 
							{"parent": selected_item.master_bom}, 
							"diamond_type"
						)
						if diamond_type:
							row.diamond_type = diamond_type

			# If workflow_state == "Approved", design_type is mandatory for non-finding
			if (
				not row.is_finding_order
				and not row.design_type
				and self.workflow_state == "Approved"
			):
				missing_fields.append("'Design Type' (required in 'Creating Item & BOM')")

			if missing_fields:
				errors.append(f"Row {row.idx} is missing: {', '.join(missing_fields)}")

		if errors:
			frappe.throw("<br>".join(errors))

		# Enforce all status as 'Done' if workflow_state is Approved
		if self.workflow_state == "Approved":
			for row in self.order_details:
				if row.status != "Done":
					frappe.throw(f"Row {row.idx}: Status must be 'Done' before you approve. Please update it.")

	else:
		# is_mannual is unchecked validate design_type for non-finding orders
		missing_design_type_rows = []
		for row in self.order_details:
			if not row.is_finding_order and not row.design_type:
				missing_design_type_rows.append(
					f"Row {row.idx}: Design Type is mandatory when 'Is Finding Order' is unchecked and 'Is Mannual' is also unchecked."
				)

		if missing_design_type_rows:
			frappe.throw("<br>".join(missing_design_type_rows))



def set_data(self):
	if self.order_details:
		for i in self.order_details:
			if i.design_type in ['As Per Design Type','Mod - Old Stylebio & Tag No'] and i.design_id:
				try:
					design_id = i.design_id
					item_subcategory = frappe.db.get_value("Item", design_id, "item_subcategory")
					master_bom = i.bom

					# Prepare a list to hold the item attribute names formatted as per your requirements
					all_item_attributes = []
					
					# Retrieve all item attributes for the given item subcategory
					for item_attr in frappe.get_doc("Attribute Value", item_subcategory).item_attributes:
						# Format the item attribute names by replacing spaces with underscores, removing '/', and converting to lower case
						formatted_attr = item_attr.item_attribute.replace(' ', '_').replace('/', '').lower()
						all_item_attributes.append(formatted_attr)
					
					# Retrieve the values for the specified attributes from the BOM
					attribute_values = frappe.db.get_value("BOM", master_bom, all_item_attributes, as_dict=1)
					
					# Dynamically set the attributes on self with the retrieved values
					for key, value in attribute_values.items():
						if str(key) == "item_category":
							key = "category"
						if str(key) == "item_subcategory":
							key = "subcategory"
						a = getattr(i, key, value)
						if a:
							continue
						# frappe.throw(f"{a}")
						else:
							setattr(i, key, value)
						# Prepare a list to hold the item attribute names formatted as per your requirements
						all_item_attributes = []
						
						# Retrieve all item attributes for the given item subcategory
						for item_attr in frappe.get_doc("Attribute Value", item_subcategory).item_attributes:
							# Format the item attribute names by replacing spaces with underscores, removing '/', and converting to lower case
							formatted_attr = item_attr.item_attribute.replace(' ', '_').replace('/', '').lower()
							
							all_item_attributes.append(formatted_attr)
						
						# Retrieve the values for the specified attributes from the BOM
						attribute_values = frappe.db.get_value("BOM", master_bom, all_item_attributes, as_dict=1)
						# Dynamically set the attributes on self with the retrieved values
						for key, value in attribute_values.items():
							if str(key) == "item_category":
								key = "category"
							if str(key) == "item_subcategory":
								key = "subcategory"
							a = getattr(i, key, value)
							if a:
								continue
							else:
								setattr(i, key, value)
				except:
					frappe.throw(f"Row {i.idx} has Issue.Check BOM first.")


def create_po(self):
	qty = 0
	po_doc = frappe.new_doc("Purchase Order")
	po_doc.supplier = self.supplier
	po_doc.transaction_date = self.delivery_date
	po_doc.company = self.company
	po_doc.branch = self.branch
	po_doc.project = self.project
	po_doc.purchase_type = 'Subcontracting'
	# po_doc.custom_form = "Order Form"
	# po_doc.custom_form_id = self.name
	po_doc.schedule_date = self.delivery_date

	po_item_log = po_doc.append("items", {})
	if self.purchase_type == 'Design':
		po_item_log.item_code = "Design Expness"
	elif self.purchase_type == 'RPT':
		po_item_log.item_code = "RPT Expness"
	elif self.purchase_type == 'Model':
		total_weight = 0
		# qty_18 = 0
		# qty_22 = 0
		item_code = ''
		for i in self.order_details:
			# total_weight += i.total_weight
			if i.metal_touch == '18KT':
				item_code = "Semi Finish Goods 18KT"
			if i.metal_touch == '22KT':
				item_code = "Semi Finish Goods 22KT"
		po_item_log.item_code = item_code
		# po_item_log.total_weight = total_weight
		# po_item_log.weight_per_unit = total_weight/qty_22
	elif self.purchase_type == 'Mould':
		po_item_log.item_code = "Mould Expness"
	
	if self.purchase_type in ['Model']:
		qty_18 = 0
		qty_22 = 0
		for i in self.order_details:
			if i.metal_touch == '18KT':
				qty_18 += i.qty
			if i.metal_touch == '22KT':
				qty_22 += i.qty
		if qty_18:
			qty = qty_18
		else:
			qty = qty_22
	else:
		for i in self.order_details:
			qty+=i.qty
	
	po_item_log.qty = qty
	po_item_log.schedule_date = self.delivery_date
	po_item_log.schedule_date = self.delivery_date
	po_item_log.qty = len(self.order_details)

	# po_doc.set("payment_schedule", [])
	# po_trn_log = po_doc.append("payment_schedule", {})
	# po_trn_log.due_date = self.delivery_date
	# po_trn_log.invoice_portion = 100.0

	
	po_doc.save()
	po_name = po_doc.name
	frappe.db.set_value("Purchase Order",po_name,"custom_form","Order Form")
	frappe.db.set_value("Purchase Order",po_name,"custom_form_id",self.name)
	msg = _("The following {0} is created: {1}").format(
			frappe.bold(_("Purchase Order")), "<br>" + get_link_to_form("Purchase Order", po_name)
		)
	
	frappe.msgprint(msg)

@frappe.whitelist()
def make_from_pre_order_form(source_name, target_doc=None):
	if isinstance(target_doc, str):
		target_doc = json.loads(target_doc)
	target_doc = frappe.new_doc("Order Form") if not target_doc else frappe.get_doc(target_doc)

	if source_name:
				customer_order_form = frappe.db.sql(f"""SELECT * FROM `tabPre Order Form Details` 
							WHERE parent = '{source_name}'""", as_dict=1)
		# customer_order_form = frappe.db.sql(f"""SELECT * FROM `tabPre Order Form Details` 
		# 					WHERE parent = '{source_name}' AND docstatus = 1""", as_dict=1)

	# parent = frappe.db.get_value("Pre Order Form Details",source_name)
	target_doc.customer_code = frappe.db.get_value("Pre Order Form",source_name,"customer_code")
	target_doc.order_date = frappe.db.get_value("Pre Order Form",source_name,"order_date")
	target_doc.salesman_name = frappe.db.get_value("Pre Order Form",source_name,"sales_person")
	target_doc.diamond_quality = frappe.db.get_value("Pre Order Form",source_name,"diamond_quality")
	target_doc.branch = frappe.db.get_value("Pre Order Form",source_name,"branch")
	target_doc.order_type = frappe.db.get_value("Pre Order Form",source_name,"order_type")
	target_doc.due_days = frappe.db.get_value("Pre Order Form",source_name,"due_days")
	target_doc.po_no = frappe.db.get_value("Pre Order Form",source_name,"po_no")
	target_doc.delivery_date = frappe.db.get_value("Pre Order Form",source_name,"delivery_date")
	target_doc.pre_order_form = source_name
	# target_doc.branch = frappe.db.get_value("Pre Order Form",source_name,"customer")
	service_types = frappe.db.get_values("Service Type 2", {"parent": source_name},"service_type1")
	for service_type in service_types:
		target_doc.append("service_type",{"service_type1": service_type[0]})
	
	shipping_territories = frappe.db.get_values("Territory Multi Select", {"parent": source_name},"territory")
	for shipping_territory in shipping_territories:
		target_doc.append("parcel_place",{"territory": shipping_territory[0]})

	for i in customer_order_form:
		# frappe.throw(f"{i}")
		# item, order_id = i.get("design_code"), i.get("order_id")
		# order_data = frappe.db.sql(f"SELECT * FROM `tabOrder` WHERE name = '{order_id}'", as_dict=1)
		# customer_design_code = frappe.db.sql(f"SELECT * FROM `tabBOM` WHERE item = '{item}' AND name = '{i.get('design_code_bom')}'", as_dict=1)
		# item_serial = frappe.db.get_value("Serial No", {'item_code': item}, 'name')

		# data_source = order_data if order_data else customer_design_code
		# if data_source:
		# 	for j in data_source:
		if i.status == 'Done':
			design_id = i.item_variant
			item_subcategory = frappe.db.get_value("Item", design_id, "item_subcategory")
			master_bom = i.bom

			extra_fields = {}

			if item_subcategory and master_bom:
				all_item_attributes = []
				for item_attr in frappe.get_doc("Attribute Value", item_subcategory).item_attributes:
					formatted_attr = item_attr.item_attribute.replace(' ', '_').replace('/', '').lower()
					all_item_attributes.append((item_attr.item_attribute, formatted_attr))

				# Build dict from Item Variant Attributes
				variant_attributes = frappe.db.get_all("Item Variant Attribute",
					filters={"parent": design_id},
					fields=["attribute", "attribute_value"]
				)
				variant_attr_map = {d.attribute.replace(' ', '_').replace('/', '').lower(): d.attribute_value for d in variant_attributes}

				# Fetch fallback values from BOM
				attribute_values = frappe.db.get_value("BOM", master_bom, [f[1] for f in all_item_attributes], as_dict=1)

				for original_name, formatted_name in all_item_attributes:
					value = variant_attr_map.get(formatted_name) or attribute_values.get(formatted_name)
					if value:
						fieldname = "category" if formatted_name == "item_category" else \
									"subcategory" if formatted_name == "item_subcategory" else formatted_name
						extra_fields[fieldname] = value

			target_doc.append("order_details", {
				"design_by":i.design_by,
				"design_type":i.design_type,
				"order_type":i.order_type,
				"delivery_date":frappe.db.get_value("Pre Order Form",source_name,"delivery_date"),
				"diamond_quality":frappe.db.get_value("Pre Order Form",source_name,"diamond_quality"),
				"design_id": i.item_variant,
				"mod_reason":i.mod_reason,
				"bom":i.bom,
				"category":i.new_category,
				"subcategory":i.new_sub_category,
				"metal_target":i.gold_target,
				"diamond_target":i.diamond_target,
				"setting_type":i.bom_setting_type,
				"pre_order_form_details":i.name,
				"diamond_type":"Natural",
				"jewelex_batch_no":i.bulk_order_no,
				"design_image_1":i.design_image,
				**({"metal_touch": i.metal_touch} if i.design_type == "New Design" else {}),
				**({"metal_colour": i.metal_color} if i.design_type == "New Design" else {}),
				**extra_fields
			})
			
	return target_doc


# customer order form
# for customer order form 
@frappe.whitelist()
def gc_export_to_excel(order_form, doc):
	order_form_doc = frappe.get_doc('Order Form', order_form)
	doc = json.loads(doc)  
	order_date_str = getdate(order_form_doc.order_date).strftime("%Y-%m-%d")
	
	file_name = f"GC_Format_{order_date_str}.xlsx" 

	workbook = openpyxl.Workbook()
	sheet = workbook.active
	sheet.title = 'GC Format'
	headers = [
		# 'Date','Customer','Customer Name','Company',
		'BS_ID',
		'FILE_ID',
		'FILE_NAME',
		'UNIQUE_ID',
		'M_VENDOR_REF_CODE',
		'M_VENDOR',
		'M_BRAND',
		'M_CFA_CODE',
		'M_STONE_CODE',
		'M_QUANTITY',
		'M_TOTAL_WEIGHT','KARAT',
		'CATEGORY',
		'WTSETTING_TYPE_1',
		'SETTING_TYPE_2',
		'SETTING_TYPE_3'
		]
	
	sheet.append(headers)

	# Store all rows in a list before writing to the sheet
	rows_data = []
	design_item = []
	
	for row in doc.get('order_details', []):
		if row.get('design_id') in design_item:
			continue  

		else:
			if row.get('design_id'):
				bom_list = ''
				if row.get('tag_no'): 
					# frappe.throw(f"{row.get('tag_no')}")
					bom_list = frappe.db.get_list("BOM", filters={'item': row["design_id"], 'bom_type': 'Finish Goods'}, fields=['*'])
				else:  
					bom_list = frappe.db.get_list("BOM", filters={'item': row["design_id"], 'name': row['bom']}, fields=['*'])

				if bom_list:
					bom_diamond = frappe.db.get_all("BOM Diamond Detail", 
									filters={'parent': bom_list[0].get('name')}, fields=['*'])
					max_rows = len(bom_diamond) or 1
					for i in range(max_rows):
						diamond = bom_diamond[i] if i < len(bom_diamond) else {}

					if row.get('uomset_of') == "SET":
						diamond_pcs = []
						set_item =  frappe.db.get_all("Set Item Table",filters={'parent': row.get('design_id')},fields=['item_code'])
						item_sub = ''
						metal_weight = 0
						if set_item:
							for item in set_item:
								set_item_bom =  frappe.db.get_value("Item",item.get('item_code'),'master_bom')
								item_sub =  frappe.db.get_value("Item",item.get('item_code'),'item_subcategory')
								metal_weight =  frappe.db.get_value("BOM",set_item_bom ,'total_metal_weight')
								
								diamond_no =  frappe.db.get_all("BOM Diamond Detail",filters={'parent': set_item_bom},fields=['diamond_sieve_size','pcs','sub_setting_type','diamond_type','size_in_mm','stone_shape'])
								design_item.append(item.get('item_code'))
								for pcs in diamond_no:
									diamond_pcs.append(pcs)
						design_item_code = frappe.db.get_all("Customer Category Detail",
											filters={
												'parent': order_form_doc.customer_code,
												'gk_sub_category': row.get('subcategory')+ " + " + item_sub
											},
											fields=['article'])
						code = design_item_code[0].get('article','') if design_item_code else ''
						# frappe.throw(str(design_item_code))
						diamond_count =  frappe.db.get_all("BOM Diamond Detail",filters={'parent': row.get('bom')},fields=['diamond_sieve_size','pcs','sub_setting_type','diamond_type','size_in_mm','stone_shape'])
						for d in diamond_count:
							found = False

							for existing in diamond_pcs:
								if existing.get('diamond_sieve_size') == d.get('diamond_sieve_size'):
									existing['pcs'] = existing.get('pcs', 0) + d.get('pcs', 0)
									found = True
									break

							if not found:
								diamond_pcs.append(d)
							
						for diamond in diamond_pcs:
							stone_code = frappe.db.sql("""
							SELECT rm.customer_code
							FROM `tabCustomer RM Code Detail` rm
							LEFT JOIN `tabCustomer RM Code` rmc ON rm.parent = rmc.name
							WHERE 
								IFNULL(rm.stone_shape, '') = IFNULL(%s, '')
								AND rmc.customer = %s
								AND IFNULL(rm.diamond_type, '') = IFNULL(%s, '')
								AND IFNULL(rm.diamond_quality, '') = IFNULL(%s, '')
								AND CAST(IFNULL(rm.size_in_mm, 0) AS DECIMAL(10,3)) = CAST(%s AS DECIMAL(10,3))
									
							""",
							(
								diamond.get("stone_shape"),
								order_form_doc.customer_code,
								diamond.get("diamond_type"),
								order_form_doc.diamond_quality,
								diamond.get("size_in_mm"),
							),
							as_dict=1)
							# frappe.throw(str(stone_code[0].get('customer_code')))
							code_ctg = ['N&J','N&D','N&C','E&D']

							row_data = [
								'-',
								'-', 
								'-',
								'-',
								row.get('design_id','') + code,
								'GURU',
								'TANISHQ',
								'-',
								stone_code[0].get('customer_code'),
								f"{(diamond.get('pcs', 0))}",
								float(bom_list[0].get('total_metal_weight') or 0) + float(metal_weight or 0),
								row.get('metal_touch', ''),
								# row.get('category',''),
								'SET2' if code in code_ctg else 'SET1',
								diamond.get('sub_setting_type' or ''),
								'-',
								'-',
							]

							rows_data.append(row_data)
					else:	
						# customer_code_list = []
						# stone_value = frappe.db.sql("""
						# 	SELECT *
						# 	FROM `tabBOM Diamond Detail`
						# 	WHERE parent = %s
						# """, bom_list[0].get("name"), as_dict=True)

						# # frappe.throw(str(stone))
						# for stone in stone_value:
						# 	stone_code = frappe.db.sql("""
						# 		SELECT rm.customer_code
						# 		FROM `tabCustomer RM Code Detail` rm
						# 		LEFT JOIN `tabCustomer RM Code` rmc ON rm.parent = rmc.name
						# 		WHERE 
						# 			IFNULL(rm.stone_shape, '') = IFNULL(%s, '')
						# 			AND rmc.customer = %s
						# 			AND IFNULL(rm.diamond_type, '') = IFNULL(%s, '')
						# 			AND IFNULL(rm.diamond_quality, '') = IFNULL(%s, '')
						# 			AND CAST(IFNULL(rm.size_in_mm, 0) AS DECIMAL(10,3)) = CAST(%s AS DECIMAL(10,3))
									
						# 	""",
						# 	(
						# 		stone.get("stone_shape"),
						# 		order_form_doc.customer_code,
						# 		stone.get("diamond_type"),
						# 		order_form_doc.diamond_quality,
						# 		stone.get("size_in_mm"),
						# 	),
						# 	as_dict=1)
						# 	if stone_code:
						# 		for code_row in stone_code:
						# 			customer_code_list.append(code_row.customer_code)
						# frappe.throw(str(customer_code_list))
						stylebio =  frappe.db.get_value("Item",row.get('design_id'),'stylebio')
						multiplier = 2 if row.get('category', '') == "Bangles" else 1
						for diamond in bom_diamond:
							stone_code = frappe.db.sql("""
							SELECT rm.customer_code
							FROM `tabCustomer RM Code Detail` rm
							LEFT JOIN `tabCustomer RM Code` rmc ON rm.parent = rmc.name
							WHERE 
								IFNULL(rm.stone_shape, '') = IFNULL(%s, '')
								AND rmc.customer = %s
								AND IFNULL(rm.diamond_type, '') = IFNULL(%s, '')
								AND IFNULL(rm.diamond_quality, '') = IFNULL(%s, '')
								AND CAST(IFNULL(rm.size_in_mm, 0) AS DECIMAL(10,3)) = CAST(%s AS DECIMAL(10,3))
									
							""",
							(
								diamond.get("stone_shape"),
								order_form_doc.customer_code,
								diamond.get("diamond_type"),
								order_form_doc.diamond_quality,
								diamond.get("size_in_mm"),
							),
							as_dict=1)
							row_data = [
								'-',
								'-', 
								'-',
								'-',
								row.get('design_id',''),
								'GURU',
								'TANISHQ',
								'-',
								stone_code[0].get('customer_code'),
								f"{(diamond.get('pcs', 0)) * multiplier}",
								f"{(bom_list[0].get('total_metal_weight', 0)) * multiplier:0.3f}",
								row.get('metal_touch', ''),
								row.get('category',''),
								diamond.get('sub_setting_type' or ''),
								'-',
								'-',
							]

							rows_data.append(row_data)

	if rows_data:
		for row in rows_data:
			sheet.append(row)
	else:
		frappe.throw("GC Sheet Can Not Download")

	output = BytesIO()
	workbook.save(output)
	output.seek(0)

	file_doc = save_file(
		file_name,
		output.getvalue(),
		order_form_doc.doctype,
		order_form_doc.name,
		is_private=0
	)
	
	return file_doc.file_url



@frappe.whitelist()
def creation_export_to_excel(order_form, doc):
	order_form_doc = frappe.get_doc('Order Form', order_form)
	doc = json.loads(doc)  
	
	order_date_str = getdate(order_form_doc.order_date).strftime("%Y-%m-%d")
	file_name = f"Code_Creation_File_{order_date_str}.xlsx" 

	workbook = openpyxl.Workbook()
	sheet = workbook.active
	sheet.title = 'Code Creation File'
	headers = [
		"BS_ID",
		"FILE_ID",
		"FILE_NAME",
		"VENDOR_REF_CODE",
		"UNIQUE_ID",
		"COLLECTION_NAME",
		"NEED_STATE",
		"M_BS_CLUSTER",
		"M_BS_CATEGORY",
		"BS_VENDOR",
		"M_CFA_CODE",
		"M_MATERIALS",
		"M_KARAT",
		"M_SIZES",
		"FINDINGS",
		"DORI_BLACK_CHIAN",
		"SHAPE","M_UOM",
		"M_GENDER",
		"M_COLOR",
		"CHILD_1",
		"CHILD_2",
		"CHILD_3",
		"CHILD_4",
		"CHILD_5",
		"CHILD_6",
		"CHILD_7",
		"CHILD_1_WT",
		"CHILD_2_WT",
		"CHILD_3_WT",
		"CHILD_4_WT",
		"CHILD_5_WT",
		"CHILD_6_WT",
		"CHILD_7_WT",
		"M_TOTAL_WEIGHT",
		"CHILD_1_D.WT",
		"CHILD_2_D.WT",
		"M_TOTAL_D.WEIGHT",
		"M_DIA QLTY",
		"STONE COMBINATION",
		"M_LC_GRAM",
		"M_LOSS_PERCENTAGE",
		"STONE_GROUP"
	]

	sheet.append(headers) 
	
	rows_data = []
	design_item = []

	for row in doc.get('order_details', []):
		if row.get('design_id') in design_item:
			continue  
		
		else:
			if row['design_id']:
				finish_bom_list = frappe.db.get_list("BOM", filters={'item': row["design_id"], 'bom_type': 'Finish Goods'}, fields=['name'])
				
				finish_bom = ''
				if len(finish_bom_list) > 1:
					order = frappe.db.get_value("Order", 
						{'cad_order_form': order_form, 'item': row['design_id']},'name')	
					pmo = frappe.db.get_value("Parent Manufacturing Order",{'order_form_id': order}, 'name')
					snc = frappe.db.get_value("Serial Number Creator",{'parent_manufacturing_order': pmo}, 'name')
					fg_bom = frappe.db.get_value("BOM", {'custom_serial_number_creator': snc, 'item': row['design_id'], 'bom_type': 'Finish Goods'}, 'name')
					finish_bom = fg_bom		
				else:
					for fg in finish_bom_list:
						finish_bom = fg.get('name')
				
				final_bom = ''
				if finish_bom:
					final_bom = finish_bom
				else:
					final_bom = frappe.db.get_value("Item", {'name': row["design_id"],}, ['master_bom'])
				
				
				if final_bom:
					# item_bom = frappe.db.get_list("BOM", filters={'item': row["design_id"], 'name': final_bom}, fields=['*'])
					if row.get('uomset_of') == "SET":
						metal_touch = row.get('metal_touch', '')
						metal_touch = metal_touch.replace('KT', '').replace('kt', '').strip()
						gemstone = frappe.db.get_value('BOM',row.get('bom'),'total_gemstone_pcs')
						diamond = frappe.db.get_value('BOM',row.get('bom'),'total_diamond_pcs')
						order_date = frappe.utils.formatdate(order_form_doc.order_date, "dd.MM.yyyy")

						set_item =  frappe.db.get_all("Set Item Table",filters={'parent': row.get('design_id')},fields=['item_code'])
						item_sub = None
						item_sub_bom = None
						# if set_item:
						# 	item_sub =  frappe.db.get_value("Item",set_item[0].get('item_code'),'item_subcategory')
						# 	item_sub_bom =  frappe.db.get_value('Item',set_item[0].get('item_code'),'master_bom')
						# 	design_item.append(set_item[0].get('item_code'))
						item_sub = ''
						metal_weight = 0
						if set_item:
							for item in set_item:
								item_sub_bom =  frappe.db.get_value("Item",item.get('item_code'),'master_bom')
								item_sub =  frappe.db.get_value("Item",item.get('item_code'),'item_subcategory')
								metal_weight =  frappe.db.get_value("BOM",item_sub_bom ,'total_metal_weight')
								
								# diamond_no =  frappe.db.get_all("BOM Diamond Detail",filters={'parent': item_sub_bom},fields=['diamond_sieve_size','pcs','sub_setting_type','diamond_type','size_in_mm','stone_shape'])
								design_item.append(item.get('item_code'))
						child_1 =  frappe.db.get_all("Customer Category Detail",
									filters = {
										'parent':order_form_doc.customer_code,
										'gk_category':row.get('category')
									},fields=['customer_category'])
						child_2 = []
						if item_sub:
							child_2 =  frappe.db.get_all("Customer Category Detail",
										filters = {
											'parent':order_form_doc.customer_code,
											'gk_sub_category':item_sub
										},fields=['customer_category'])

						design_item_code = frappe.db.get_all("Customer Category Detail",
											filters={
												'parent': order_form_doc.customer_code,
												'gk_sub_category': row.get('subcategory')+ " + " + (item_sub or '')
											},
											fields=['article'])
						code = design_item_code[0].get("article", "") if design_item_code else ""
						product_size =  frappe.db.get_value("Item",row.get('design_id'),'master_bom')
						size =  frappe.db.get_value('BOM',product_size,'product_size')
						product_size_item = frappe.db.get_all(
								"Titan Size Master",
								filters={
									"item_category": row.get("category"),
									"customer": order_form_doc.customer_code,
									"product_size": ["like", f"%{size}%"]
								},
								fields=['code']
							)
						code_ctg = ['N&J','N&D','N&C','E&D']
						finding =  frappe.get_all("BOM Finding Detail",filters={'parent':row.get('bom')},fields=['finding_type'])
						finding_code = []
						if finding:
							for fnd in finding:
								fnd_code = frappe.db.get_value("Customer Finding Detail",
										{'parent':order_form_doc.customer_code,'gk_finding_sub_category':fnd.get('finding_type')},['code_finding'])
								if fnd_code:
									finding_code.append(fnd_code)
						row_data = [
							'-',
							'-',
							'-',
							row.get('design_id','') + code,
							'-',
							row.get('collection_name', '') ,
							'-',
							'STUDDED HIGH VALUE',
							'SET2' if code in code_ctg else 'SET1',
							'GURU',
							'-',
							'GO' if row.get('metal_type') == 'Gold' else '',
							metal_touch,
							product_size_item[0].get('code')if product_size_item else '',
							", ".join(finding_code) if finding_code else '',
							'N/A',
							'Round' if row.get('category') in ['Bangles', 'Ring'] else 'Oval' if row.get('category') == 'Bracelet' else 'N/A',
							row.get('uomset_of',''),
							row.get('gender',''),
							'YEL' if row.get('metal_colour','') == 'Yellow' else 'ROS',
							child_1[0].get('customer_category')if child_1 else '',
							child_2[0].get('customer_category') if child_2 else '', 
							'-', 
							'-', 
							'-',
							'-',
							'-',
							frappe.db.get_value('BOM',row.get('bom'),'metal_and_finding_weight'),
							frappe.db.get_value('BOM',item_sub_bom,'metal_and_finding_weight'),
							'-',
							'-',
							'-',
							'-',
							'-', 
							(float(frappe.db.get_value("BOM", row.get("bom"), "metal_and_finding_weight") or 0) + float(frappe.db.get_value("BOM", item_sub_bom, "metal_and_finding_weight") or 0)),
       						frappe.db.get_value('BOM',row.get('bom'),'diamond_weight'),
							frappe.db.get_value('BOM',item_sub_bom,'diamond_weight'),
							(frappe.db.get_value("BOM", row.get("bom"), "diamond_weight") or 0)+ (frappe.db.get_value("BOM", item_sub_bom, "diamond_weight") or 0),
							row.get('diamond_quality', ''),
							'DIMOND + COLOURSTONE' if(diamond and gemstone) else 'DIAMOND',
							'0',
							'0',
							'STUDDED'
							
						]
						rows_data.append(row_data) 
					else:
						metal_touch = row.get('metal_touch', '')
						metal_touch = metal_touch.replace('KT', '').replace('kt', '').strip()
						gemstone = frappe.db.get_value('BOM',row.get('bom'),'total_gemstone_pcs')
						diamond = frappe.db.get_value('BOM',row.get('bom'),'total_diamond_pcs')
						order_date = frappe.utils.formatdate(order_form_doc.order_date, "dd.MM.yyyy")
						# product_size =  frappe.db.get_value("Item",row.get('design_id'),'master_bom')
						# size =  frappe.db.get_value('BOM',product_size,'product_size')
						product_size = frappe.db.get_value(
								"Item Variant Attribute",
								{
									"parent": row.get("design_id"),
									"attribute": "Product Size",
								},
								['attribute_value']
							)
						product_size_item = None

						if product_size:
							product_size_item = frappe.db.get_all(
									"Titan Size Master",
									filters={
										"item_category": row.get("category"),
										"customer": order_form_doc.customer_code,
										"gk_product_size": ["like", f"%{product_size}%"]
									},
									fields=['code']
								)

						set_item =  frappe.db.get_all("Set Item Table",filters={'parent': row.get('design_id')},fields=['item_code'])
						item_sub = None
						if set_item:
							item_sub =  frappe.db.get_value("Item",set_item[0].get('item_code'),'item_subcategory')
						child_1 =  frappe.db.get_all("Customer Category Detail",
									filters = {
										'parent':order_form_doc.customer_code,
										'gk_category':row.get('category')
									},fields=['customer_category'])
						child_2 = []
						if item_sub:
							child_2 =  frappe.db.get_all("Customer Category Detail",
										filters = {
											'parent':order_form_doc.customer_code,
											'gk_sub_category':item_sub
										},fields=['customer_category'])
						finding =  frappe.get_all("BOM Finding Detail",filters={'parent':row.get('bom')},fields=['finding_type'])
						finding_code = []
						if finding:
							for fnd in finding:
								fnd_code = frappe.db.get_value("Customer Finding Detail",
										{'parent':order_form_doc.customer_code,'gk_finding_sub_category':fnd.get('finding_type')},['code_finding'])
								if fnd_code:
									finding_code.append(fnd_code)
						row_data = [
							'-',
							'-',
							'-',
							row.get('design_id',''),
							'-',
							row.get('collection_name', '') ,
							'-',
							'STUDDED HIGH VALUE',
							row.get('category', ''),
							'GURU',
							'-',
							'GO' if row.get('metal_type') == 'Gold' else '',
							metal_touch,
							product_size_item[0].get('code')if product_size_item else '',
							", ".join(finding_code) if finding_code else '',
							'N/A',
							'Round' if row.get('category') in ['Bangles', 'Ring'] else 'Oval' if row.get('category') == 'Bracelet' else 'N/A',
							row.get('uomset_of',''),
							row.get('gender',''),
							'YEL' if row.get('metal_colour','') == 'Yellow' else 'ROS',
							child_1[0].get('customer_category')if child_1 else '',
							child_2[0].get('customer_category') if child_2 else '', 
							'-', 
							'-', 
							'-',
							'-', 
							'-',
							'-',
							'-',
							'-',
							'-',
							'-',
							'-',
							'-', 
							frappe.db.get_value('BOM',row.get('bom'),'metal_and_finding_weight'),
							'-',
							'-',
							frappe.db.get_value('BOM',row.get('bom'),'diamond_weight'),
							row.get('diamond_quality', ''),
							'DIMOND + COLOURSTONE' if(diamond and gemstone) else 'DIAMOND',
							'0',
							'0',
							'STUDDED'
							
						]
						rows_data.append(row_data)
       

	# Write all rows to the Excel sheet at once
	if rows_data:
		for row in rows_data:
			sheet.append(row)
	else:
		frappe.throw("Code creation Sheet Can Not Download , Check all details..")
	output = BytesIO()
	workbook.save(output)
	output.seek(0)

	file_doc = save_file(
		file_name,
		output.getvalue(),
		order_form_doc.doctype,
		order_form_doc.name,
		is_private=0
	)
	
	return file_doc.file_url

@frappe.whitelist()
def proto_export_to_excel(order_form, doc):

	order_form_doc = frappe.get_doc('Order Form', order_form)
	doc = json.loads(doc)

	order_date_str = getdate(order_form_doc.order_date).strftime("%Y-%m-%d")
	file_name = f"Proto_Sheet_{order_date_str}.xlsx"

	workbook = openpyxl.Workbook()
	sheet = workbook.active
	sheet.title = 'Proto Sheet'
	
	# Store all rows in a list before writing to the sheet
	rows_data = []
	# frappe.throw(f"heree{order_form_doc.customer_name}")
	if 'Caratlane' in order_form_doc.customer_name:
		# Define headers
		# headers = [
		# 	"Design Selecion Date", "Collection Name", "Vendor Name", "Order Type",
		# 	"Image", "Theme Code", "Vendor/ Designer Ref Code", "Set Code", "Product Group",
		# 	"Product SubGroup", "Product Category", "Sub Category ",
		# 	"Category, Sub-Category Code", "Size", "Size (UOM)", "KT", "Metal Color",
		# 	"Diamond Quality", "Stone Proliferation", "Qty", "UOM", "Findings", "Proto Remarks in PO",
		# 	"Metal Purity", "Gross Wt.", "Gold Weight", "Diamond Carat Weight", "Polki Wt.",
		# 	"Other Stone Weight", "Polki Quality", "Gender", "Design Source/Route","TOTAL LABOUR AMOUNT", 
		# 	"DIAMOND HANDLING AMOUNT", "TOTAL DIAMOND AMOUNT", "COLORSTONE HANDLING AMOUNT",
		# 	"COLORSTONE AMOUNT", "GOLD AMOUNT", "LOSS AMOUNT",
		# 	"ADDITIONAL CHARGES", "TOTAL VALUE", "Design Complexity", "Need state",
		# 	"Primary Design language", "Name of the Design Motif", "Modularity Flag ", "Modularity description", "Finish Type ",
		# 	"Colour Stone Name", "Colour Stone Type","Colorstone Color Family","Enamel Color Family","Bangle"
		# ]
		headers = [
			"Caratlane SKU Code", "Item Code", "Vendor Style Code", "Images",
			"Gold Kt", "Gold Colour", "Product Type", "Product Size", "Stone Type",
			"Diamond Sieve Size/Col Stone", "Diamond Shape", "Diamond Sieve Size(mm Size)",
			"Quantity", "Individual Stone Wt", "Total Stone Wt", "Setting Type", "Type",
			"Stone Quality", "Stone Colour", "Cut", "Rate PCT", "Value", "Gross Weight",
			"Metal Colour", "Metal Karat", "Quantity", "Gold Weight", "Finding Name",
			"Finding Quantity", "Finding Weight", "Finding Colour", "Finding Karat","Finding Type", 
			"Net Weight(Min)", "Net Weight(Avg)", "Net Weight(Max)",
			"Diamond Weight(Min)", "Diamond Weight(Avg)", "Diamond Weight(Max)",
			# "Gemstone Type", "Gemstone Shape", "Gemstone Quantity", "Gemstone Weight",
			"Finishing Information", "Shipping Days", "Metal Rate", "Total Dia",
			"Cent per gm", "Labor", "Per Pc Labor", "Wastage", "Total", "Total Price", "Technique"
		]
		sheet.append(headers)

		# Loop through order details
		for row in doc.get('order_details', []):
			if row['design_id']:
				finish_bom_list = frappe.db.get_list("BOM", filters={'item': row["design_id"], 'bom_type': 'Template'}, fields=['name'])
				
				finish_bom = ''
				if len(finish_bom_list) > 1:
					order = frappe.db.get_value("Order", 
						{'cad_order_form': order_form, 'item': row['design_id']},'name')	
					pmo = frappe.db.get_value("Parent Manufacturing Order",{'order_form_id': order}, 'name')
					snc = frappe.db.get_value("Serial Number Creator",{'parent_manufacturing_order': pmo}, 'name')
					fg_bom = frappe.db.get_value("BOM", {'custom_serial_number_creator': snc, 'item': row['design_id'], 'bom_type': 'Finish Goods'}, 'name')
					finish_bom = fg_bom		
				else:
					for fg in finish_bom_list:
						finish_bom = fg.get('name')
				
				# frappe.throw(f"{finish_bom}")
				
				if finish_bom:
					item_image = frappe.db.get_value("Item", {'name': row["design_id"]}, ['image'])
					item_bom = frappe.db.get_list("BOM", filters={'item': row["design_id"], 'name': finish_bom}, fields=['*'])
					bom_metal = frappe.db.get_all("BOM Metal Detail", filters={'parent': finish_bom}, fields=['*'])
					bom_diamond = frappe.db.get_all("BOM Diamond Detail", filters={'parent': finish_bom}, fields=['*'])
					bom_finding = frappe.db.get_all("BOM Finding Detail", filters={'parent': finish_bom}, fields=['*'])
					bom_gems = frappe.db.get_all("BOM Gemstone Detail", filters={'parent': finish_bom}, fields=['*'])
									
					# Get the maximum number of rows needed for this item
					max_rows = max(len(bom_diamond), len(bom_finding), len(bom_metal), len(bom_gems)) or 1

					for i in range(max_rows):
						diamond = bom_diamond[i] if i < len(bom_diamond) else {}
						finding = bom_finding[i] if i < len(bom_finding) else {}
						metal = bom_metal[i] if i < len(bom_metal) else {}
						gemstone = bom_gems[i] if i < len(bom_gems) else {}
						
						diamond_tolerance = set_tolerance(diamond.get('quantity', 0), order_form_doc.customer_code)
						
						row_data = [
							"",  # Caratlane SKU Code
							row.get('design_id', '') if i == 0 else "",  # Item Code (only first row)
							"",  # Vendor Style Code
							item_image if i == 0 else "",  # Images (only first row)
							metal.get('metal_touch', '') if i == 0 else "",  # Gold Kt (only first row)
							metal.get('metal_colour', '') if i == 0 else "",  # Gold Colour (only first row)
							row.get('category', '') if i == 0 else "",  # Product Type (only first row)
							row.get("product_size", "") if i == 0 else "",  # Product Size (only first row)
							diamond.get('diamond_type', ''),  # Stone Type
							diamond.get('sieve_size_range', ''),  # Diamond Sieve Size/Col Stone
							diamond.get('stone_shape', ''),  # Diamond Shape
							diamond.get('size_in_mm', ''),  # Diamond Sieve Size(mm Size)
							diamond.get('pcs', ''),  # Quantity
							diamond.get('weight_per_pcs', ''),  # Individual Stone Wt
							"",  # Total Stone Wt (not available in the given structure)
							diamond.get('sub_setting_type', ''),  # Setting Type
							"",  # Type
							diamond.get('quality', ''),  # Stone Quality
							diamond.get('sieve_size_color', ''),  # Stone Colour
							"", "", "",  # Cut, Rate PCT, Value
							item_bom[0].get('gross_weight', '') if i == 0 else "",  # Gross Weight (only first row)
							metal.get('metal_colour', ''),  # Metal Colour
							metal.get('metal_touch', ''),  # Metal Karat
							metal.get('quantity', ''),  # Quantity
							metal.get('actual_quantity', ''),  # Gold Weight
							finding.get('finding_category', ''),  # Finding Name
							finding.get('qty', ''),  # Finding Quantity
							finding.get('quantity', ''),  # Finding Weight
							finding.get('metal_colour', ''),  # Finding Colour
							finding.get('metal_touch', ''),  # Finding Karat
							"", # Finding Type
							"", "", # Net Weights
							item_bom[0].get('metal_and_finding_weight', '') if i == 0 else "",
							# "", "", "", #, Diamond Weights
							diamond_tolerance.get('min_diamond', ''),  # Diamond Weight (Min)
							diamond_tolerance.get('diamond_weight', ''),  # Diamond Weight (Avg)
							diamond_tolerance.get('max_diamond', ''),  # Diamond Weight (Max)
							# "", "", "", "",  
							# gemstone.get('stone_type', ''),  # Gemstone Type
							# gemstone.get('stone_shape', ''),  # Gemstone Shape
							# gemstone.get('pcs', ''),  # Gemstone Quantity
							# gemstone.get('total_weight', ''),  # Gemstone Weight
							"", "", 
							metal.get('rate'), #metal rate 
							"", "", "", "", "", "", "", "", ""  # Remaining empty fields
						]
						rows_data.append(row_data)

						# if i == 0 and item_image:
						# 	try:
						# 		file_path = get_file(item_image)  # Fetch image file path
						# 		img = Image(file_path)
						# 		img.width, img.height = 100, 100  # Resize image
						# 		img_cell = f"D{row_index}"  # Column "D" (Images), row based on row_index
						# 		sheet.add_image(img, img_cell)
						# 	except Exception as e:
						# 		frappe.log_error(f"Image Insert Error: {str(e)}", "Proto Export")
	
	elif 'Reliance Retail Limited' in order_form_doc.customer_name:
		from openpyxl.styles import Alignment, Font
		# Define headers
		headers = [
			"Sr. NO.","Collection Name", "Vendor Name", "Vendor Design Code","TAG NO", "Proto Image", "Article", 
			"Metal Color", "Purity", "Stone Clarity", "Approx Net Wt (gms)", "Approx Dia Wt (cts)", 
			"Approx Color Stone Wt (cts)", "Size", "Findings", "Design Approved By", "Catrgory Approved By", 
			"Sourcing Approved By", "NPD Approved By", "QA Approved By", "QA Remarks", "Remark"
		]
		sheet.append(headers)
		
		# Header Style
		for cell in sheet[1]:
			cell.font = Font(bold=True)
			cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

		sheet.row_dimensions[1].height = 40


		# Column Widths
		sheet.column_dimensions['A'].width = 8
		sheet.column_dimensions['B'].width = 18
		sheet.column_dimensions['C'].width = 25
		sheet.column_dimensions['D'].width = 30
		sheet.column_dimensions['E'].width = 18
		sheet.column_dimensions['F'].width = 35   # IMAGE COLUMN
		sheet.column_dimensions['G'].width = 15
		sheet.column_dimensions['H'].width = 25
		sheet.column_dimensions['I'].width = 25
		sheet.column_dimensions['J'].width = 25
		sheet.column_dimensions['K'].width = 25
		sheet.column_dimensions['L'].width = 25
		sheet.column_dimensions['M'].width = 25
		sheet.column_dimensions['N'].width = 25
		sheet.column_dimensions['O'].width = 25
		sheet.column_dimensions['P'].width = 25
		sheet.column_dimensions['Q'].width = 25
		sheet.column_dimensions['R'].width = 25
		sheet.column_dimensions['S'].width = 25
		sheet.column_dimensions['T'].width = 25
		sheet.column_dimensions['U'].width = 25
		sheet.column_dimensions['V'].width = 25
		sheet.column_dimensions['W'].width = 25
		sheet.column_dimensions['X'].width = 25
		sheet.column_dimensions['Y'].width = 25
		sheet.column_dimensions['Z'].width = 25

		# Loop through order details
		for row in doc.get('order_details', []):
			if row['design_id']:
				finish_bom_list = frappe.db.get_list("BOM", filters={'item': row["design_id"], 'bom_type': 'Finish Goods'}, fields=['name'])
				
				finish_bom = ''
				if len(finish_bom_list) > 1:
					order = frappe.db.get_value("Order", 
						{'cad_order_form': order_form, 'item': row['design_id']},'name')	
					# frappe.throw(str(order_form))
					pmo = frappe.db.get_value("Parent Manufacturing Order",{'order_form_id': order}, 'name')
					snc = frappe.db.get_value("Serial Number Creator",{'parent_manufacturing_order': pmo}, 'name')
					fg_bom = frappe.db.get_value("BOM", {'custom_serial_number_creator': snc, 'item': row['design_id'], 'bom_type': 'Finish Goods'}, 'name')
					finish_bom = fg_bom		
				else:
					for fg in finish_bom_list:
						finish_bom = fg.get('name')
						
				
				final_bom = ''
				if finish_bom:
					final_bom = finish_bom
				else:
					final_bom = frappe.db.get_value("Item", {'name': row["design_id"],}, ['master_bom'])
								
				# frappe.throw(f"{finish_bom}")

				
				if final_bom:
					item_bom = frappe.db.get_list("BOM", filters={'item': row["design_id"], 'name': final_bom}, fields=['*'])

					realiance_quality = frappe.db.get_value("Customer Prolif Detail", 
						{'parent': order_form_doc.customer_code, 'gk_d': row.get('diamond_quality')  },
						['customer_prolif']
						) 
					realiance_quality if realiance_quality else ''

					product_size = row.get('product_size') if row.get('product_size') else 0
					# frappe.throw(str(prod))
					order_size = float(product_size)

					code_entry = frappe.db.get_all(
						"Reliance Size Master", 
						filters={
							'customer': order_form_doc.customer_code,
							'item_category': row.get('category'),
							'gurukrupa_size': product_size
						},
						fields=['code','product_size'],
						# as_dict=1
					)
					# frappe.throw(str(code_entry))

					code_size = code_entry[0]['code'] if code_entry else ''
					code_categories = frappe.db.get_value(
						"Customer Category Detail",
						{
							'parent': order_form_doc.customer_code,
							'gk_category': row.get('category') ,
							'gk_sub_category': row.get('subcategory') 
					  	},
						['customer_category','customer_subcategory','code_category','article'],
						as_dict=True
					)
					bom_finding = frappe.db.get_all("BOM Finding Detail",
                                    filters={
										'parent': final_bom
									},fields =['finding_type','finding_category'])
					code_list =[]
					for finding in bom_finding:
						finding_type = frappe.db.get_value("Customer Finding Detail",
							{
								# 'gk_finding_category':finding.get('finding_category'),
								'gk_finding_sub_category': finding.get('finding_type'),
								'parent': order_form_doc.customer_code,
							},
							['code_finding'],
							)
						if finding_type:
							code_list.append(finding_type)

					code_string = ", ".join(code_list)

					row_data = [
						row.get('idx') ,
						"IIJS SELECTION-2026" ,
						"GK", #order_form_doc.company,
						row.get('design_id', '') ,
						row.get('tag_no',''),  
						"",
						code_categories['article'], # row.get('category', ''),
						f"{row.get('metal_colour', '')} {row.get('metal_type', '')}",
						row.get('metal_touch', ''),
						realiance_quality, # row.get('diamond_quality', ''),
						item_bom[0].get('metal_and_finding_weight', '') ,
						item_bom[0].get('diamond_weight', '') , 
						"",
						code_size, # row.get('product_size', ''),
						# finding_type.get('code_finding') if finding_type else '',
						code_string,
						"",
						"",
						"",
						"",
						"",
						"",
						"",						
					]
					# rows_data.append(row_data)
					sheet.append(row_data)
					import os
					from openpyxl.drawing.image import Image

					current_row = sheet.max_row
					# Set row height for image
					sheet.row_dimensions[current_row].height = 150
					# Center align row
					for cell in sheet[current_row]:
						cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

					# Insert Image
					image_path = row.get("design_image_1")

					if image_path:
						full_path = frappe.get_site_path(image_path.replace("/files/", "public/files/"))
						if os.path.exists(full_path):

							img = Image(full_path)

							img.width = 180
							img.height = 180

							sheet.add_image(img, f"F{current_row}")


	elif 'Novel' in order_form_doc.customer_name:
		from openpyxl.styles import Alignment, Font
		# Define headers
		headers = [
			"SR. NO.","Design Selecion Date","Collection Name","Vendor Name","Order Type","Image","Theme Code","Vendor/ Designer Ref Code","Set Code",
			"Product Group","Product SubGroup","Product Category","Sub Category","Category, Sub-Category Code","Size","Size (UOM)","KT","Metal Color",
			"Diamond Quality","Stone Proliferation","Qty","UOM","Findings","Proto Remarks in PO","Metal Purity","Gross Wt.","Gold Weight","Diamond Carat Weight",
			"Polki Wt.","Other Stone Weight","Polki Quality","Gender","Design Source/Route","TOTAL LABOUR AMOUNT","DIAMOND HANDLING AMOUNT","TOTAL DIAMOND AMOUNT",
			"COLORSTONE HANDLING AMOUNT","COLORSTONE AMOUNT","GOLD AMOUNT","LOSS AMOUNT","ADDITIONAL CHARGES","TOTAL VALUE","Design Complexity","Need state",
			"Primary Design language","Name of the Design Motif","Modularity Flag","Modularity description","Finish Type","Colour Stone Name","Colour Stone Type",
			"Colorstone Color Family","Enamel Color Family","Bangle"
			]

		sheet.append(headers)

		# Header Style
		for cell in sheet[1]:
			cell.font = Font(bold=True)
			cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

		sheet.row_dimensions[1].height = 40


		# Column Widths
		sheet.column_dimensions['A'].width = 8
		sheet.column_dimensions['B'].width = 18
		sheet.column_dimensions['C'].width = 25
		sheet.column_dimensions['D'].width = 30
		sheet.column_dimensions['E'].width = 18
		sheet.column_dimensions['F'].width = 35   # IMAGE COLUMN
		sheet.column_dimensions['G'].width = 15
		sheet.column_dimensions['H'].width = 25
		sheet.column_dimensions['I'].width = 25
		sheet.column_dimensions['J'].width = 25
		sheet.column_dimensions['K'].width = 25
		sheet.column_dimensions['L'].width = 25
		sheet.column_dimensions['M'].width = 25
		sheet.column_dimensions['N'].width = 25
		sheet.column_dimensions['O'].width = 25
		sheet.column_dimensions['P'].width = 25
		sheet.column_dimensions['Q'].width = 25
		sheet.column_dimensions['R'].width = 25
		sheet.column_dimensions['S'].width = 25
		sheet.column_dimensions['T'].width = 25
		sheet.column_dimensions['U'].width = 25
		sheet.column_dimensions['V'].width = 25
		sheet.column_dimensions['W'].width = 25
		sheet.column_dimensions['X'].width = 25
		sheet.column_dimensions['Y'].width = 25
		sheet.column_dimensions['Z'].width = 25
		sheet.column_dimensions['AA'].width = 25
		sheet.column_dimensions['AB'].width = 25
		sheet.column_dimensions['AC'].width = 25
		sheet.column_dimensions['AD'].width = 25
		sheet.column_dimensions['AE'].width = 25
		sheet.column_dimensions['AF'].width = 25
		sheet.column_dimensions['AG'].width = 25
		sheet.column_dimensions['AH'].width = 25
		sheet.column_dimensions['AI'].width = 25
		sheet.column_dimensions['AJ'].width = 25
		sheet.column_dimensions['AK'].width = 25
		sheet.column_dimensions['AL'].width = 25
		sheet.column_dimensions['AM'].width = 25
		sheet.column_dimensions['AN'].width = 25
		sheet.column_dimensions['AO'].width = 25
		sheet.column_dimensions['AP'].width = 25
		sheet.column_dimensions['AQ'].width = 25
		sheet.column_dimensions['AR'].width = 25
		sheet.column_dimensions['AS'].width = 25
		sheet.column_dimensions['AT'].width = 25
		sheet.column_dimensions['AU'].width = 25
		sheet.column_dimensions['AV'].width = 25
		sheet.column_dimensions['AW'].width = 25
		sheet.column_dimensions['AX'].width = 25
		sheet.column_dimensions['AY'].width = 25
		sheet.column_dimensions['AZ'].width = 25
		sheet.column_dimensions['BA'].width = 25
		sheet.column_dimensions['BB'].width = 25
		# Loop through order details
		for row in doc.get('order_details', []):
			if row['design_id']:
				finish_bom_list = frappe.db.get_list("BOM", filters={'item': row["design_id"], 'bom_type': 'Finish Goods'}, fields=['name'])
				
				finish_bom = ''
				if len(finish_bom_list) > 1:
					order = frappe.db.get_value("Order", 
						{'cad_order_form': order_form, 'item': row['design_id']},'name')	
					pmo = frappe.db.get_value("Parent Manufacturing Order",{'order_form_id': order}, 'name')
					snc = frappe.db.get_value("Serial Number Creator",{'parent_manufacturing_order': pmo}, 'name')
					fg_bom = frappe.db.get_value("BOM", {'custom_serial_number_creator': snc, 'item': row['design_id'], 'bom_type': 'Finish Goods'}, 'name')
					finish_bom = fg_bom		
				else:
					for fg in finish_bom_list:
						finish_bom = fg.get('name')
				
				final_bom = ''
				if finish_bom:
					final_bom = finish_bom
				else:
					final_bom = frappe.db.get_value("Item", {'name': row["design_id"],}, ['master_bom'])
				
				# frappe.throw(f"{final_bom}")
				
				if final_bom:
					item_bom = frappe.db.get_list("BOM", filters={'item': row["design_id"], 'name': final_bom}, fields=['*'])
					order_date_fmt = frappe.utils.formatdate(order_form_doc.order_date, "dd-MM-yyyy")
					
					novel_quality = frappe.db.get_value("Customer Prolif Detail", 
						{'parent': order_form_doc.customer_code, 'gk_d': row.get('diamond_quality')  },
						['customer_prolif']
						) 
					novel_quality if novel_quality else ''
					
					product_size = row.get('product_size')
					order_size = float(product_size)
					amount = 0
					diamond_list = frappe.db.get_all("BOM Diamond Detail",filters ={"parent":final_bom},fields=["*"])
					for weight in diamond_list:
						rate_doc = frappe.db.get_value(
								"Diamond Price List",
								{
									"size_in_mm": weight.size_in_mm,
									"diamond_quality": order_form_doc.diamond_quality,
									"customer": order_form_doc.customer_code,
								},
								"rate"
							)
						if rate_doc:
							amount += rate_doc*flt(weight.get("quantity"), 3)
					
					
					finding_purity = frappe.db.get_all("BOM Finding Detail",
						 {
							'parent': item_bom[0].name,
						 },
						['metal_purity']
      					)
					finding_rate = 0
					gold_doc = frappe.db.get_value("Gold Rates",{"date":frappe.utils.today()},"name")
					gold_rate = frappe.db.get_all("Gold Rates branchs",{"parent":gold_doc,'particulars':"Jain Jewels"},"live_rate")
					gold_rate =  gold_rate[0].get("live_rate")/10
					metal_rate =  item_bom[0].get("total_metal_weight") * ((gold_rate *float(item_bom[0].get('metal_purity', 0)))/100)
					if finding_purity:
						finding_rate =  item_bom[0].get("total_finding_weight_per_gram") * ((gold_rate *float(finding_purity[0].get('metal_purity', 0)))/100)


					code_entry = frappe.db.get_value(
						"Novel Size Master",
						{
							'customer': order_form_doc.customer_code,
							'item_category': row.get('category'),
							'product_size_in': product_size
						},
						['code', 'product_size','size_umo'],
						as_dict=True
					)

					code_size = code_entry['code'] if code_entry else order_size

					metal_purity = float(item_bom[0].get('metal_purity', 0))
					converted_purity = round(metal_purity / 100, 2)

					code_categories = frappe.db.get_value(
						"Customer Category Detail",
						{
							'parent': order_form_doc.customer_code,
							'gk_category': row.get('category') ,
							'gk_sub_category': row.get('subcategory') 
					  	},
						['customer_category','customer_subcategory','code_category'],
						as_dict=True
					)
					finding_type = frappe.db.get_value("Customer Finding Detail",
						 {
							'description_finding':row.get('chain_type'),
							'parent': order_form_doc.customer_code,
						 },
						 as_dict=True
						 )
					diamond_qlty = frappe.db.get_value("Customer Prolif Detail",
						 {
							'parent': order_form_doc.customer_code,
							'gk_d' : row.get('diamond_quality')
						 },
						 ['customer_diamond_quality'],
						 as_dict=True
						 )
					making_charge_price = frappe.db.get_all("Making Charge Price",filters={
						"customer" :order_form_doc.customer_code,
						"setting_type" : row.get("setting_type"),
						"metal_touch": row.get("metal_touch")
						},
						fields=['name'])
					making_charge = frappe.db.get_all("Making Charge Price Item Subcategory",filters={
						"parent": making_charge_price[0].name,
						"mfg_complexity_code": row.get('mfg_complexity_code'),
						"subcategory":row.get("subcategory")
					},fields=['rate_per_gm'])
					# metal_rate =  item_bom[0].get("total_metal_weight") * ((gold_rate *float(item_bom[0].get('metal_purity', 0)))/100)
					# finding_rate =  item_bom[0].get("total_finding_weight_per_gram") * ((gold_rate *float(finding_purity[0].get('metal_purity', 0)))/100)
					diamond_price = frappe.db.get_all("Diamond Price List",filters={
									"customer" :order_form_doc.customer_code,
									"diamond_quality" : order_form_doc.diamond_quality,
									},
									fields=['outright_handling_charges_rate'])
					total_amt = (
								(diamond_price[0].outright_handling_charges_rate * item_bom[0].get("diamond_weight", 0) if diamond_price else 0)
								+ (making_charge[0].rate_per_gm * item_bom[0].get("metal_and_finding_weight", 0) if making_charge else 0)
								+ metal_rate + finding_rate
								+ amount
							)
					row_data = [
						row.get('idx') ,
						order_date_fmt,
						row.get('collection_name', '') ,
						order_form_doc.company,
						f"{order_form_doc.flow_type} Order",
						# row.get("design_image_1"),
						"",
						"",
						row.get('design_id', ''),
						row.get('category', ''),
						"Studded",
						"Studded-DIS",
						code_categories.get('customer_category', '') if code_categories else '',
						code_categories['customer_subcategory'] if code_categories else '',
						code_categories['code_category'] if code_categories else '', #code
						str(code_size) + '-' + str(row.get('product_size', '')),
						code_entry['size_umo'] if code_entry else '',
						row.get('metal_touch', ''),
						row.get('metal_colour', ''),
						diamond_qlty.get('customer_diamond_quality') if diamond_qlty else '',
						novel_quality,
						row.get('qty', ''),
						row.get('uomset_of', ''),
						row.get('chain_type') if finding_type else '', #finding
						"",
						converted_purity , #metal purity
						item_bom[0].get('gross_weight', '') , #gross wt
						item_bom[0].get('metal_and_finding_weight', 'metal_weight') , #gold wt
						item_bom[0].get('diamond_weight', '') , #diam wt
						"",
						"",
						"",
						row.get('gender', ''),
						row.get('design_sourceroute',''),
						making_charge[0].rate_per_gm * item_bom[0].get('metal_and_finding_weight', '') if making_charge else "", #labour amount
						diamond_price[0].outright_handling_charges_rate * item_bom[0].get('diamond_weight', '') if diamond_price else "", #diam handling amt
						amount, #diam amt
						"", #colorstone handling amt
						"", #colorstone amt
						metal_rate + finding_rate , #gold amt
						"", #loss amt
						"", #additional charge
						total_amt, #total value
						row.get('mfg_complexity_code', ''),
						"",
						"",
						"",
						"",
						"",
						"",
						"",
						"",
						"",
						"",
						"",
						]
					# rows_data.append(row_data)

					# Add row to excel
					sheet.append(row_data)

					import os
					from openpyxl.drawing.image import Image

					current_row = sheet.max_row
					# Set row height for image
					sheet.row_dimensions[current_row].height = 150
					# Center align row
					for cell in sheet[current_row]:
						cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

					# Insert Image
					image_path = row.get("design_image_1")

					if image_path:
						full_path = frappe.get_site_path(image_path.replace("/files/", "public/files/"))
						if os.path.exists(full_path):

							img = Image(full_path)

							img.width = 180
							img.height = 180

							sheet.add_image(img, f"F{current_row}")

	
	# Write all rows to the Excel sheet at once
	# if rows_data:
	# 	for row in rows_data:
	# 		sheet.append(row)
	# else:
	# 	frappe.throw("Proto Sheet Can Not Download")
	if sheet.max_row <= 1:
		frappe.throw("Proto Sheet Can Not Download")

	# Save the workbook to a BytesIO stream
	output = BytesIO()
	workbook.save(output)
	output.seek(0)

	file_doc = save_file(
		file_name,
		output.getvalue(),
		order_form_doc.doctype,
		order_form_doc.name,
		is_private=0
	)

	return file_doc.file_url

@frappe.whitelist()
def get_variant_format(order_form, doc): 
	
	order_form_doc = frappe.get_doc('Order Form', order_form)
	doc = json.loads(doc)

	order_date_str = getdate(order_form_doc.order_date).strftime("%Y-%m-%d")
	file_name = f"Variant_Format_{order_date_str}.xlsx"

	workbook = openpyxl.Workbook()
	sheet = workbook.active
	sheet.title = 'Variant Format'

	rows_data = []

	if 'Reliance Retail Limited' in order_form_doc.customer_name:
		headers = [
			"Vendor Code","Article","Vendor design code", "Purity","Set of", "Metal Color", 
			"Dia quality", "Variant Size","Net Wt(14kt)","Net Wt(18kt)","Dia pcs", "Dia Wt",
			"Color Stone pcs", "Color Stone Wt", "Gross Wt(14kt)","Gross Wt(18kt)", "Remark","","RELIANCE SIZE"
		]
		sheet.append(headers)

		# Loop through order details
		for row in doc.get('order_details', []):
			if row.get('category') in ['Bangles','Bracelet','Ring']:
				if row['design_id']:
					finish_bom_list = frappe.db.get_list("BOM", filters={'item': row["design_id"], 'bom_type': 'Finish Goods'}, fields=['name'])
					
					finish_bom = ''
					if len(finish_bom_list) > 1:
						order = frappe.db.get_value("Order", 
							{'cad_order_form': order_form, 'item': row['design_id']},'name')	
						pmo = frappe.db.get_value("Parent Manufacturing Order",{'order_form_id': order}, 'name')
						snc = frappe.db.get_value("Serial Number Creator",{'parent_manufacturing_order': pmo}, 'name')
						fg_bom = frappe.db.get_value("BOM", {'custom_serial_number_creator': snc, 'item': row['design_id'], 'bom_type': 'Finish Goods'}, 'name')
						finish_bom = fg_bom		
					else:
						for fg in finish_bom_list:
							finish_bom = fg.get('name')
					
					final_bom = ''
					if finish_bom:
						final_bom = finish_bom
					else:
						final_bom = frappe.db.get_value("Item", {'name': row["design_id"],}, ['master_bom'])
									
					# frappe.throw(f"{finish_bom}")
					
					if final_bom:
						item_bom = frappe.db.get_list("BOM", filters={'item': row["design_id"], 'name': final_bom}, fields=['*'])
						
						realiance_quality = frappe.db.get_value("Customer Prolif Detail", 
							{'parent': order_form_doc.customer_code, 'gk_d': row.get('diamond_quality')  },
							['customer_prolif']
							) 
						realiance_quality if realiance_quality else ''

						product_size = row.get('product_size') if row.get('product_size') else 0
						order_size = float(product_size)


						code_categories = frappe.db.get_value(
							"Customer Category Detail",
							{
								'parent': order_form_doc.customer_code,
								'gk_category': row.get('category') ,
								'gk_sub_category': row.get('subcategory') 
							},
							['customer_category','customer_subcategory','code_category','article'],
							as_dict=True
						)
						
						wax_to_gold_ratio_22 = frappe.db.get_value(
							"Metal Ratio",
							{
								'Parent': order_form_doc.customer_code,
								'touch': '22KT',
								'metal_color': row.get('metal_colour'),
								'setting_type': row.get('setting_type')
							},
							"wax_to_gold_ratio"
						)
						wax_to_gold_ratio = frappe.db.get_value(
							"Metal Ratio",
							{
								'Parent': order_form_doc.customer_code,
								'touch': '18KT',
								'metal_color': row.get('metal_colour'),
								'setting_type': row.get('setting_type')
							},
							"wax_to_gold_ratio"
						)
						wax_to_gold_ratio_14kt = frappe.db.get_value(
							"Metal Ratio",
							{
								'Parent': order_form_doc.customer_code,
								'touch': '14KT',
								'metal_color': row.get('metal_colour'),
								'setting_type': row.get('setting_type')
							},
							"wax_to_gold_ratio"
						)
						metal = 0
						metal_18 = 0
						gross_18 = 0
						gross_14 = 0
						if wax_to_gold_ratio_14kt and wax_to_gold_ratio:
							if row.get('metal_touch') == '22KT':	
								metal_18 =  round((item_bom[0].get('metal_and_finding_weight', '')/float(wax_to_gold_ratio_22))*float(wax_to_gold_ratio),3)
								metal =  round((metal_18/float(wax_to_gold_ratio))*float(wax_to_gold_ratio_14kt),3)
								gross_18 = round((item_bom[0].get('gross_weight', '')/float(wax_to_gold_ratio_22))*float(wax_to_gold_ratio),3)
								gross_14 =  round((gross_18/float(wax_to_gold_ratio))*float(wax_to_gold_ratio_14kt),3)

							else:
								metal =  round((item_bom[0].get('metal_and_finding_weight', '')/float(wax_to_gold_ratio))*float(wax_to_gold_ratio_14kt),3)
								gross_14 =  round((item_bom[0].get('gross_weight', '')/float(wax_to_gold_ratio))*float(wax_to_gold_ratio_14kt),3)
							
							



						all_sizes = frappe.db.get_all(
							"Reliance Size Master",
							filters={
								'customer': order_form_doc.customer_code,
								'item_category': row.get('category'),
								'item_subcategory':row.get('subcategory')
							},
							fields=['code', 'gurukrupa_size','product_size'],   # 👈 changed field
							order_by='gurukrupa_size asc'
						)

						product_size = float(row.get('product_size'))   # input

						match_index = None
						min_diff = None

						# Find closest match based on gurukrupa_size
						for i, d in enumerate(all_sizes):
							size_str = d.get('gurukrupa_size')
							
							if size_str:
								size_value = float(size_str.replace(' MM', ''))
								diff = abs(size_value - product_size)

								if min_diff is None or diff < min_diff:
									min_diff = diff
									match_index = i

						result = []

						if match_index is not None:
							start = max(match_index - 2, 0)
							end = match_index + 4
							result = all_sizes[start:end]

						if result:
							for size_row in result:
								code_size = size_row.get('code')  #
								row_data = [
									"60450001",
									code_categories['code_category'],
									row.get('design_id', '') ,  
									row.get('metal_touch', ''),
									"-",
									f"{row.get('metal_colour', '')} {row.get('metal_type', '')}",
									realiance_quality,
									code_size,
									metal,
									metal_18 if row.get('metal_touch') == '22KT' else item_bom[0].get('metal_and_finding_weight') ,
									item_bom[0].get('total_diamond_pcs', '') ,
									item_bom[0].get('diamond_weight', '') , 
									item_bom[0].get('total_gemstone_pcs', '') , 
									item_bom[0].get('total_gemstone_weight', '') , 
									gross_14,
									gross_18 if row.get('metal_touch') == '22KT' else item_bom[0].get('gross_weight', '') , 
									"",
									"",
									size_row.get('product_size','')


								]
								rows_data.append(row_data)
							rows_data.append([""] * len(row_data))
						else:
							code_size = size_row.get('code')  #
							row_data = [
								"60450001",
								code_categories['code_category'],
								row.get('design_id', '') ,  
								row.get('metal_touch', ''),
								"-",
								f"{row.get('metal_colour', '')} {row.get('metal_type', '')}",
								realiance_quality,
								code_size,
								metal,
								metal_18 if row.get('metal_touch') == '22KT' else item_bom[0].get('metal_and_finding_weight') ,
								item_bom[0].get('total_diamond_pcs', '') ,
								item_bom[0].get('diamond_weight', '') , 
								item_bom[0].get('total_gemstone_pcs', '') , 
								item_bom[0].get('total_gemstone_weight', '') , 
								gross_14,
								gross_18 if row.get('metal_touch') == '22KT' else item_bom[0].get('gross_weight', '') , 
								"",
								"",
								"",

							]
							rows_data.append(row_data)
						rows_data.append([""] * len(row_data))
					# Write all rows to the Excel sheet at once
	
	if rows_data:
		for row in rows_data:
			sheet.append(row)
	else:
		frappe.throw("Proto Sheet Can Not Download")

	# Save the workbook to a BytesIO stream
	output = BytesIO()
	workbook.save(output)
	output.seek(0)

	file_doc = save_file(
		file_name,
		output.getvalue(),
		order_form_doc.doctype,
		order_form_doc.name,
		is_private=0
	)

	return file_doc.file_url

def set_tolerance(diamond_weight, customer):
	data_json = {}
	if diamond_weight:
		tolerance_data = frappe.db.get_all('Diamond Tolerance Table',
			filters={'weight_type': 'Weight wise', 'parent': customer}, 
			fields=['from_diamond', 'to_diamond', 'plus_percent', 'minus_percent'])

		for row in tolerance_data:
			if row['from_diamond'] <= diamond_weight <= row['to_diamond']:
				plus_percent = row['plus_percent']
				minus_percent = row['minus_percent']

				max_diamond_weight = diamond_weight + plus_percent
				min_diamond_weight = diamond_weight - minus_percent
				
				data_json['diamond_weight'] = round(diamond_weight, 3)
				data_json['max_diamond'] = round(max_diamond_weight, 3)
				data_json['min_diamond'] = round(min_diamond_weight, 3)
				
	return data_json







@frappe.whitelist()
def get_design_creation(order_form, doc):
	from openpyxl.drawing.image import Image
	from frappe.utils.file_manager import get_file
	import requests
	import os

	order_form_doc = frappe.get_doc('Order Form', order_form)
	doc = json.loads(doc)

	order_date_str = getdate(order_form_doc.order_date).strftime("%Y-%m-%d")
	file_name = f"Code_Creation{order_date_str}.xlsx"

	workbook = openpyxl.Workbook()
	sheet = workbook.active
	sheet.title = 'Code Creation'

	rows_data = []
	headers = [
		"SR. NO.",
		"PRODUCT DIMENTION / METAL CONFIGURATION",
		"PRODUCT TYPE",
		"MANUFACTURING TYPE (MFG_Code)",
		"SET REF-PEARL COLOUR (WatchDialColor)",
		"METAL COLOUR (WatchDialMetal)",
		"GENDER (Gender)",
		"APPROX COST $",
		"GROSS WT",
		"TOTAL GOLD wt.(Net Wt for SKU)",
		"Total DIA CT",
		"SUPPLIER Style No",
		"VENDOR CODE",
		"Product Pcs",
		"METAL ITEM",
		"NET GOLD WEIGHT",
		"CHAIN ITEM",
		"Chain Wt",
		"Finding item",
		"FINDING WT",
		"Diamond Stone Item",
		"Diamond Size",
		"Diamond Code (Cut/Colour/Clarity)",
		"Diamond Pcs",
		"Diamond Cts",
		"Diamond Stone Ct Unit",
		"Colour Stone Item",
		"Colour Stone Size",
		"Colour Stone Code",
		"Colour Stone Pcs",
		"Colour Stone Ct",
		"Colour Stone Ct Unit",
		"Color stone price Per ct.($)",
		"Manufaturing Style*",
	]

	sheet.append(headers)
	for cell in sheet[1]:
		cell.font = Font(bold=True)
		cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

	sheet.row_dimensions[1].height = 40

	# Column Widths
	sheet.column_dimensions['A'].width = 8
	sheet.column_dimensions['B'].width = 18
	sheet.column_dimensions['C'].width = 25
	sheet.column_dimensions['D'].width = 30
	sheet.column_dimensions['E'].width = 18
	sheet.column_dimensions['F'].width = 35
	sheet.column_dimensions['G'].width = 15
	sheet.column_dimensions['H'].width = 25
	sheet.column_dimensions['I'].width = 25
	sheet.column_dimensions['J'].width = 25
	sheet.column_dimensions['K'].width = 25
	sheet.column_dimensions['L'].width = 25
	sheet.column_dimensions['M'].width = 25
	sheet.column_dimensions['N'].width = 25
	sheet.column_dimensions['O'].width = 25
	sheet.column_dimensions['P'].width = 25
	sheet.column_dimensions['Q'].width = 25
	sheet.column_dimensions['R'].width = 25
	sheet.column_dimensions['S'].width = 25
	sheet.column_dimensions['T'].width = 25
	sheet.column_dimensions['U'].width = 25
	sheet.column_dimensions['V'].width = 25
	sheet.column_dimensions['W'].width = 25
	sheet.column_dimensions['X'].width = 25
	sheet.column_dimensions['Y'].width = 25
	sheet.column_dimensions['Z'].width = 25

	# Loop through order details
	for row in doc.get('order_details', []):
		if row['design_id']:
			serial_bom = frappe.db.get_value("Serial No", row.get("tag_no"), "custom_bom_no")
			final_bom = frappe.get_doc("BOM", serial_bom)

			if final_bom:
				metal_touch = frappe.db.get_all("Customer Metal Touch Detail", filters={
					"gk_metal_touch": row.get('metal_touch'),
					"parent": order_form_doc.customer_code
				}, fields=['customer_metal_touch'])

				product_type = frappe.get_all("Customer Category Detail", filters={
					"gk_category": row.get('category'),
					"parent": order_form_doc.customer_code
				}, fields=['customer_category'])

				metal_color = frappe.get_all("Customer Metal Color Detail", filters={
					"gk_metal_color": row.get('metal_colour'),
					"parent": order_form_doc.customer_code
				}, fields=['code_color'])

				metal_type = frappe.get_all("Customer Metal Type Detail", filters={
					"gk_metal_type": row.get('metal_type'),
					"parent": order_form_doc.customer_code
				}, fields=['customer_metal_type'])

				chain_weight = frappe.get_all("BOM Finding Detail", filters={
					"metal_type": row.get('metal_type'),
					"parent": final_bom
				}, fields=['quantity'])

				# Fetch setting ONCE per order row
				setting = frappe.get_all("Customer Setting Detail", filters={
					"gk_setting_type": row.get("setting_type"),
					"gk_sub_setting_type": row.get("sub_setting_type1"),
					"parent": order_form_doc.customer_code
				}, fields=['customer_setting_size'])

				gemstone_shape = []
				diamond_shape = []

				# Build stone_item_list ONCE per order row
				# Merge when resolved group + code are same
				stone_items = {}
				for d in final_bom.diamond_detail:
					diamond_stone_item = frappe.get_all(
						"Customer Diamond Size Details",
						filters={
							"gk_diamond_shape": d.get("stone_shape"),
							"gk_diamond_size_in_mm": ["like", f"%{d.get('size_in_mm')}%"],
							"parent": order_form_doc.customer_code
						},
						fields=["customer_diamond_group_size", "code"],
						limit_page_length=1
					)

					if not diamond_stone_item:
						continue

					diamond_stone_item = diamond_stone_item[0]

					resolved_code = diamond_stone_item.get("code") or ''
					resolved_group = diamond_stone_item.get("customer_diamond_group_size") or ''

					# Merge when resolved group + code are same
					key = (resolved_group, resolved_code)

					if key in stone_items:
						stone_items[key]["pcs"] += d.get("pcs") or 0
						stone_items[key]["quantity"] += d.get("quantity") or 0
					else:
						stone_items[key] = {
							"stone_shape": d.get("stone_shape"),
							"size_in_mm": d.get("size_in_mm"),
							"customer_diamond_group_size": resolved_group,
							"code": resolved_code,
							"pcs": d.get("pcs") or 0,
							"quantity": d.get("quantity") or 0
						}

				stone_item_list = list(stone_items.values())

				# Fetch diamond_shape ONCE per order row using first diamond detail
				if len(final_bom.diamond_detail) > 0:
					diamond_shape = frappe.get_all("Customer Diamond Shape Detail",
						filters={
							"gk_diamond_shape": final_bom.diamond_detail[0].get('stone_shape'),
							"parent": order_form_doc.customer_code
						}, fields=['data'])

				# Total rows = longer of merged stone list vs gemstone detail
				total_rows = max(len(stone_item_list), len(final_bom.gemstone_detail))
				if total_rows == 0:
					total_rows = 1

				for item in range(total_rows):
					# Gemstone shape changes per item so fetch inside loop
					gemstone_shape = []
					if len(final_bom.gemstone_detail) > item:
						gemstone_shape = frappe.get_all("Customer Gemstone Detail",
							filters={
								"gk_gemstone_type": final_bom.gemstone_detail[item].get('gemstone_type'),
								"parent": order_form_doc.customer_code
							}, fields=['customer_gemstone_type', 'customer_gemstone_size', 'code'])

					# Only show stone data when item is within stone_item_list range
					# Do NOT fall back to last item — show empty instead
					has_stone = len(stone_item_list) > item
					inr_amount = final_bom.get("total_bom_amount") or 0
					exchange_rate = frappe.db.get_value(
						"Currency Exchange",
						filters={
							"from_currency": "USD",
							"to_currency": "INR"
						},
						fieldname="exchange_rate",
						order_by="date desc"
					)

					exchange_rate = exchange_rate or 0
					gemstone_rate = None

					if len(final_bom.gemstone_detail) > item:
						gemstone_type = final_bom.gemstone_detail[item].get("gemstone_type")
						qty = final_bom.gemstone_detail[item].get("quantity")

						result = frappe.db.sql("""
							SELECT gpl.rate
							FROM `tabGemstone Multiplier` gm
							INNER JOIN `tabGemstone Price List` gpl
								ON gpl.name = gm.parent
							WHERE gm.gemstone_type LIKE %(gemstone_type)s
							AND gm.from_weight <= %(qty)s
							AND gm.to_weight >= %(qty)s
							AND gpl.customer = %(customer)s
							LIMIT 1
						""", {
							"gemstone_type": f"%{gemstone_type}%",
							"qty": qty,
							"customer": order_form_doc.customer_code
						}, as_dict=1)

						if result:
							gemstone_rate = result[0].rate
					usd_amount = 0
					if exchange_rate:
						usd_amount = round(inr_amount/exchange_rate,3)
					stone_cols = [
						stone_item_list[item].get('code') if has_stone else '',
						stone_item_list[item].get('customer_diamond_group_size') if has_stone else '',
						diamond_shape[0].get('data', '') + '/' + order_form_doc.diamond_quality if (diamond_shape and has_stone) else '',
						stone_item_list[item].get('pcs') if has_stone else '',
						stone_item_list[item].get('quantity') if has_stone else '',
						"CT" if has_stone else '',
						gemstone_shape[0].get("customer_gemstone_type", '') if gemstone_shape else '',
						gemstone_shape[0].get("customer_gemstone_size", '') if gemstone_shape else '',
						gemstone_shape[0].get("code", '') if gemstone_shape else '',
						final_bom.gemstone_detail[item].get('pcs') if len(final_bom.gemstone_detail) > item else "",
						final_bom.gemstone_detail[item].get('quantity') if len(final_bom.gemstone_detail) > item else "",
						"CT" if len(final_bom.gemstone_detail) > item else '',
						# final_bom.gemstone_detail[item].get('total_gemstone_rate') if len(final_bom.gemstone_detail) > item else "",
						gemstone_rate or "" ,
						# Setting only on first row
						setting[0].get('customer_setting_size') if (setting and item == 0) else "",
					]

					if item == 0:
						# First iteration: full base row + stone cols
						row_data = [
							row.get('idx'),
							metal_touch[0].get('customer_metal_touch') if metal_touch else '',
							product_type[0].get('customer_category') if product_type else "",
							"",  # mgf_code
							row.get('set_ref_pearl_color',''),  # set color like stand alone
							metal_color[0].get('code_color') if metal_color else '',
							row.get('gender', ''),
							usd_amount,
							round(final_bom.get("gross_weight"),3) or 0,
							round(final_bom.get("metal_and_finding_weight"),2) or 0,
							final_bom.get("diamond_weight") or 0,
							row.get('design_id', ''),
							"106-VENDOR ACCOUNT NO",
							row.get('qty', ''),
							metal_type[0].get('customer_metal_type') if metal_type else '',
							final_bom.get('total_metal_weight') or 0,
							"CHAIN" if chain_weight else "-",
							chain_weight[0].get('quantity') if chain_weight else 0,
							"FINDINGS" if final_bom.get('finding_weight')  else "",
							final_bom.get('finding_weight') or "",
						] + stone_cols
					else:
						# Subsequent iterations: 20 empty base columns + stone cols only
						row_data = [""] * 20 + stone_cols

					rows_data.append(row_data)

	# Write all rows to the Excel sheet at once
	if rows_data:
		for row in rows_data:
			sheet.append(row)
	else:
		frappe.throw("Design Sheet Can Not Download")

	# Save the workbook to a BytesIO stream
	output = BytesIO()
	workbook.save(output)
	output.seek(0)

	file_doc = save_file(
		file_name,
		output.getvalue(),
		order_form_doc.doctype,
		order_form_doc.name,
		is_private=0
	)

	return file_doc.file_url



@frappe.whitelist()
def design_quotation_file(order_form, doc):
	from openpyxl.drawing.image import Image
	from frappe.utils.file_manager import get_file
	from openpyxl.styles import Alignment, Font

	order_form_doc = frappe.get_doc('Order Form', order_form)
	doc = json.loads(doc)

	order_date_str = getdate(order_form_doc.order_date).strftime("%Y-%m-%d")
	file_name = f"Code_Creation{order_date_str}.xlsx"

	workbook = openpyxl.Workbook()
	sheet = workbook.active
	sheet.title = 'Quotation File'
	
	# Store all rows in a list before writing to the sheet
	rows_data = []
	headers = [
		"Sr No",
  		"Vendor Name",
		"Vendor Design Code",
		"MGD Code",
		"Image",
		"Article",
		"Qty",
		"Set Reference",
		"Product Shape",
		"Style",
		"Primary Theme",
		"Diamond Setting Type",
		"Product Size & NK,PN Length",
		"Metal Karat",
		"Metal Color",
		"Gross Wt",
		"Dmd Wt",
		"Dmd Quality",
		"CST Wt",
		"Vendor Cost (USD)",
		"Remarks",
		"Selected By",

		]

	sheet.append(headers)
	# Loop through order details
	for cell in sheet[1]:
		cell.font = Font(bold=True)
		cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

		sheet.row_dimensions[1].height = 40


		# Column Widths
		sheet.column_dimensions['A'].width = 8
		sheet.column_dimensions['B'].width = 18
		sheet.column_dimensions['C'].width = 25
		sheet.column_dimensions['D'].width = 30
		sheet.column_dimensions['E'].width = 18
		sheet.column_dimensions['F'].width = 35   # IMAGE COLUMN
		sheet.column_dimensions['G'].width = 15
		sheet.column_dimensions['H'].width = 25
		sheet.column_dimensions['I'].width = 25
		sheet.column_dimensions['J'].width = 25
		sheet.column_dimensions['K'].width = 25
		sheet.column_dimensions['L'].width = 25
		sheet.column_dimensions['M'].width = 25
		sheet.column_dimensions['N'].width = 25
		sheet.column_dimensions['O'].width = 25
		sheet.column_dimensions['P'].width = 25
		sheet.column_dimensions['Q'].width = 25
		sheet.column_dimensions['R'].width = 25
		sheet.column_dimensions['S'].width = 25
		sheet.column_dimensions['T'].width = 25
		sheet.column_dimensions['U'].width = 25
		sheet.column_dimensions['V'].width = 25
		sheet.column_dimensions['W'].width = 25
		sheet.column_dimensions['X'].width = 25
		sheet.column_dimensions['Y'].width = 25
		sheet.column_dimensions['Z'].width = 25

	for row in doc.get('order_details', []):
		if row['design_id']:
			serial_bom =  frappe.db.get_value("Serial No",row.get("tag_no"),"custom_bom_no")
			final_bom =  frappe.get_doc("BOM",serial_bom)
			
			if final_bom:
				metal_touch =  frappe.db.get_all("Customer Metal Touch Detail",filters = {
					"gk_metal_touch":row.get('metal_touch'),
					"parent":order_form_doc.customer_code
				},fields=['customer_metal_touch'])

				setting = frappe.get_all("Customer Setting Detail",
							filters = {
									"gk_setting_type":row.get("setting_type"),
									"gk_sub_setting_type": row.get("sub_setting_type1"),
									"parent": order_form_doc.customer_code
							},fields=['customer_setting_size'])
    
				metal_color = frappe.get_all("Customer Metal Color Detail",
                                  filters = {
									  	"gk_metal_color":row.get('metal_colour'),
										"parent":order_form_doc.customer_code
								  },fields=['code_color'])

				inr_amount = final_bom.get("total_bom_amount") or 0
				exchange_rate = frappe.db.get_value(
					"Currency Exchange",
					filters={
						"from_currency": "USD",
						"to_currency": "INR"
					},
					fieldname="exchange_rate",
					order_by="date desc"
				)
				usd_amount = 0
				if exchange_rate:
					usd_amount = round(inr_amount/exchange_rate,3)
			
				row_data = [
					row.get('idx') ,
					"106-VND-000000022",
					row.get('design_id'),
					"", #mgf_code
					"",
					row.get('category'),
					row.get('qty', ''),
					row.get('set_ref_pearl_color'),
					"",
					"",
					"",
					setting[0].get('customer_setting_size') if setting else "",
					row.get('qty', ''),
					metal_touch[0].get('customer_metal_touch' or '') if metal_touch else '',
					metal_color[0].get('code_color') if metal_color else '',
     				round(final_bom.get("gross_weight"),3) or 0,
					final_bom.get("diamond_weight" or 0),
					order_form_doc.diamond_quality,
					final_bom.get("gemstone_weight" or ''),
					usd_amount,
					"",
					"",
					]
				rows_data.append(row_data)
    
				import os
				from openpyxl.drawing.image import Image

				current_row = sheet.max_row
				# Set row height for image
				# sheet.row_dimensions[current_row].height = 150

				# Center align row
				for cell in sheet[current_row]:
					cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

				# ---------------- INSERT IMAGE IN COLUMN E ----------------
				image_path = final_bom.get("finish_front_view_preview")

				if image_path:
					# Convert /files/xyz.png -> public/files/xyz.png
					full_path = frappe.get_site_path(image_path.replace("/files/", "public/files/"))

					if os.path.exists(full_path):
						img = Image(full_path)
						img.width = 180
						img.height = 180

						# Insert image in Column E
						sheet.add_image(img, f"E{current_row}")
	
	# Write all rows to the Excel sheet at once
	if rows_data:
		for row in rows_data:
			sheet.append(row)
	else:
		frappe.throw("Design Quotation Can Not Download")

	# Save the workbook to a BytesIO stream
	output = BytesIO()
	workbook.save(output)
	output.seek(0)

	file_doc = save_file(
		file_name,
		output.getvalue(),
		order_form_doc.doctype,
		order_form_doc.name,
		is_private=0
	)

	return file_doc.file_url











@frappe.whitelist()
def bom_format(order_form, doc):
	from openpyxl.drawing.image import Image
	from frappe.utils.file_manager import get_file
	from openpyxl.styles import Alignment, Font

	order_form_doc = frappe.get_doc('Order Form', order_form)
	doc = json.loads(doc)

	order_date_str = getdate(order_form_doc.order_date).strftime("%Y-%m-%d")
	file_name = f"Code_Creation{order_date_str}.xlsx"

	workbook = openpyxl.Workbook()
	sheet = workbook.active
	sheet.title = 'Code Creation'
	
	# Store all rows in a list before writing to the sheet
	rows_data = []
	headers = [
		"VENDORCODE",
  		"VENDORDESIGN",
		"MATCHING DESIGN",
		"PRODUCTTYPE",
		"ITEM",
		"IMAGE",
		"FORM",
		"COLLECTION",
		"SUBCOLLECTION",
		"MANUFACTURING TYPE",
		"THEME",
		"GENDER",
		"PLAIN/STUDDED",
		"SETTING TYPE",
		"SOLITAIRE",
		"CODE",
		"INGRADIANT TYPE",
		"STONETYPE",
		"SHAPE",
		"SIEVE",
		"STONEQTY",
		"CTS",
		"GRM",
		"GROSS",
		"NET",
		"NODD",
		"DDCTS",
		"COLCTS",
		"CUT",
		"CLARITY",
		"COLOR",
		"PURITY",
		"MGDFITTINGTYPECODE",
		"MGDSETTINGTYPECODE",
		"FINISH",
		"ORNAMENTSIZE",
		"POLISHTYPE",
		"METALCOLOR",
		"HEIGHT",
		"WIDTH",
		"LENGTH",
		"NECKLACE PENDANT HIGHT",

		]

	sheet.append(headers)
	# Loop through order details
	for cell in sheet[1]:
		cell.font = Font(bold=True)
		cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

		sheet.row_dimensions[1].height = 40


		# Column Widths
		sheet.column_dimensions['A'].width = 8
		sheet.column_dimensions['B'].width = 18
		sheet.column_dimensions['C'].width = 25
		sheet.column_dimensions['D'].width = 30
		sheet.column_dimensions['E'].width = 18
		sheet.column_dimensions['F'].width = 35   # IMAGE COLUMN
		sheet.column_dimensions['G'].width = 15
		sheet.column_dimensions['H'].width = 25
		sheet.column_dimensions['I'].width = 25
		sheet.column_dimensions['J'].width = 25
		sheet.column_dimensions['K'].width = 25
		sheet.column_dimensions['L'].width = 25
		sheet.column_dimensions['M'].width = 25
		sheet.column_dimensions['N'].width = 25
		sheet.column_dimensions['O'].width = 25
		sheet.column_dimensions['P'].width = 25
		sheet.column_dimensions['Q'].width = 25
		sheet.column_dimensions['R'].width = 25
		sheet.column_dimensions['S'].width = 25
		sheet.column_dimensions['T'].width = 25
		sheet.column_dimensions['U'].width = 25
		sheet.column_dimensions['V'].width = 25
		sheet.column_dimensions['W'].width = 25
		sheet.column_dimensions['X'].width = 25
		sheet.column_dimensions['Y'].width = 25
		sheet.column_dimensions['Z'].width = 25
		sheet.column_dimensions['AA'].width = 25
		sheet.column_dimensions['AB'].width = 25
		sheet.column_dimensions['AC'].width = 25
		sheet.column_dimensions['AD'].width = 25
		sheet.column_dimensions['AE'].width = 25
		sheet.column_dimensions['AF'].width = 25
		sheet.column_dimensions['AG'].width = 25
		sheet.column_dimensions['AH'].width = 25
		sheet.column_dimensions['AI'].width = 25
		sheet.column_dimensions['AJ'].width = 25
		sheet.column_dimensions['AK'].width = 25
		sheet.column_dimensions['AL'].width = 25
		sheet.column_dimensions['AM'].width = 25
		sheet.column_dimensions['AN'].width = 25
		sheet.column_dimensions['AO'].width = 25
		sheet.column_dimensions['AP'].width = 25
		sheet.column_dimensions['AQ'].width = 25
		sheet.column_dimensions['AR'].width = 25
		sheet.column_dimensions['AS'].width = 25
		sheet.column_dimensions['AT'].width = 25
		sheet.column_dimensions['AU'].width = 25
		sheet.column_dimensions['AV'].width = 25
		sheet.column_dimensions['AW'].width = 25
		sheet.column_dimensions['AX'].width = 25
		sheet.column_dimensions['AY'].width = 25
		sheet.column_dimensions['AZ'].width = 25

	for row in doc.get('order_details', []):
		if row.get('design_id') and row.get('bom'):

			final_bom = frappe.get_doc("BOM", row.get('bom'))
			if not final_bom:
				continue

			metal_touch = frappe.db.get_all(
				"Customer Metal Touch Detail",
				filters={
					"gk_metal_touch": row.get('metal_touch'),
					"parent": order_form_doc.customer_code
				},
				fields=['customer_metal_touch']
			)

			category = frappe.get_all(
				"Customer Category Detail",
				filters={
					"gk_category": row.get("category"),
					"parent": order_form_doc.customer_code
				},
				fields=['customer_category']
			)

			setting_type = frappe.get_all(
				"Customer Setting Detail",
				filters={
					"gk_setting_type": row.get("setting_type"),
					"gk_sub_setting_type": row.get("sub_setting_type1"),
					"parent": order_form_doc.customer_code
				},
				fields=['customer_setting_size']
			)

			metal_color = frappe.get_all(
				"Customer Metal Color Detail",
				filters={
					"gk_metal_color": row.get('metal_colour'),
					"parent": order_form_doc.customer_code
				},
				fields=['customer_metal_color']
			)
			metal_purity = frappe.get_all(
				"Customer Metal Touch Detail",
				filters={
					"gk_metal_touch": final_bom.get('metal_touch'),
					"parent": order_form_doc.customer_code
				},
				fields=['customer_metal_touch']
			)
			# ---------------------------
			# COMMON DATA (Same for all rows)
			# ---------------------------
			common_row_data = [
				"66426",
				row.get('design_id'),
				"",  # MATCHING DESIGN
				"DIAMOND ORNAMENTS",  # PRODUCTTYPE
				category[0].get('customer_category') if category else "",  # ITEM
				"",  # IMAGE
				"",  # FORM
				"",  # COLLECTION
				"",  # SUBCOLLECTION
				row.get("setting_type"),  # MANUFACTURING TYPE
				"",  # THEME
				row.get('gender', ''),
				"STUDDED",  # PLAIN/STUDDED
				setting_type[0].get('customer_setting_size') if setting_type else "",  # setting size
				"",  # SOLITAIRE
				row.get("design_id"),  # CODE
			]

			# ---------------------------
			# METAL DETAIL ROWS
			# ---------------------------
			for m in final_bom.get("metal_detail", []):
				row_data = common_row_data + [
					"METAL",
					"RAW GOLD",
					"",#Shape
					"",#Size
					"",#PCs
					final_bom.get('total_metal_weight' or ''),
					final_bom.get('total_metal_weight' or ''),
					final_bom.get('gross_weight' or ''),
					final_bom.get('metal_and_finding_weight' or ''),
					final_bom.get('total_diamond_pcs' or ''),
					final_bom.get('diamond_weight' or ''),
					final_bom.get('gemstone_weight' or ''),
					"NONE",
					"NONE",
					"NONE",
					metal_purity[0].get('customer_metal_touch') if metal_purity else '',
					"",
					"",
					"",
					"",
					"",
					metal_color[0].get('customer_metal_color') if metal_color else "",


					
				]
				rows_data.append(row_data)


			
			diamond_grouped = {}
			for g in final_bom.get("diamond_detail", []):

				diamond_code = frappe.db.get_value(
					"Customer Diamond Shape Detail",
					{
						"gk_diamond_shape": g.get("stone_shape"),
						"parent": order_form_doc.customer_code
					},
					"data"
				)

				diamond_size = frappe.db.get_value(
					"Customer Diamond Size Details",
					{
						"gk_diamond_shape": g.get("stone_shape"),
						"gk_diamond_size_in_mm": ["like", f"%{g.get('size_in_mm')}%"],
						"parent": order_form_doc.customer_code
					},
					"customer_diamond_group_size"
				)

				if not diamond_size:
					diamond_size = g.get("size_in_mm")  # fallback if not found

				# Group Key (same diamond_size means same row)
				key = (diamond_code or "", diamond_size)

				if key not in diamond_grouped:
					diamond_grouped[key] = {
						"pcs": 0,
						"weight_in_gms": 0,
						"quantity": 0,
						"diamond_code": diamond_code or "",
						"diamond_size": diamond_size
					}

				diamond_grouped[key]["pcs"] += flt(g.get("pcs"))
				diamond_grouped[key]["weight_in_gms"] += flt(g.get("weight_in_gms"))
				diamond_grouped[key]["quantity"] += flt(g.get("quantity"))


			# Now append only one row per group
			for key, values in diamond_grouped.items():

				row_data = common_row_data + [
					"STONE",
					"Diamond",
					values["diamond_code"],
					values["diamond_size"],
					values["pcs"],
					values["weight_in_gms"],
					values["quantity"]
				]

				rows_data.append(row_data)

			# ---------------------------
			# GEMSTONE DETAIL ROWS
			# ---------------------------
			gemstone_grouped = {}
			for g in final_bom.get("gemstone_detail", []):

				gemstone_type = frappe.db.get_value(
					"customer_gemstone_detail",
					{
						"gk_gemstone_type": g.get("gemstone_type"),
						"parent": order_form_doc.customer_code
					},
					"customer_gemstone_type"
				)

				gemstone_code = frappe.db.get_value(
					"Customer Diamond Shape Detail",
					{
						"gk_diamond_shape": g.get("stone_shape"),
						"parent": order_form_doc.customer_code
					},
					"data"
				)

				gemstone_size = frappe.db.get_value(
					"Customer Diamond Size Details",
					{
						"gk_diamond_shape": g.get("stone_shape"),
						"gk_diamond_size_in_mm": ["like", f"%{g.get('gemstone_size')}%"],
						"parent": order_form_doc.customer_code
					},
					"customer_diamond_group_size"
				)

				if not gemstone_size:
					gemstone_size = g.get("gemstone_size")

				key = (gemstone_type or "", gemstone_code or "", gemstone_size)

				if key not in gemstone_grouped:
					gemstone_grouped[key] = {
						"pcs": 0,
						"weight_in_gms": 0,
						"quantity": 0,
						"gemstone_type": gemstone_type or "",
						"gemstone_code": gemstone_code or "",
						"gemstone_size": gemstone_size or ""
					}

				gemstone_grouped[key]["pcs"] += flt(g.get("pcs"))
				gemstone_grouped[key]["weight_in_gms"] += flt(g.get("weight_in_gms"))
				gemstone_grouped[key]["quantity"] += flt(g.get("quantity"))


			# Append only one row per group
			for key, values in gemstone_grouped.items():

				row_data = common_row_data + [
					"STONE",
					values["gemstone_type"],
					values["gemstone_code"],
					values["gemstone_size"],
					values["pcs"],
					values["weight_in_gms"],
					values["quantity"]
				]

				rows_data.append(row_data)
				
    
    
    
    
    

			# ---------------------------
			# OTHER DETAIL ROWS
			# ---------------------------
			for o in final_bom.get("other_detail", []):
				row_data = common_row_data + [
					o.get("weight") or '',
					o.get("pcs") or '',
					o.get("shape") or "",
					"OTHER"
				]
				rows_data.append(row_data)
				import os
				from openpyxl.drawing.image import Image

				current_row = sheet.max_row
				# Set row height for image
				# sheet.row_dimensions[current_row].height = 150

				# Center align row
				for cell in sheet[current_row]:
					cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

				# ---------------- INSERT IMAGE IN COLUMN E ----------------
				image_path = final_bom.get("finish_front_view_preview")

				if image_path:
					# Convert /files/xyz.png -> public/files/xyz.png
					full_path = frappe.get_site_path(image_path.replace("/files/", "public/files/"))

					if os.path.exists(full_path):
						img = Image(full_path)
						img.width = 180
						img.height = 180

						# Insert image in Column E
						sheet.add_image(img, f"E{current_row}")
	
	# Write all rows to the Excel sheet at once
	if rows_data:
		for row in rows_data:
			sheet.append(row)
	else:
		frappe.throw("Proto Sheet Can Not Download")

	# Save the workbook to a BytesIO stream
	output = BytesIO()
	workbook.save(output)
	output.seek(0)

	file_doc = save_file(
		file_name,
		output.getvalue(),
		order_form_doc.doctype,
		order_form_doc.name,
		is_private=0
	)

	return file_doc.file_url





