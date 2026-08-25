"""
    /api/method/your_app.api.attendance_comparison.generate_comparison_excel?month=7&year=2026

or via frappe.call({method: "your_app.api.attendance_comparison.generate_comparison_excel", args: {month, year}})


The Excel file is built entirely in memory and streamed back as the
HTTP response (Content-Disposition: attachment) — nothing is written
to disk and nothing is saved as a Frappe File doc.
"""

import calendar
import io
import re

import frappe
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# =====================================================
# Config
# =====================================================

SOFTWARE_API_TEMPLATE = (
    "https://gkexport.frappe.cloud/api/method/get_attendance_report_data"
    "?month={month}&year={year}"
)

REQUIRED_COLS = [
    "employee", "employee_name", "sp_branch_name", "sp_branch_code",
    "old_employee_code", "old_punch_id", "status",
    "payment_days_slip", "sp_bank_days_api",
    "gross_pay_slip", "sp_gross_salary_api",
    "pt_slip", "sp_pt_api",
    "pf_slip", "sp_pf_api",
    "esi_slip", "sp_esi_api",
    "tds_slip", "sp_tds_api",
    "total_deduction_slip", "sp_total_deduct_api",
    "net_pay_slip", "sp_net_payable_api",
]

# Display field -> (ERP/slip key, SmartTime/API key) in the raw record dict
DIFF_MAPPING = {
    "Bank Days": ("payment_days_slip", "sp_bank_days_api"),
    "Gross Salary": ("gross_pay_slip", "sp_gross_salary_api"),
    "PT": ("pt_slip", "sp_pt_api"),
    "PF": ("pf_slip", "sp_pf_api"),
    "ESI": ("esi_slip", "sp_esi_api"),
    "TDS": ("tds_slip", "sp_tds_api"),
    "Total Deduct": ("total_deduction_slip", "sp_total_deduct_api"),
    "Net Pay": ("net_pay_slip", "sp_net_payable_api"),
}

IDENTIFIER_COLS = {
    "Employee", "Employee_name", "Branch Name", "Branch Code",
    "Employee Code", "Punch ID", "Status",
}

TRAILING_DOT_ZERO = re.compile(r"\.0$")

CACHE_EXPIRY_SECONDS = 15 * 60  # cached raw records are considered fresh for 15 min


def _cache_key(month, year):
    return f"attendance_comparison_raw::{month}::{year}"


# =====================================================
# Helpers
# =====================================================

def clean_str(value):
    """Trim whitespace and strip a trailing '.0' (from numeric-looking IDs)."""
    if value is None:
        return ""
    s = str(value).strip()
    return TRAILING_DOT_ZERO.sub("", s)


def to_number(value):
    """Best-effort numeric coercion; returns None (like NaN) if it can't convert."""
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _fetch_software_data(month, year):
    cache_key = _cache_key(month, year)
    cached = frappe.cache().get_value(cache_key)
    if cached:
        return cached["records"], cached["month"], cached["year"]

    url = SOFTWARE_API_TEMPLATE.format(month=month, year=year)

    response = requests.get(url, timeout=30)
    response.raise_for_status()

    try:
        json_data = response.json()
    except ValueError as e:
        raise ValueError(
            f"API did not return valid JSON. Status={response.status_code}. "
            f"First 300 chars of response: {response.text[:300]}"
        ) from e

    if isinstance(json_data, dict) and isinstance(json_data.get("message"), dict) and "data" in json_data["message"]:
        payload = json_data["message"]
    elif isinstance(json_data, dict) and "data" in json_data:
        payload = json_data
    else:
        raise ValueError(
            f"API JSON missing expected 'data' key. Got keys: "
            f"{list(json_data.keys()) if isinstance(json_data, dict) else type(json_data)}"
        )

    records = payload["data"]
    resp_month = payload.get("month", month)
    resp_year = payload.get("year", year)

    frappe.cache().set_value(
        cache_key,
        {"records": records, "month": resp_month, "year": resp_year},
        expires_in_sec=CACHE_EXPIRY_SECONDS,
    )

    return records, resp_month, resp_year


def _fetch_and_compare(month, year):
    """
    Core logic: fetch raw records, build the comparison rows.
    Returns (rows, columns, summary_dict, period_display)

    rows    -> list of dicts, one per employee (in `columns` order)
    columns -> ordered list of column names for the sheet
    """
    records, resp_month, resp_year = _fetch_software_data(month, year)

    try:
        period_display = f"{calendar.month_name[int(resp_month)]} {resp_year}"
    except (ValueError, IndexError, TypeError):
        period_display = f"{resp_month} {resp_year}"

    if not records:
        raise ValueError("API returned no records for the given month/year.")

    missing = [c for c in REQUIRED_COLS if c not in records[0]]
    if missing:
        raise ValueError(f"API data missing expected columns: {missing}. Available: {list(records[0].keys())}")

    columns = ["Employee", "Employee_name", "Branch Name", "Branch Code",
               "Employee Code", "Punch ID", "Status"]
    for field in DIFF_MAPPING:
        columns += [f"{field} (ERP)", f"{field} (SmartTime)", f"{field} Diff"]
    columns.append("Overall Status")

    rows = []
    matched_count = 0
    mismatch_count = 0
    missing_employee_count = 0

    for rec in records:
        if rec.get("employee_name") in (None, ""):
            missing_employee_count += 1

        row = {
            "Employee": rec.get("employee"),
            "Employee_name": clean_str(rec.get("employee_name")),
            "Branch Name": rec.get("sp_branch_name"),
            "Branch Code": rec.get("sp_branch_code"),
            "Employee Code": clean_str(rec.get("old_employee_code")),
            "Punch ID": clean_str(rec.get("old_punch_id")),
            "Status": clean_str(rec.get("status")),
        }

        all_zero = True
        for field, (erp_key, smart_key) in DIFF_MAPPING.items():
            erp_val = to_number(rec.get(erp_key))
            smart_val = to_number(rec.get(smart_key))
            diff = None if (erp_val is None or smart_val is None) else erp_val - smart_val

            row[f"{field} (ERP)"] = erp_val
            row[f"{field} (SmartTime)"] = smart_val
            row[f"{field} Diff"] = diff

            if diff is None or diff != 0:
                all_zero = False

        row["Overall Status"] = "Matched" if all_zero else "Mismatch"
        if all_zero:
            matched_count += 1
        else:
            mismatch_count += 1

        rows.append(row)

    summary = {
        "records": len(records),
        "matched": matched_count,
        "mismatch": mismatch_count,
        "missing_employee_name": missing_employee_count,
    }

    return rows, columns, summary, period_display


def _build_excel_bytes(rows, columns):
    """Builds the styled workbook fully in memory and returns raw bytes."""
    numeric_cols = {
        c for c in columns
        if c.endswith("(ERP)") or c.endswith("(SmartTime)") or c.endswith("Diff")
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"
    ws.append(columns)

    for row in rows:
        ws.append([row.get(col) for col in columns])

    # blank spacer row
    ws.append([""] * len(columns))

    # TOTAL row
    totals = {col: "" for col in columns}
    totals["Employee"] = "TOTAL"
    for col in numeric_cols:
        totals[col] = sum((r.get(col) or 0) for r in rows)
    ws.append([totals.get(col, "") for col in columns])

    data_row_count = len(rows)

    # Styling
    blue_fill = PatternFill(start_color="0078D7", end_color="0078D7", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    white_bold = Font(color="FFFFFF", bold=True)
    black_bold = Font(color="000000", bold=True)

    for cell in ws[1]:  # header row
        col_name = str(cell.value)
        if col_name in IDENTIFIER_COLS:
            cell.fill, cell.font = blue_fill, white_bold
        elif col_name.endswith("(ERP)"):
            cell.fill, cell.font = yellow_fill, black_bold
        elif col_name.endswith("(SmartTime)"):
            cell.fill, cell.font = green_fill, black_bold
        elif col_name.endswith("Diff"):
            cell.fill, cell.font = red_fill, white_bold

    for cell in ws[ws.max_row]:  # TOTAL row
        cell.font = Font(bold=True)

    last_col_letter = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A1:{last_col_letter}{data_row_count + 1}"

    out_buffer = io.BytesIO()
    wb.save(out_buffer)
    return out_buffer.getvalue()


# =====================================================
# Whitelisted API endpoints
# =====================================================

@frappe.whitelist()
def get_comparison_summary(month, year):
    """
    Lightweight endpoint — fetches + compares only, no Excel is built.
    Use this to show the counts first (Records/Matched/Mismatch/Missing)
    before the user decides to generate the file.
    """
    try:
        _, _, summary, period_display = _fetch_and_compare(month, year)
        return {
            "state": "Completed",
            "period": period_display,
            "records": summary["records"],
            "matched": summary["matched"],
            "mismatch": summary["mismatch"],
            "missing_employee_name": summary["missing_employee_name"],
        }
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Attendance Comparison Summary Failed")
        return {"state": "Failed", "error": str(e)}


@frappe.whitelist()
def generate_comparison_excel(month, year):
    """
    Pass month & year -> the HTTP response itself is the .xlsx file
    (Content-Disposition: attachment), so the browser downloads it
    directly.

    IMPORTANT: hit this as a direct URL/navigation, not frappe.call():
        window.open(
          "/api/method/your_app.api.attendance_comparison.generate_comparison_excel"
          + "?month=" + month + "&year=" + year
        )
    or a plain <a href="...">Download</a> link. An AJAX call (frappe.call)
    receives the bytes back as data but the browser won't show a save
    dialog for it — only a direct navigation triggers the download.

    Nothing is saved to disk or to the Frappe File doctype — the file
    only exists as bytes in memory for this request.

    On failure: returns a normal JSON error dict instead (no file).
    """
    try:
        rows, columns, summary, period_display = _fetch_and_compare(month, year)
        file_bytes = _build_excel_bytes(rows, columns)

        file_name = f"Software_Comparison_Report_{month}_{year}.xlsx"

        frappe.local.response.filename = file_name
        frappe.local.response.filecontent = file_bytes
        frappe.local.response.type = "binary"
        # counts ride along as response headers since the body is now the file
        frappe.local.response.headers = {
            "X-Report-Period": period_display,
            "X-Records": summary["records"],
            "X-Matched": summary["matched"],
            "X-Mismatch": summary["mismatch"],
            "X-Missing-Employee-Name": summary["missing_employee_name"],
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Attendance Comparison Excel Failed")
        frappe.local.response["http_status_code"] = 417
        frappe.response["message"] = {"state": "Failed", "error": str(e)}